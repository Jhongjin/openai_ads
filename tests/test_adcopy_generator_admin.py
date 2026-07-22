from __future__ import annotations

from io import BytesIO
import os
import unittest
from unittest.mock import AsyncMock, patch

import httpx
from fastapi.testclient import TestClient

from app import AdsAdcopyGenerateRequest, _adcopy_import_trace, _adcopy_is_contextual_hint, _adcopy_user_prompt, _adcopy_warning_penalty, app


ADMIN_HEADERS = {"x-admin-password": "nas2026@"}


def adcopy_payload() -> dict:
    return {
        "advertiser_name": "캐츠잉글리시",
        "industry": "교육",
        "campaign_name": "01_학습자료",
        "objective": "Views",
        "budget_max": 102600,
        "budget_type": "daily",
        "launch_date": "2026-07-01",
        "end_date": "2026-07-31",
        "target_countries": ["KR"],
        "product_name": "초등 영어 학습 앱",
        "landing_url": "https://example.com/landing",
        "image_link": "https://example.com/image.png",
        "audience": "초등학생 영어 학습을 돕는 학부모",
        "selling_points": "짧은 반복 훈련, 수준별 학습, 학습 리포트 제공",
        "tone": "담백하고 신뢰감 있게",
        "banned_terms": "무조건\n업계 1위",
        "required_phrases": "",
        "adgroup_count": 1,
        "ads_per_adgroup": 2,
    }


def generated_payload() -> dict:
    trace = {
        "source_type": "AI 생성",
        "source_url": "",
        "source_excerpt": "",
        "generation_basis": "초등 영어 반복 학습",
        "confidence_score": 0.82,
        "validation_status": "담당자 확인 필요",
        "review_comment": "should be removed",
        "exclusion_reason": "should be removed",
    }
    return {
        "policy": {"banned_terms": []},
        "campaigns": [],
        "adgroups": [
            {
                "campaign_name": "AI가 바꾼 이름",
                "adgroup_name": "01_반복훈련",
                "max_bid": 7000,
                "keywords": [
                    {"text": "초등 영어 앱", "origin": "customer_data"},
                    {"text": "영어 반복 학습", "origin": "ai_inferred"},
                    {"text": "초등 파닉스", "origin": "ai_inferred"},
                    {"text": "초등 영어 리포트", "origin": "ai_inferred"},
                    {"text": "집에서 영어 공부", "origin": "ai_inferred"},
                ],
                "required_phrases": [],
                "trace": trace,
            }
        ],
        "ads": [
            {
                "ad_name": "KID_01_001",
                "adgroup_name": "01_반복훈련",
                "title": "초등 영어 반복 훈련",
                "copy": "짧은 학습 루틴으로 매일 영어 자신감을 키우세요",
                "link": "https://wrong.example",
                "image_link": "https://wrong.example/image.png",
                "trace": trace,
            },
            {
                "ad_name": "KID_01_002",
                "adgroup_name": "01_반복훈련",
                "title": "수준별 영어 학습",
                "copy": "아이 수준에 맞춘 영어 훈련과 리포트를 확인하세요",
                "link": "https://wrong.example",
                "image_link": "https://wrong.example/image.png",
                "trace": trace,
            },
        ],
    }


class AdcopyGeneratorAdminTests(unittest.TestCase):
    def test_ai_validation_and_blank_human_review_remain_separate(self) -> None:
        trace = _adcopy_import_trace(
            {
                "validation_status": "광고주 확인 필요: 이미지 소재 미제공",
                "검수상태": "",
            }
        )

        self.assertEqual(trace["validation_status"], "광고주 확인 필요: 이미지 소재 미제공")
        self.assertEqual(trace["review_status"], "담당자 확인 필요")

    def test_admin_adcopy_engines_requires_admin_password(self) -> None:
        client = TestClient(app, raise_server_exceptions=False)

        response = client.get("/api/admin/adcopy/engines")

        self.assertEqual(response.status_code, 403)

    def test_admin_adcopy_engines_defaults_to_external_ai_team_plugin(self) -> None:
        client = TestClient(app, raise_server_exceptions=False)

        with patch.dict(os.environ, {}, clear=True):
            response = client.get("/api/admin/adcopy/engines", headers=ADMIN_HEADERS)

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertEqual(body["default_engine"], "ai_team_plugin")
        self.assertEqual(body["effective_default_engine"], "ai_team_plugin")
        self.assertFalse(body["default_fallback_used"])
        plugin = next(item for item in body["engines"] if item["id"] == "ai_team_plugin")
        self.assertTrue(plugin["available"])
        self.assertEqual(plugin["execution_mode"], "external_plugin")

    def test_admin_adcopy_engines_reports_admate_server_availability(self) -> None:
        client = TestClient(app, raise_server_exceptions=False)

        with patch.dict(os.environ, {"OPENAI_API_KEY": "test-admate-key"}, clear=True):
            response = client.get("/api/admin/adcopy/engines", headers=ADMIN_HEADERS)

        self.assertEqual(response.status_code, 200)
        body = response.json()
        admate = next(item for item in body["engines"] if item["id"] == "admate")
        self.assertTrue(admate["available"])
        self.assertEqual(admate["execution_mode"], "server")

    def test_admin_adcopy_generation_routes_ai_team_plugin_to_local_workflow(self) -> None:
        client = TestClient(app, raise_server_exceptions=False)
        payload = {**adcopy_payload(), "engine": "ai_team_plugin"}

        with patch("app._call_openai_adcopy", AsyncMock()) as admate_call:
            response = client.post("/api/admin/adcopy/generate", json=payload, headers=ADMIN_HEADERS)

        self.assertEqual(response.status_code, 409)
        self.assertIn("Claude Code 플러그인", response.json()["detail"])
        admate_call.assert_not_awaited()

    def test_admin_adcopy_generation_rejects_unknown_engine(self) -> None:
        client = TestClient(app, raise_server_exceptions=False)
        payload = {**adcopy_payload(), "engine": "unknown"}

        with patch("app._call_openai_adcopy", AsyncMock()) as admate_call:
            response = client.post("/api/admin/adcopy/generate", json=payload, headers=ADMIN_HEADERS)

        self.assertEqual(response.status_code, 400)
        admate_call.assert_not_awaited()

    def test_admin_adcopy_generation_requires_admin_password(self) -> None:
        client = TestClient(app, raise_server_exceptions=False)

        response = client.post("/api/admin/adcopy/generate", json=adcopy_payload())
        empty_response = client.post("/api/admin/adcopy/generate", json={})

        self.assertEqual(response.status_code, 403)
        self.assertEqual(empty_response.status_code, 403)

    def test_admin_adcopy_generation_normalizes_and_validates_generated_json(self) -> None:
        client = TestClient(app, raise_server_exceptions=False)
        mocked_call = AsyncMock(
            return_value={
                "model": "gpt-test",
                "raw_response_id": "resp_test",
                "generated": generated_payload(),
            }
        )

        with patch("app._call_openai_adcopy", mocked_call):
            response = client.post("/api/admin/adcopy/generate", json=adcopy_payload(), headers=ADMIN_HEADERS)

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body["ok"])
        self.assertEqual(body["model"], "gpt-test")
        self.assertEqual(body["generated"]["campaigns"][0]["campaign_name"], "01_학습자료")
        self.assertEqual(body["generated"]["campaigns"][0]["budget_max"], 102600)
        self.assertNotIn("max_bid", body["generated"]["adgroups"][0])
        self.assertEqual(body["generated"]["ads"][0]["link"], "https://example.com/landing")
        self.assertEqual(body["generated"]["ads"][0]["image_link"], "https://example.com/image.png")
        self.assertEqual(body["generated"]["ads"][0]["trace"]["review_comment"], "")
        self.assertEqual(body["summary"]["adgroups"], 1)
        self.assertEqual(body["summary"]["ads"], 2)
        self.assertIn("quality", body["validation_report"])
        self.assertIn("creative_checks", body["validation_report"])
        self.assertGreaterEqual(body["validation_report"]["quality"]["score"], 1)
        self.assertEqual(body["validation_report"]["creative_checks"][0]["ad_name"], "KID_01_001")
        self.assertIn("readiness", body["summary"])

    def test_admin_adcopy_generation_reports_policy_errors(self) -> None:
        client = TestClient(app, raise_server_exceptions=False)
        generated = generated_payload()
        generated["ads"][0]["title"] = "업계 1위 영어 학습"
        mocked_call = AsyncMock(
            return_value={
                "model": "gpt-test",
                "raw_response_id": "resp_test",
                "generated": generated,
            }
        )

        with patch("app._call_openai_adcopy", mocked_call):
            response = client.post("/api/admin/adcopy/generate", json=adcopy_payload(), headers=ADMIN_HEADERS)

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertFalse(body["ok"])
        self.assertIn("banned_term", {item["rule"] for item in body["validation_report"]["errors"]})

    def test_admin_adcopy_generation_reports_quality_warnings(self) -> None:
        client = TestClient(app, raise_server_exceptions=False)
        generated = generated_payload()
        generated["ads"][0]["copy"] = "학습을 고려한 영어 앱"
        mocked_call = AsyncMock(
            return_value={
                "model": "gpt-test",
                "raw_response_id": "resp_test",
                "generated": generated,
            }
        )

        with patch("app._call_openai_adcopy", mocked_call):
            response = client.post("/api/admin/adcopy/generate", json=adcopy_payload(), headers=ADMIN_HEADERS)

        self.assertEqual(response.status_code, 200)
        body = response.json()
        warning_rules = {item["rule"] for item in body["validation_report"]["warnings"]}
        self.assertIn("awkward_phrase", warning_rules)
        self.assertLess(body["validation_report"]["quality"]["score"], 100)
        self.assertEqual(body["validation_report"]["creative_checks"][0]["status"], "warning")

    def test_adcopy_prompt_requires_title_copy_continuity_and_role_diversity(self) -> None:
        prompt = _adcopy_user_prompt(AdsAdcopyGenerateRequest(**adcopy_payload()))

        self.assertIn("user's tension", prompt)
        self.assertIn("continue the title instead of paraphrasing it", prompt)
        self.assertIn("solution direction -> benefit or judgment point", prompt)
        self.assertIn("Assign distinct roles before writing", prompt)

    def test_adcopy_context_hint_accepts_long_situation_sentences(self) -> None:
        self.assertTrue(_adcopy_is_contextual_hint("아이 영어 숙제를 봐주다가 문법 설명에서 막혔어요"))
        self.assertTrue(_adcopy_is_contextual_hint("사 준 영어 원서가 너무 어려워서 아이가 덮어버려요"))
        self.assertFalse(_adcopy_is_contextual_hint("초등 영어 원서 추천"))

    def test_adcopy_warning_penalty_caps_repeated_rule_instances(self) -> None:
        warnings = [
            {"rule": "policy_risk_phrase"}
            for _ in range(20)
        ]

        self.assertEqual(_adcopy_warning_penalty(warnings), 8)

    def test_admin_adcopy_validate_warns_when_copy_repeats_title(self) -> None:
        client = TestClient(app, raise_server_exceptions=False)
        generated = generated_payload()
        generated["ads"][0]["title"] = "아이 영어 수준이 궁금해요"
        generated["ads"][0]["copy"] = "아이 영어 수준이 궁금해요 지금 확인해보세요"

        response = client.post("/api/admin/adcopy/validate", json={"generated": generated}, headers=ADMIN_HEADERS)

        self.assertEqual(response.status_code, 200)
        warning_rules = {item["rule"] for item in response.json()["validation_report"]["warnings"]}
        self.assertIn("copy_repeats_title", warning_rules)

    def test_admin_adcopy_validate_warns_on_repeated_title_and_copy_forms(self) -> None:
        client = TestClient(app, raise_server_exceptions=False)
        generated = generated_payload()
        generated["ads"] = [
            {
                **generated["ads"][0],
                "ad_name": f"KID_REPEAT_{index}",
                "title": title,
                "copy": copy,
            }
            for index, (title, copy) in enumerate(
                [
                    ("학습 순서가 자꾸 막힌다면", "로드맵을 따라 매일 순서를 확인할 수 있어요"),
                    ("영어 진도가 늘 고민이라면", "로드맵을 따라 다음 단계를 확인할 수 있어요"),
                    ("다음 공부가 늘 막막하다면", "로드맵을 따라 학습 방향을 확인할 수 있어요"),
                ],
                start=1,
            )
        ]

        response = client.post("/api/admin/adcopy/validate", json={"generated": generated}, headers=ADMIN_HEADERS)

        self.assertEqual(response.status_code, 200)
        warning_rules = {item["rule"] for item in response.json()["validation_report"]["warnings"]}
        self.assertIn("title_form_repetition", warning_rules)
        self.assertIn("copy_opening_repetition", warning_rules)
        self.assertIn("copy_role_diversity", warning_rules)

    def test_admin_adcopy_validate_requires_admin_password(self) -> None:
        client = TestClient(app, raise_server_exceptions=False)

        response = client.post("/api/admin/adcopy/validate", json={"generated": generated_payload()})

        self.assertEqual(response.status_code, 403)

    def test_admin_adcopy_validate_reviews_edited_generated_json(self) -> None:
        client = TestClient(app, raise_server_exceptions=False)
        generated = generated_payload()
        generated["ads"][0]["title"] = "짧음"

        response = client.post("/api/admin/adcopy/validate", json={"generated": generated}, headers=ADMIN_HEADERS)

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertIn("validation_report", body)
        warning_rules = {item["rule"] for item in body["validation_report"]["warnings"]}
        self.assertIn("title_len_recommended", warning_rules)
        self.assertEqual(body["validation_report"]["creative_checks"][0]["status"], "warning")

    def test_admin_adcopy_import_requires_admin_password(self) -> None:
        client = TestClient(app, raise_server_exceptions=False)

        response = client.post("/api/admin/adcopy/import", json={"generated": generated_payload()})

        self.assertEqual(response.status_code, 403)

    def test_admin_adcopy_sample_workbook_requires_admin_and_downloads_xlsx(self) -> None:
        from openpyxl import load_workbook

        client = TestClient(app, raise_server_exceptions=False)
        blocked = client.get("/api/admin/adcopy/sample-workbook")
        self.assertEqual(blocked.status_code, 403)

        response = client.get("/api/admin/adcopy/sample-workbook", headers=ADMIN_HEADERS)
        self.assertEqual(response.status_code, 200)
        self.assertIn("openai_ads_adcopy_review_sample.xlsx", response.headers.get("content-disposition", ""))
        workbook = load_workbook(BytesIO(response.content), data_only=True, read_only=True)
        try:
            self.assertEqual(workbook.sheetnames, ["campaigns_검수", "adgroups_검수", "ads_검수"])
            self.assertEqual(workbook["campaigns_검수"]["A2"].value, "campaign_name")
            self.assertEqual(workbook["adgroups_검수"]["A2"].value, "adgroup_name")
            self.assertEqual(workbook["adgroups_검수"]["A3"].value, "영어학습*초등학부모*학습루틴고민_문제정의")
            self.assertIn("어떻게 시작하면 좋을까", workbook["adgroups_검수"]["B3"].value)
            self.assertEqual(workbook["ads_검수"]["B3"].value, "영어학습*초등학부모*학습루틴고민_문제정의")
            self.assertEqual(workbook["ads_검수"]["A2"].value, "ad_name")
        finally:
            workbook.close()

    def test_admin_adcopy_import_workbook_explains_protected_office_file(self) -> None:
        client = TestClient(app, raise_server_exceptions=False)

        response = client.post(
            "/api/admin/adcopy/import-workbook",
            files={"file": ("protected.xlsx", b"SCDSA004-protected", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=ADMIN_HEADERS,
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("민감도 레이블", response.json()["detail"])
        self.assertIn("표준 .xlsx 복사본", response.json()["detail"])

    def test_admin_adcopy_import_normalizes_native_generated_json(self) -> None:
        client = TestClient(app, raise_server_exceptions=False)
        generated = generated_payload()
        generated["campaigns"] = [
            {
                "campaign_name": "01_학습자료",
                "budget_max": 102600,
                "budget_type": "daily",
                "launch_date": "2026-07-01",
                "end_date": "2026-07-31",
                "objective": "Views",
                "target_countries": ["KR"],
            }
        ]

        response = client.post(
            "/api/admin/adcopy/import",
            json={
                "generated": generated,
                "advertiser_name": "캐츠잉글리시",
                "landing_url": "https://example.com/landing",
                "image_link": "https://example.com/image.png",
            },
            headers=ADMIN_HEADERS,
        )

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertEqual(body["model"], "기존 작업 불러오기")
        self.assertEqual(body["import_report"]["source_format"], "native")
        self.assertEqual(body["generated"]["campaigns"][0]["campaign_name"], "01_학습자료")
        self.assertNotIn("max_bid", body["generated"]["adgroups"][0])
        self.assertEqual(body["summary"]["ads"], 2)

    def test_admin_adcopy_import_accepts_ai_team_style_json(self) -> None:
        client = TestClient(app, raise_server_exceptions=False)
        external = {
            "campaign": {
                "캠페인명": "초등영어_테스트",
                "예산": "25,000원",
                "예산유형": "일 예산",
                "시작일": "2026-07-10",
                "종료일": "2026-07-31",
                "목표": "노출",
                "국가": "KR",
            },
            "ad_groups": [
                {
                    "광고그룹명": "01_학습습관",
                    "확장 키워드": ["초등 영어 앱", "매일 영어 학습", "파닉스 연습", "영어 루틴", "학부모 영어"],
                    "필수 포함 문구": ["영어 루틴"],
                }
            ],
            "creatives": [
                {
                    "소재명": "AD_001",
                    "광고그룹명": "01_학습습관",
                    "제목": "초등 영어 습관 만들기",
                    "카피": "영어 루틴으로 매일 짧게 학습을 이어가요",
                    "랜딩 URL": "https://example.com/landing",
                    "이미지 URL": "https://example.com/image.png",
                    "휴먼 체크": "확인 완료 검수",
                }
            ],
            "policy": {"banned_terms": ["무조건"]},
        }

        response = client.post(
            "/api/admin/adcopy/import",
            json={"generated": external, "advertiser_name": "캐츠잉글리시"},
            headers=ADMIN_HEADERS,
        )

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body["ok"])
        self.assertEqual(body["generated"]["campaigns"][0]["budget_type"], "daily")
        self.assertEqual(body["generated"]["campaigns"][0]["objective"], "Views")
        self.assertEqual(body["generated"]["adgroups"][0]["keywords"][0]["text"], "초등 영어 앱")
        self.assertEqual(body["generated"]["ads"][0]["trace"]["validation_status"], "승인")
        self.assertIn("무조건", body["generated"]["policy"]["banned_terms"])

    def test_admin_adcopy_import_accepts_ai_team_workbook_dump_json(self) -> None:
        client = TestClient(app, raise_server_exceptions=False)
        external = {
            "file": "review.xlsx",
            "sheets": [
                {
                    "name": "campaigns",
                    "rows": [
                        {
                            "campaign_name": "워크북_캠페인",
                            "budget_max": "50,000",
                            "budget_type": "daily",
                            "launch_date": "2026-07-10",
                            "end_date": "2026-07-12",
                            "objective": "Views",
                            "target_countries": '["KR"]',
                        }
                    ],
                },
                {
                    "name": "adgroups_검수",
                    "rows": [
                        {
                            "campaign_name": "워크북_캠페인",
                            "adgroup_name": "01_검수그룹",
                            "keywords": '["초등 영어", "학습 습관", "파닉스", "영어 루틴", "학부모"]',
                            "keywords_origin": '["customer_data", "ai_inferred", "ai_inferred", "customer_data", "ai_inferred"]',
                            "validation_status": "통과",
                            "검수상태": "수정 후 승인",
                            "review_comment": "문구 확인 완료",
                        }
                    ],
                },
                {
                    "name": "ads_검수",
                    "rows": [
                        {
                            "ad_name": "AD_001",
                            "adgroup_name": "01_검수그룹",
                            "title": "초등 영어 루틴 시작",
                            "copy": "매일 짧게 반복하는 영어 학습 습관",
                            "link": "https://example.com/landing",
                            "image_link": "https://example.com/image.png",
                            "validation_status": "통과",
                            "검수상태": "무수정 승인",
                            "review_comment": "사용 가능",
                        },
                        {
                            "ad_name": "AD_002",
                            "adgroup_name": "01_검수그룹",
                            "title": "초등 영어 무료 체험",
                            "copy": "무료 혜택은 광고주 확인 후 사용",
                            "link": "https://alt.example.com/landing",
                            "image_link": "https://example.com/image.png",
                            "validation_status": "의미 중복 후보 / 경고",
                            "검수상태": "무수정 승인",
                        },
                        {
                            "ad_name": "AD_003",
                            "adgroup_name": "01_검수그룹",
                            "title": "사용 불가 소재",
                            "copy": "제외 상태 확인용 문구입니다",
                            "link": "https://example.com/landing",
                            "image_link": "https://example.com/image.png",
                            "검수상태": "사용 불가",
                        },
                    ],
                },
            ],
        }

        response = client.post(
            "/api/admin/adcopy/import",
            json={"generated": external, "advertiser_name": "캐츠잉글리시"},
            headers=ADMIN_HEADERS,
        )

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertEqual(body["import_report"]["adgroups"], 1)
        self.assertEqual(body["import_report"]["ads"], 3)
        self.assertEqual(body["generated"]["adgroups"][0]["trace"]["validation_status"], "통과")
        self.assertEqual(body["generated"]["adgroups"][0]["trace"]["review_status"], "승인")
        self.assertEqual(body["generated"]["adgroups"][0]["keywords"][0]["origin"], "customer_data")
        self.assertEqual(body["generated"]["ads"][0]["trace"]["validation_status"], "통과")
        self.assertEqual(body["generated"]["ads"][0]["trace"]["review_status"], "승인")
        self.assertEqual(body["generated"]["ads"][1]["trace"]["validation_status"], "의미 중복 후보 / 경고")
        self.assertEqual(body["generated"]["ads"][1]["trace"]["review_status"], "승인")
        self.assertEqual(body["generated"]["ads"][1]["trace"]["review_comment"], "")
        self.assertEqual(body["generated"]["ads"][2]["trace"]["validation_status"], "제외")
        self.assertEqual(body["generated"]["ads"][2]["trace"]["review_status"], "제외")
        self.assertGreaterEqual(body["validation_report"]["summary"]["policy_risk_count"], 1)
        self.assertGreaterEqual(body["validation_report"]["summary"]["landing_domain_count"], 2)

    def test_admin_adcopy_import_canonicalizes_legacy_underscore_adgroup_names(self) -> None:
        client = TestClient(app, raise_server_exceptions=False)
        generated = generated_payload()
        legacy_name = "무료학습_초등학부모_체험조건확인_신청전환"
        generated["adgroups"][0]["adgroup_name"] = legacy_name
        for ad in generated["ads"]:
            ad["adgroup_name"] = legacy_name

        response = client.post(
            "/api/admin/adcopy/import",
            json={"generated": generated},
            headers=ADMIN_HEADERS,
        )

        self.assertEqual(response.status_code, 200)
        body = response.json()
        canonical_name = "무료학습*초등학부모*체험조건확인_신청전환"
        self.assertEqual(body["generated"]["adgroups"][0]["adgroup_name"], canonical_name)
        self.assertTrue(all(ad["adgroup_name"] == canonical_name for ad in body["generated"]["ads"]))

    def test_admin_adcopy_import_derives_campaigns_from_adgroup_rows(self) -> None:
        client = TestClient(app, raise_server_exceptions=False)
        external = {
            "adgroups_검수": [
                {"campaign_name": "01_학습자료", "adgroup_name": "훈련앱_학부모_반복연습_제품발견", "keywords": ["반복 연습이 필요해요"]},
                {"campaign_name": "02_학습가이드", "adgroup_name": "로드맵_학부모_학습순서_제품발견", "keywords": ["다음 단계가 궁금해요"]},
            ],
            "ads_검수": [
                {
                    "ad_name": "AD_001",
                    "adgroup_name": "훈련앱_학부모_반복연습_제품발견",
                    "title": "반복 연습이 필요한 순간",
                    "copy": "교재와 연동된 앱으로 배운 내용을 이어서 연습해요",
                    "link": "https://example.com",
                    "image_link": "https://example.com/image.png",
                },
                {
                    "ad_name": "AD_002",
                    "adgroup_name": "로드맵_학부모_학습순서_제품발견",
                    "title": "다음 학습 순서가 고민이라면",
                    "copy": "영역별 로드맵을 보며 다음 학습 방향을 정할 수 있어요",
                    "link": "https://example.com",
                    "image_link": "https://example.com/image.png",
                },
            ],
        }

        response = client.post(
            "/api/admin/adcopy/import",
            json={"generated": external},
            headers=ADMIN_HEADERS,
        )

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertEqual([item["campaign_name"] for item in body["generated"]["campaigns"]], ["01_학습자료", "02_학습가이드"])
        self.assertEqual(
            [item["campaign_name"] for item in body["generated"]["adgroups"]],
            ["01_학습자료", "02_학습가이드"],
        )
        self.assertIn("캠페인 2개를 복원", body["import_report"]["warnings"][0])

    def test_admin_adcopy_import_workbook_normalizes_review_xlsx(self) -> None:
        from openpyxl import Workbook

        client = TestClient(app, raise_server_exceptions=False)
        workbook = Workbook()
        ws = workbook.active
        ws.title = "campaigns"
        ws.append(["campaign_name", "advertiser_name", "budget_max", "budget_type", "launch_date", "end_date", "objective", "target_countries"])
        ws.append(["엑셀흡수_캠페인", "캐츠잉글리시", "102600", "daily", "2026-07-01", "2026-07-31", "Views", "KR"])

        groups = workbook.create_sheet("adgroups")
        groups.append(["adgroup_name", "keywords", "required_phrases", "검수상태"])
        groups.append(["01_검수그룹", "초등 영어 앱, 반복 학습, 파닉스, 무료 체험, 학부모", "초등 영어", "확인 완료 검수"])

        ads = workbook.create_sheet("ads")
        ads.append(["ad_name", "adgroup_name", "title", "copy", "link", "image_link", "검수상태"])
        ads.append([
            "AD_001",
            "01_검수그룹",
            "초등 영어 반복 학습",
            "학습 흐름을 확인하고 시작해 보세요",
            "https://example.com/landing",
            "https://example.com/image.png",
            "확인 완료 검수",
        ])
        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()
        buffer.seek(0)

        response = client.post(
            "/api/admin/adcopy/import-workbook",
            files={"file": ("review.xlsx", buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=ADMIN_HEADERS,
        )

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertEqual(body["model"], "엑셀 불러오기")
        self.assertEqual(body["workbook"]["sheets"], 3)
        self.assertEqual(body["import_report"]["adgroups"], 1)
        self.assertEqual(body["import_report"]["ads"], 1)
        self.assertEqual(body["generated"]["campaigns"][0]["campaign_name"], "엑셀흡수_캠페인")
        self.assertEqual(body["generated"]["adgroups"][0]["trace"]["validation_status"], "승인")
        self.assertEqual(body["generated"]["adgroups"][0]["trace"]["review_status"], "승인")
        self.assertEqual(body["generated"]["ads"][0]["trace"]["validation_status"], "승인")
        self.assertEqual(body["generated"]["ads"][0]["trace"]["review_status"], "승인")

    def test_admin_adcopy_workbook_to_paused_draft_e2e(self) -> None:
        from openpyxl import Workbook

        client = TestClient(app, raise_server_exceptions=False)
        workbook = Workbook()
        campaign_sheet = workbook.active
        campaign_sheet.title = "campaigns_검수"
        campaign_sheet.append(["광고팀 전달 워크북"])
        campaign_sheet.append(["campaign_name", "advertiser_name", "budget_max", "budget_type", "launch_date", "end_date", "objective", "target_countries"])
        campaign_sheet.append(["워크북E2E_캠페인", "캐츠잉글리시", "150000", "total", "2026-07-20", "2026-08-02", "Views", "KR"])

        groups_sheet = workbook.create_sheet("adgroups_검수")
        groups_sheet.append(["adgroup_name", "keywords", "required_phrases", "검수상태"])
        groups_sheet.append(["01_반복훈련", "초등 영어 앱, 영어 반복 학습, 파닉스, 학습 리포트, 집 공부", "", "확인 완료 검수"])
        groups_sheet.append(["02_학부모설득", "초등 영어 학부모, 수준별 영어, 영어 습관, 학습 진단, 영어 자신감", "", "확인 완료 검수"])

        ads_sheet = workbook.create_sheet("ads_검수")
        ads_sheet.append(["ad_name", "adgroup_name", "title", "copy", "link", "image_link", "검수상태"])
        ads_sheet.append([
            "AD_001",
            "01_반복훈련",
            "초등 영어 반복 루틴",
            "매일 짧게 이어지는 영어 학습 흐름을 확인하세요",
            "https://example.com/cats/landing",
            "https://example.com/cats/image.png",
            "확인 완료 검수",
        ])
        ads_sheet.append([
            "AD_002",
            "01_반복훈련",
            "반복 학습 보류 소재",
            "운영자가 제외한 소재는 draft 생성에서 제외됩니다",
            "https://example.com/cats/landing",
            "https://example.com/cats/excluded.png",
            "사용 불가",
        ])
        ads_sheet.append([
            "AD_003",
            "02_학부모설득",
            "학부모가 보는 영어",
            "수준별 학습 리포트로 아이의 변화를 살펴보세요",
            "https://example.com/cats/landing",
            "https://example.com/cats/image.png",
            "확인 완료 검수",
        ])
        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()
        buffer.seek(0)

        import_response = client.post(
            "/api/admin/adcopy/import-workbook",
            files={"file": ("ad-team-review.xlsx", buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=ADMIN_HEADERS,
        )
        self.assertEqual(import_response.status_code, 200)
        imported = import_response.json()
        self.assertEqual(imported["import_report"]["adgroups"], 2)
        self.assertEqual(imported["import_report"]["ads"], 3)
        generated = imported["generated"]

        save_response = client.post(
            "/api/admin/adcopy/review-state",
            json={
                "advertiser_name": "캐츠잉글리시",
                "campaign_name": "워크북E2E_캠페인",
                "source_label": "워크북 E2E",
                "generated": generated,
            },
            headers=ADMIN_HEADERS,
        )
        self.assertEqual(save_response.status_code, 200)
        snapshot_id = save_response.json()["snapshot"]["id"]

        snapshot_response = client.get(f"/api/admin/adcopy/review-state/{snapshot_id}", headers=ADMIN_HEADERS)
        self.assertEqual(snapshot_response.status_code, 200)
        snapshot_generated = snapshot_response.json()["item"]["generated"]
        account = {"ok": True, "ad_account": {"name": "Catsenglish", "currency_code": "KRW", "timezone": "Asia/Seoul"}}

        with patch("admin_store.get_ads_api_key_credential", return_value={"advertiser_name": "캐츠잉글리시", "industry": "교육", "api_key": "sk-test"}), patch(
            "rag_chatbot.ads_api.fetch_ad_account_metadata",
            AsyncMock(return_value=account),
        ), patch("app._ads_draft_api_request", AsyncMock()) as no_mutation_request:
            preflight_response = client.post(
                "/api/admin/adcopy/draft-preflight",
                json={
                    "advertiser_name": "캐츠잉글리시",
                    "generated": snapshot_generated,
                    "default_max_bid_krw": 7000,
                    "location_ids": ["2000043"],
                },
                headers=ADMIN_HEADERS,
            )

        self.assertEqual(preflight_response.status_code, 200)
        preflight = preflight_response.json()
        self.assertTrue(preflight["ok"])
        self.assertEqual(preflight["plan"]["summary"]["ads"], 2)
        self.assertEqual(preflight["plan"]["summary"]["excluded_ads"], 1)
        self.assertFalse(no_mutation_request.await_args_list)

        calls: list[tuple[str, dict]] = []

        async def mocked_request(api_key: str, method: str, path: str, json_body: dict | None = None) -> dict:
            calls.append((path, json_body or {}))
            if path == "/v1/campaigns":
                return {"id": "cmpn_e2e", "status": "paused"}
            if path == "/v1/upload":
                return {"file_id": "file_shared"}
            if path == "/v1/ad_groups":
                return {"id": f"adgrp_{len([item for item in calls if item[0] == '/v1/ad_groups'])}", "status": "paused"}
            if path == "/v1/ads":
                return {"id": f"ad_{len([item for item in calls if item[0] == '/v1/ads'])}", "status": "paused"}
            raise AssertionError(path)

        state: dict = dict(preflight.get("state_patch") or {})
        base_payload = {
            "advertiser_name": "캐츠잉글리시",
            "generated": snapshot_generated,
            "default_max_bid_krw": 7000,
            "location_ids": ["2000043"],
        }
        with patch("admin_store.get_ads_api_key_credential", return_value={"api_key": "sk-test", "industry": "교육"}), patch("app._ads_draft_api_request", mocked_request):
            for action in ["create_campaign", "upload_assets", "create_ad_groups", "create_ads"]:
                response = client.post(
                    "/api/admin/adcopy/draft-execute",
                    json={**base_payload, "action": action, "state": state, "confirm": True},
                    headers=ADMIN_HEADERS,
                )
                self.assertEqual(response.status_code, 200)
                state = response.json()["state"]

        paths = [path for path, _payload in calls]
        self.assertEqual(paths, ["/v1/campaigns", "/v1/upload", "/v1/ad_groups", "/v1/ad_groups", "/v1/ads", "/v1/ads"])
        self.assertTrue(all("/activate" not in path for path in paths))
        self.assertEqual(calls[0][1]["status"], "paused")
        self.assertEqual(calls[2][1]["status"], "paused")
        self.assertEqual(calls[4][1]["status"], "paused")
        self.assertEqual(calls[4][1]["creative"]["file_id"], "file_shared")
        self.assertEqual(len(state["ad_group_ids"]), 2)
        self.assertEqual(len(state["ad_ids"]), 2)

        delete_response = client.delete(f"/api/admin/adcopy/review-state/{snapshot_id}", headers=ADMIN_HEADERS)
        self.assertEqual(delete_response.status_code, 200)

    def test_admin_adcopy_review_state_saves_current_review_snapshot(self) -> None:
        client = TestClient(app, raise_server_exceptions=False)
        generated = generated_payload()
        generated["campaigns"] = [
            {
                "campaign_name": "검수저장_캠페인",
                "budget_max": 50000,
                "budget_type": "daily",
                "launch_date": "2026-07-10",
                "end_date": "2026-07-12",
                "objective": "Views",
                "target_countries": ["KR"],
            }
        ]
        generated["adgroups"][0].pop("max_bid", None)
        generated["ads"][0]["trace"]["validation_status"] = "승인"

        response = client.post(
            "/api/admin/adcopy/review-state",
            json={
                "advertiser_name": "캐츠잉글리시",
                "campaign_name": "검수저장_캠페인",
                "source_label": "unit-test",
                "generated": generated,
            },
            headers=ADMIN_HEADERS,
        )

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body["ok"])
        self.assertTrue(body["snapshot"]["id"])
        self.assertEqual(body["snapshot"]["campaign_name"], "검수저장_캠페인")
        self.assertIn("validation_report", body)
        snapshot_id = body["snapshot"]["id"]

        list_response = client.get("/api/admin/adcopy/review-state?limit=20", headers=ADMIN_HEADERS)
        self.assertEqual(list_response.status_code, 200)
        self.assertIn(snapshot_id, {item["id"] for item in list_response.json()["items"]})

        get_response = client.get(f"/api/admin/adcopy/review-state/{snapshot_id}", headers=ADMIN_HEADERS)
        self.assertEqual(get_response.status_code, 200)
        item = get_response.json()["item"]
        self.assertEqual(item["campaign_name"], "검수저장_캠페인")
        self.assertEqual(item["generated"]["campaigns"][0]["campaign_name"], "검수저장_캠페인")

        delete_response = client.delete(f"/api/admin/adcopy/review-state/{snapshot_id}", headers=ADMIN_HEADERS)
        self.assertEqual(delete_response.status_code, 200)

        missing_response = client.get(f"/api/admin/adcopy/review-state/{snapshot_id}", headers=ADMIN_HEADERS)
        self.assertEqual(missing_response.status_code, 404)

    def test_admin_adcopy_draft_audit_requires_admin_and_lists_recent_logs(self) -> None:
        from admin_store import save_adcopy_draft_audit_log

        client = TestClient(app, raise_server_exceptions=False)
        blocked = client.get("/api/admin/adcopy/draft-audit")
        self.assertEqual(blocked.status_code, 403)

        campaign_name = f"감사_{os.urandom(4).hex()}"
        saved = save_adcopy_draft_audit_log(
            {
                "advertiser_name": "캐츠잉글리시",
                "campaign_name": campaign_name,
                "action": "create_campaign",
                "status": "success",
                "message": "테스트 감사 로그",
                "state": {"campaign_id": "cmpn_audit", "ad_group_ids": {"01": "adgrp_1"}, "ad_ids": {}, "file_ids": {}},
                "logs": [{"level": "success", "message": "비활성 캠페인 임시 등록", "id": "cmpn_audit"}],
            }
        )
        self.assertTrue(saved["ok"])

        response = client.get("/api/admin/adcopy/draft-audit?limit=10", headers=ADMIN_HEADERS)
        self.assertEqual(response.status_code, 200)
        items = response.json()["items"]
        match = next(item for item in items if item["campaign_name"] == campaign_name)
        self.assertEqual(match["action"], "create_campaign")
        self.assertEqual(match["state"]["campaign_id"], "cmpn_audit")
        self.assertEqual(match["state"]["ad_group_count"], 1)
        invalid_limit_response = client.get("/api/admin/adcopy/draft-audit?limit=abc", headers=ADMIN_HEADERS)
        self.assertEqual(invalid_limit_response.status_code, 200)
        self.assertTrue(invalid_limit_response.json()["ok"])

    def test_admin_adcopy_draft_audit_csv_requires_admin_and_downloads_summary(self) -> None:
        from admin_store import save_adcopy_draft_audit_log

        client = TestClient(app, raise_server_exceptions=False)
        blocked = client.get("/api/admin/adcopy/draft-audit.csv")
        self.assertEqual(blocked.status_code, 403)

        campaign_name = f"CSV감사_{os.urandom(4).hex()}"
        save_adcopy_draft_audit_log(
            {
                "advertiser_name": "동서식품",
                "campaign_name": campaign_name,
                "action": "upload_assets",
                "status": "success",
                "message": "이미지 업로드 완료",
                "state": {"campaign_id": "cmpn_csv", "file_ids": {"img": "file_1"}, "ad_group_ids": {}, "ad_ids": {}},
                "logs": [{"level": "success", "message": "이미지 파일 등록", "file_id": "file_1"}],
            }
        )

        response = client.get("/api/admin/adcopy/draft-audit.csv?limit=10", headers=ADMIN_HEADERS)
        self.assertEqual(response.status_code, 200)
        self.assertIn("text/csv", response.headers.get("content-type", ""))
        self.assertIn("openai_ads_adcopy_draft_audit_", response.headers.get("content-disposition", ""))
        text = response.content.decode("utf-8-sig")
        self.assertIn("생성일,광고주,캠페인,단계,상태", text)
        self.assertIn(campaign_name, text)
        self.assertIn("이미지 업로드 완료", text)
        invalid_limit_response = client.get("/api/admin/adcopy/draft-audit.csv?limit=oops", headers=ADMIN_HEADERS)
        self.assertEqual(invalid_limit_response.status_code, 200)
        self.assertIn("text/csv", invalid_limit_response.headers.get("content-type", ""))

    def test_admin_adcopy_draft_execute_blocks_live_activation_by_default(self) -> None:
        client = TestClient(app, raise_server_exceptions=False)
        generated = generated_payload()
        generated["campaigns"] = [
            {
                "campaign_name": "live차단_캠페인",
                "budget_max": 100000,
                "budget_type": "total",
                "launch_date": "2026-07-10",
                "end_date": "2026-07-12",
                "objective": "Views",
                "target_countries": ["KR"],
            }
        ]
        generated["adgroups"][0].pop("max_bid", None)
        state = {
            "campaign_id": "cmpn_test",
            "ad_group_ids": {"01_반복훈련": "adgrp_test"},
            "ad_ids": {"01_반복훈련::AD_001": "ad_test"},
        }

        with patch("admin_store.get_ads_api_key_credential", return_value={"api_key": "sk-test"}), patch.dict(os.environ, {"ADS_DRAFT_ALLOW_ACTIVATION": "0"}):
            response = client.post(
                "/api/admin/adcopy/draft-execute",
                json={
                    "advertiser_name": "캐츠잉글리시",
                    "generated": generated,
                    "default_max_bid_krw": 7000,
                    "location_ids": [],
                    "action": "activate_all",
                    "state": state,
                    "confirm": True,
                },
                headers=ADMIN_HEADERS,
            )

        self.assertEqual(response.status_code, 400)
        self.assertIn("게시 전환은 비활성화", response.json()["detail"])

    def test_admin_adcopy_landing_inspect_requires_admin_password(self) -> None:
        client = TestClient(app, raise_server_exceptions=False)

        response = client.post("/api/admin/adcopy/inspect-landing", json={"landing_url": "https://example.com"})
        empty_response = client.post("/api/admin/adcopy/inspect-landing", json={})

        self.assertEqual(response.status_code, 403)
        self.assertEqual(empty_response.status_code, 403)

    def test_admin_adcopy_landing_inspect_extracts_public_metadata(self) -> None:
        client = TestClient(app, raise_server_exceptions=False)
        html = """
        <html>
          <head>
            <title>카누 디카페인 캡슐</title>
            <meta property="og:title" content="카누 디카페인 캡슐 커피">
            <meta name="description" content="부드러운 캐러멜 향과 돌체구스토 호환 캡슐">
            <meta property="og:image" content="/image.jpg">
            <link rel="Canonical" href="/canonical-product">
          </head>
        </html>
        """
        mocked_fetch = AsyncMock(return_value=(html, "https://example.com/product"))

        with patch("app._fetch_landing_html", mocked_fetch):
            response = client.post(
                "/api/admin/adcopy/inspect-landing",
                json={"landing_url": "https://example.com/product"},
                headers=ADMIN_HEADERS,
            )

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body["ok"])
        self.assertEqual(body["metadata"]["title"], "카누 디카페인 캡슐 커피")
        self.assertIn("돌체구스토 호환", body["metadata"]["description"])
        self.assertEqual(body["metadata"]["image_url"], "https://example.com/image.jpg")
        self.assertEqual(body["metadata"]["canonical_url"], "https://example.com/canonical-product")

    def test_admin_adcopy_landing_inspect_rejects_private_ip_url(self) -> None:
        client = TestClient(app, raise_server_exceptions=False)

        response = client.post(
            "/api/admin/adcopy/inspect-landing",
            json={"landing_url": "http://127.0.0.1:8000/private"},
            headers=ADMIN_HEADERS,
        )

        self.assertEqual(response.status_code, 400)

    def test_admin_adcopy_draft_plan_builds_paused_payloads(self) -> None:
        client = TestClient(app, raise_server_exceptions=False)
        generated = generated_payload()
        generated["campaigns"] = [
            {
                "campaign_name": "01_학습자료",
                "budget_max": 102600,
                "budget_type": "daily",
                "launch_date": "2026-07-01",
                "end_date": "2026-07-03",
                "objective": "Views",
                "target_countries": ["KR"],
            }
        ]
        generated["adgroups"][0].pop("max_bid", None)

        response = client.post(
            "/api/admin/adcopy/draft-plan",
            json={
                "advertiser_name": "캐츠잉글리시",
                "generated": generated,
                "default_max_bid_krw": 7000,
                "location_ids": ["2000043"],
            },
            headers=ADMIN_HEADERS,
        )

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body["ok"])
        self.assertEqual(body["safety"]["default_status"], "paused")
        self.assertEqual(body["campaign"]["api_payload"]["status"], "paused")
        self.assertEqual(body["campaign"]["api_payload"]["budget"]["lifetime_spend_limit_micros"], 307800000000)
        self.assertEqual(body["ad_groups"][0]["api_payload"]["status"], "paused")
        self.assertEqual(body["ad_groups"][0]["api_payload"]["bidding_config"]["max_bid_micros"], 7000000)
        self.assertEqual(body["ads"][0]["api_payload"]["status"], "paused")
        self.assertEqual(body["summary"]["ads"], 2)

    def test_admin_adcopy_draft_preflight_is_read_only(self) -> None:
        client = TestClient(app, raise_server_exceptions=False)
        generated = generated_payload()
        generated["campaigns"] = [
            {
                "campaign_name": "01_학습자료",
                "budget_max": 102600,
                "budget_type": "daily",
                "launch_date": "2026-07-01",
                "end_date": "2026-07-03",
                "objective": "Views",
                "target_countries": ["KR"],
            }
        ]
        generated["adgroups"][0].pop("max_bid", None)
        account = {"ok": True, "ad_account": {"name": "Catsenglish", "currency_code": "KRW", "timezone": "Asia/Seoul"}}

        with patch("admin_store.get_ads_api_key_credential", return_value={"advertiser_name": "캐츠잉글리시", "industry": "교육", "api_key": "sk-test"}), patch(
            "rag_chatbot.ads_api.fetch_ad_account_metadata",
            AsyncMock(return_value=account),
        ), patch("app._ads_draft_api_request", AsyncMock()) as mocked_request:
            response = client.post(
                "/api/admin/adcopy/draft-preflight",
                json={
                    "advertiser_name": "캐츠잉글리시",
                    "generated": generated,
                    "default_max_bid_krw": 7000,
                    "location_ids": ["2000043"],
                },
                headers=ADMIN_HEADERS,
            )

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body["ok"])
        self.assertEqual(body["readiness"], "draft_ready")
        self.assertTrue(body["state_patch"]["account_verified_at"])
        self.assertTrue(any(item["code"] == "activation_blocked" for item in body["checks"]))
        self.assertFalse(mocked_request.await_args_list)

    def test_admin_adcopy_draft_execute_requires_confirm_for_mutation(self) -> None:
        client = TestClient(app, raise_server_exceptions=False)
        generated = generated_payload()
        campaign_name = f"확인누락_{os.urandom(4).hex()}"
        generated["campaigns"] = [
            {
                "campaign_name": campaign_name,
                "budget_max": 100000,
                "budget_type": "total",
                "launch_date": "2026-07-01",
                "end_date": "2026-07-03",
                "objective": "Views",
                "target_countries": ["KR"],
            }
        ]
        generated["adgroups"][0].pop("max_bid", None)

        with patch("admin_store.get_ads_api_key_credential", return_value={"api_key": "sk-test", "industry": "교육"}):
            response = client.post(
                "/api/admin/adcopy/draft-execute",
                json={
                    "advertiser_name": "캐츠잉글리시",
                    "generated": generated,
                    "action": "create_campaign",
                    "confirm": False,
                },
                headers=ADMIN_HEADERS,
            )

        self.assertEqual(response.status_code, 400)
        self.assertIn("실행 확인", response.json()["detail"])
        audit_response = client.get("/api/admin/adcopy/draft-audit?limit=10", headers=ADMIN_HEADERS)
        self.assertEqual(audit_response.status_code, 200)
        audit_item = next(item for item in audit_response.json()["items"] if item["campaign_name"] == campaign_name)
        self.assertEqual(audit_item["status"], "error")
        self.assertIn("실행 확인", audit_item["message"])

    def test_admin_adcopy_draft_execute_audits_ads_api_errors(self) -> None:
        client = TestClient(app, raise_server_exceptions=False)
        generated = generated_payload()
        campaign_name = f"API오류_{os.urandom(4).hex()}"
        generated["campaigns"] = [
            {
                "campaign_name": campaign_name,
                "budget_max": 100000,
                "budget_type": "total",
                "launch_date": "2026-07-01",
                "end_date": "2026-07-03",
                "objective": "Views",
                "target_countries": ["KR"],
            }
        ]
        request = httpx.Request("POST", "https://api.ads.openai.com/v1/campaigns")
        response = httpx.Response(403, json={"error": {"message": "test ads api blocked"}}, request=request)
        error = httpx.HTTPStatusError("blocked", request=request, response=response)

        with patch("admin_store.get_ads_api_key_credential", return_value={"api_key": "sk-test", "industry": "교육"}):
            with patch("app._ads_draft_api_request", AsyncMock(side_effect=error)):
                draft_response = client.post(
                    "/api/admin/adcopy/draft-execute",
                    json={
                        "advertiser_name": "캐츠잉글리시",
                        "generated": generated,
                        "action": "create_campaign",
                        "confirm": True,
                    },
                    headers=ADMIN_HEADERS,
                )

        self.assertEqual(draft_response.status_code, 502)
        self.assertIn("test ads api blocked", draft_response.json()["detail"])
        audit_response = client.get("/api/admin/adcopy/draft-audit?limit=10", headers=ADMIN_HEADERS)
        audit_item = next(item for item in audit_response.json()["items"] if item["campaign_name"] == campaign_name)
        self.assertEqual(audit_item["status"], "error")
        self.assertIn("test ads api blocked", audit_item["message"])

    def test_admin_adcopy_draft_execute_creates_campaign_paused(self) -> None:
        client = TestClient(app, raise_server_exceptions=False)
        generated = generated_payload()
        generated["campaigns"] = [
            {
                "campaign_name": "01_학습자료",
                "budget_max": 100000,
                "budget_type": "total",
                "launch_date": "2026-07-01",
                "end_date": "2026-07-03",
                "objective": "Views",
                "target_countries": ["KR"],
            }
        ]
        generated["adgroups"][0].pop("max_bid", None)
        mocked_request = AsyncMock(return_value={"id": "cmpn_test", "status": "paused"})

        with patch("admin_store.get_ads_api_key_credential", return_value={"api_key": "sk-test", "industry": "교육"}):
            with patch("app._ads_draft_api_request", mocked_request):
                response = client.post(
                    "/api/admin/adcopy/draft-execute",
                    json={
                        "advertiser_name": "캐츠잉글리시",
                        "generated": generated,
                        "action": "create_campaign",
                        "confirm": True,
                    },
                    headers=ADMIN_HEADERS,
                )

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertEqual(body["state"]["campaign_id"], "cmpn_test")
        called_args = mocked_request.await_args.args
        self.assertEqual(called_args[1], "POST")
        self.assertEqual(called_args[2], "/v1/campaigns")
        self.assertEqual(called_args[3]["status"], "paused")

    def test_admin_adcopy_draft_execute_completes_paused_setup_sequence(self) -> None:
        client = TestClient(app, raise_server_exceptions=False)
        generated = generated_payload()
        generated["campaigns"] = [
            {
                "campaign_name": "01_학습자료",
                "budget_max": 100000,
                "budget_type": "total",
                "launch_date": "2026-07-01",
                "end_date": "2026-07-03",
                "objective": "Views",
                "target_countries": ["KR"],
            }
        ]
        generated["adgroups"][0].pop("max_bid", None)
        calls: list[tuple[str, dict]] = []

        async def mocked_request(api_key: str, method: str, path: str, json_body: dict | None = None) -> dict:
            calls.append((path, json_body or {}))
            if path == "/v1/campaigns":
                return {"id": "cmpn_test", "status": "paused"}
            if path == "/v1/upload":
                return {"file_id": "file_test"}
            if path == "/v1/ad_groups":
                return {"id": "adgrp_test", "status": "paused"}
            if path == "/v1/ads":
                return {"id": f"ad_test_{sum(1 for item in calls if item[0] == '/v1/ads')}", "status": "paused"}
            raise AssertionError(path)

        base_payload = {
            "advertiser_name": "캐츠잉글리시",
            "generated": generated,
            "default_max_bid_krw": 7000,
            "location_ids": ["2000043"],
        }
        state: dict = {}

        with patch("admin_store.get_ads_api_key_credential", return_value={"api_key": "sk-test", "industry": "교육"}), patch("app._ads_draft_api_request", mocked_request):
            for action in ["create_campaign", "upload_assets", "create_ad_groups", "create_ads"]:
                response = client.post(
                    "/api/admin/adcopy/draft-execute",
                    json={
                        **base_payload,
                        "action": action,
                        "state": state,
                        "confirm": True,
                    },
                    headers=ADMIN_HEADERS,
                )
                self.assertEqual(response.status_code, 200)
                state = response.json()["state"]

        paths = [path for path, _payload in calls]
        self.assertEqual(paths, ["/v1/campaigns", "/v1/upload", "/v1/ad_groups", "/v1/ads", "/v1/ads"])
        self.assertTrue(all("/activate" not in path for path in paths))
        self.assertEqual(calls[0][1]["status"], "paused")
        self.assertEqual(calls[2][1]["status"], "paused")
        self.assertEqual(calls[2][1]["campaign_id"], "cmpn_test")
        self.assertEqual(calls[3][1]["status"], "paused")
        self.assertEqual(calls[3][1]["ad_group_id"], "adgrp_test")
        self.assertEqual(calls[3][1]["creative"]["file_id"], "file_test")
        self.assertEqual(len(state["ad_ids"]), 2)


if __name__ == "__main__":
    unittest.main()

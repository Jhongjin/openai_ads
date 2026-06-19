from __future__ import annotations

from datetime import datetime, timedelta
import unittest
from zoneinfo import ZoneInfo

from pydantic import ValidationError

from intake import IntakeSubmission, build_sheet_payload, create_workbook_bytes, inspect_workbook_bytes


KST = ZoneInfo("Asia/Seoul")


def valid_payload() -> dict:
    tomorrow = datetime.now(KST).date() + timedelta(days=1)
    end = tomorrow + timedelta(days=14)
    return {
        "opsMeta": {
            "advertiserName": "테스트 광고주",
            "adsManagerAccount": "테스트 브랜드",
            "salesOwner": "케이티나스 담당자",
            "salesOwnerEmail": "owner@nasmedia.co.kr",
            "ownerHeadquarters": "미디어본부",
            "ownerOffice": "미디어채널실",
            "ownerTeam": "미디어채널1팀",
            "notes": "테스트",
            "honeypot": "",
            "formStartedAt": int(datetime.now(KST).timestamp() * 1000) - 3000,
        },
        "campaign": {
            "campaign_name": "ChatGPT_Ads_Test",
            "budget_max": 4000000,
            "budget_type": "lifetime",
            "launch_date": tomorrow.isoformat(),
            "end_date": end.isoformat(),
            "objective": "clicks",
            "target_countries": [],
        },
        "adgroup": {
            "campaign_name": "ChatGPT_Ads_Test",
            "adgroup_name": "test_group",
            "max_bid": 1200,
            "keywords": ["밀키트", "저녁"],
        },
        "ads": [
            {
                "adgroup_name": "test_group",
                "ad_name": "test_ad_1",
                "title": "건강한 저녁 준비",
                "copy": "빠르게 차리는 균형 잡힌 한 끼",
                "link": "https://example.com/landing?utm_source=openai",
                "image_link": "https://example.com/image1.jpg",
            },
            {
                "adgroup_name": "test_group",
                "ad_name": "test_ad_2",
                "title": "퇴근 후 간편 식사",
                "copy": "오늘 저녁 고민을 줄여주는 밀키트",
                "link": "https://example.com/landing-2",
                "image_link": "https://example.com/image2.png",
            },
        ],
    }


class IntakeValidationTests(unittest.TestCase):
    def test_valid_payload_builds_four_sheet_sections(self) -> None:
        submission = IntakeSubmission.model_validate(valid_payload())

        payload = build_sheet_payload(
            submission,
            shared_secret="secret",
        )

        self.assertEqual(payload["secret"], "secret")
        self.assertEqual(payload["data"]["primary_campaign"]["campaign_name"], "ChatGPT_Ads_Test")
        self.assertEqual(payload["data"]["campaign"][0]["campaign_name"], "ChatGPT_Ads_Test")
        self.assertEqual(payload["data"]["campaign"][0]["budget_max"], "4000000")
        self.assertEqual(payload["data"]["campaign"][0]["target_countries"], [])
        self.assertEqual(payload["data"]["campaigns"][0]["campaign_name"], "ChatGPT_Ads_Test")
        self.assertEqual(len(payload["data"]["adgroups"]), 1)
        self.assertEqual(len(payload["data"]["ads"]), 2)
        self.assertEqual(payload["data"]["ads"][0]["ad_name"], "test_ad_1")
        self.assertEqual(payload["data"]["adgroups"][0]["keywords"], ["밀키트", "저녁"])
        self.assertEqual(payload["data"]["adgroups"][0]["max_bid"], "1200")
        self.assertEqual(payload["data"]["ops"]["advertiser_name"], "테스트 광고주")
        self.assertEqual(payload["data"]["ops"]["brand_name"], "테스트 브랜드")
        self.assertEqual(payload["data"]["ops"]["sales_owner_email"], "owner@nasmedia.co.kr")
        self.assertNotIn("upload_mode", payload["data"]["ops"])
        self.assertNotIn("route", payload["data"]["ops"])
        self.assertNotIn("legal_name", payload["data"]["ops"])
        self.assertNotIn("brn", payload["data"]["ops"])
        self.assertNotIn("homepage", payload["data"]["ops"])
        self.assertNotIn("invoice_email", payload["data"]["ops"])
        self.assertNotIn("submitter_name", payload["data"]["ops"])
        self.assertNotIn("submitter_email", payload["data"]["ops"])
        self.assertNotIn("honeypot", payload["data"]["ops"])
        self.assertNotIn("receiptNumber", payload)

    def test_multi_campaign_payload_preserves_each_campaign_and_adgroup_link(self) -> None:
        payload = valid_payload()
        tomorrow = datetime.now(KST).date() + timedelta(days=1)
        end = tomorrow + timedelta(days=14)
        payload.pop("campaign")
        payload.pop("adgroup")
        payload.pop("ads")
        payload["campaigns"] = [
            {
                "campaign": {
                    "campaign_name": "campaign_cpm",
                    "budget_max": 4000000,
                    "budget_type": "lifetime",
                    "launch_date": tomorrow.isoformat(),
                    "end_date": end.isoformat(),
                    "objective": "views",
                    "target_countries": [],
                },
                "adgroups": [
                    {
                        "adgroup": {
                            "adgroup_name": "ag_cpm",
                            "max_bid": None,
                            "keywords": ["노출"],
                        },
                        "ads": [
                            {
                                "ad_name": "cpm_ad_1",
                                "title": "노출 캠페인 소재",
                                "copy": "브랜드 메시지를 넓게 알립니다",
                                "link": "https://example.com/cpm",
                                "image_link": "https://example.com/cpm.png",
                            }
                        ],
                    }
                ],
            },
            {
                "campaign": {
                    "campaign_name": "campaign_cpc",
                    "budget_max": 4000000,
                    "budget_type": "daily",
                    "launch_date": tomorrow.isoformat(),
                    "end_date": end.isoformat(),
                    "objective": "clicks",
                    "target_countries": [],
                },
                "adgroups": [
                    {
                        "adgroup": {
                            "adgroup_name": "ag_cpc",
                            "max_bid": 4100,
                            "keywords": ["클릭"],
                        },
                        "ads": [
                            {
                                "ad_name": "cpc_ad_1",
                                "title": "클릭 캠페인 소재",
                                "copy": "랜딩 페이지 방문을 유도합니다",
                                "link": "https://example.com/cpc",
                                "image_link": "https://example.com/cpc.png",
                            }
                        ],
                    }
                ],
            },
        ]

        submission = IntakeSubmission.model_validate(payload)
        sheet_payload = build_sheet_payload(submission, shared_secret="secret")

        self.assertEqual(
            [item["campaign_name"] for item in sheet_payload["data"]["campaigns"]],
            ["campaign_cpm", "campaign_cpc"],
        )
        self.assertEqual(
            [item["campaign_name"] for item in sheet_payload["data"]["campaign"]],
            ["campaign_cpm", "campaign_cpc"],
        )
        self.assertEqual(sheet_payload["data"]["primary_campaign"]["campaign_name"], "campaign_cpm")
        self.assertEqual(
            [item["objective"] for item in sheet_payload["data"]["campaigns"]],
            ["views", "clicks"],
        )
        self.assertEqual(
            [(item["campaign_name"], item["adgroup_name"]) for item in sheet_payload["data"]["adgroups"]],
            [("campaign_cpm", "ag_cpm"), ("campaign_cpc", "ag_cpc")],
        )
        self.assertEqual(len(sheet_payload["data"]["ads"]), 2)

    def test_rejects_max_bid_for_views_campaign(self) -> None:
        payload = valid_payload()
        payload["campaign"]["objective"] = "views"
        payload["adgroup"]["max_bid"] = 91000

        with self.assertRaises(ValidationError):
            IntakeSubmission.model_validate(payload)

    def test_rejects_missing_max_bid_for_clicks_campaign(self) -> None:
        payload = valid_payload()
        payload["campaign"]["objective"] = "clicks"
        payload["adgroup"]["max_bid"] = None

        with self.assertRaises(ValidationError):
            IntakeSubmission.model_validate(payload)

    def test_rejects_duplicate_ad_names(self) -> None:
        payload = valid_payload()
        payload["ads"][1]["ad_name"] = payload["ads"][0]["ad_name"]

        with self.assertRaises(ValidationError):
            IntakeSubmission.model_validate(payload)

    def test_allows_space_hyphen_and_underscore_in_names(self) -> None:
        payload = valid_payload()
        payload["campaign"]["campaign_name"] = "campaign 1-test_main"
        payload["adgroup"]["campaign_name"] = "campaign 1-test_main"
        payload["adgroup"]["adgroup_name"] = "adgroup 1-test_main"
        for ad in payload["ads"]:
            ad["adgroup_name"] = "adgroup 1-test_main"

        IntakeSubmission.model_validate(payload)

    def test_rejects_dot_and_other_special_characters_in_names(self) -> None:
        payload = valid_payload()
        payload["campaign"]["campaign_name"] = "campaign.1"

        with self.assertRaises(ValidationError):
            IntakeSubmission.model_validate(payload)

    def test_rejects_reversed_dates(self) -> None:
        payload = valid_payload()
        payload["campaign"]["end_date"] = payload["campaign"]["launch_date"]
        payload["campaign"]["launch_date"] = (
            datetime.fromisoformat(payload["campaign"]["end_date"]).date() + timedelta(days=1)
        ).isoformat()

        with self.assertRaises(ValidationError):
            IntakeSubmission.model_validate(payload)

    def test_rejects_missing_campaign_operational_required_fields(self) -> None:
        for field in ("budget_type", "launch_date", "end_date", "objective"):
            payload = valid_payload()
            payload["campaign"].pop(field)
            with self.subTest(field=field):
                with self.assertRaises(ValidationError):
                    IntakeSubmission.model_validate(payload)

    def test_legacy_criteo_route_no_longer_changes_objective(self) -> None:
        payload = valid_payload()
        payload["opsMeta"]["executionRoute"] = "criteo"
        payload["campaign"]["objective"] = "clicks"

        IntakeSubmission.model_validate(payload)

    def test_openai_title_and_copy_limits_are_official_workbook_limits(self) -> None:
        payload = valid_payload()
        payload["ads"][0]["title"] = "가" * 25

        with self.assertRaises(ValidationError):
            IntakeSubmission.model_validate(payload)

        payload = valid_payload()
        payload["ads"][0]["copy"] = "나" * 49

        with self.assertRaises(ValidationError):
            IntakeSubmission.model_validate(payload)

    def test_legacy_criteo_route_uses_workbook_limits(self) -> None:
        payload = valid_payload()
        payload["opsMeta"]["executionRoute"] = "criteo"
        payload["campaign"]["objective"] = "views"
        payload["adgroup"]["max_bid"] = None
        payload["ads"][0]["title"] = "가" * 24
        payload["ads"][0]["copy"] = "나" * 48
        IntakeSubmission.model_validate(payload)

        payload["ads"][0]["title"] = "가" * 25
        with self.assertRaises(ValidationError):
            IntakeSubmission.model_validate(payload)

    def test_new_creative_upload_meta_payload_is_supported(self) -> None:
        payload = valid_payload()
        payload["opsMeta"] = {
            "advertiserName": "테스트 광고주",
            "adsManagerAccount": "Test Ads Account",
            "salesOwner": "케이티나스 담당자",
            "salesOwnerEmail": "owner@nasmedia.co.kr",
            "ownerHeadquarters": "미디어본부",
            "ownerOffice": "미디어채널실",
            "ownerTeam": "미디어채널1팀",
            "notes": "",
            "honeypot": "",
            "formStartedAt": int(datetime.now(KST).timestamp() * 1000) - 3000,
        }

        submission = IntakeSubmission.model_validate(payload)
        sheet_payload = build_sheet_payload(submission, shared_secret="secret")

        self.assertNotIn("upload_mode", sheet_payload["data"]["ops"])
        self.assertEqual(sheet_payload["data"]["ops"]["brand_name"], "Test Ads Account")
        self.assertEqual(sheet_payload["data"]["ops"]["sales_owner_email"], "owner@nasmedia.co.kr")
        self.assertEqual(sheet_payload["data"]["ops"]["owner_office"], "미디어채널실")

    def test_brand_is_optional_for_creative_upload_meta(self) -> None:
        payload = valid_payload()
        payload["opsMeta"].pop("adsManagerAccount", None)

        submission = IntakeSubmission.model_validate(payload)
        sheet_payload = build_sheet_payload(submission, shared_secret="secret")

        self.assertEqual(sheet_payload["data"]["ops"]["brand_name"], "")
        self.assertNotIn("submitter_name", sheet_payload["data"]["ops"])
        self.assertNotIn("submitter_email", sheet_payload["data"]["ops"])

    def test_can_create_and_inspect_ads_manager_workbook(self) -> None:
        submission = IntakeSubmission.model_validate(valid_payload())
        workbook = create_workbook_bytes(submission)
        summary = inspect_workbook_bytes(workbook)

        self.assertTrue(summary["ok"])
        self.assertEqual(summary["sheets"]["campaigns"]["rows"], 1)
        self.assertEqual(summary["sheets"]["adgroups"]["rows"], 1)
        self.assertEqual(summary["sheets"]["ads"]["rows"], 2)
        self.assertEqual(summary["data"]["campaigns"][0]["campaign_name"], "ChatGPT_Ads_Test")
        self.assertEqual(summary["data"]["adgroups"][0]["adgroup_name"], "test_group")
        self.assertEqual(summary["data"]["ads"][0]["ad_name"], "test_ad_1")
        self.assertEqual(summary["data"]["ads"][0]["title"], "건강한 저녁 준비")

    def test_inspect_workbook_rejects_duplicate_ad_names(self) -> None:
        from io import BytesIO

        from openpyxl import Workbook

        wb = Workbook()
        wb.remove(wb.active)
        rows_by_sheet = {
            "campaigns": [
                ["campaign_name", "budget_max", "budget_type", "launch_date", "end_date", "objective", "target_countries"],
                ["nasmedia2188", 5000, "lifetime", "2026-06-30", "2026-07-06", "views", ""],
            ],
            "adgroups": [
                ["campaign_name", "adgroup_name", "max_bid", "keywords"],
                ["nasmedia2188", "nasmedia2188_ag1", "", ""],
            ],
            "ads": [
                ["adgroup_name", "ad_name", "title", "copy", "link", "image_link"],
                ["nasmedia2188_ag1", "nasmedia2188_ad1", "첫번째", "소재 설명입니다", "https://example.com/1", "https://example.com/1.png"],
                ["nasmedia2188_ag1", "nasmedia2188_ad1", "두번째", "소재 설명입니다", "https://example.com/2", "https://example.com/2.png"],
            ],
        }
        for sheet_name, rows in rows_by_sheet.items():
            ws = wb.create_sheet(sheet_name)
            for row in rows:
                ws.append(row)
        buffer = BytesIO()
        wb.save(buffer)

        summary = inspect_workbook_bytes(buffer.getvalue())

        self.assertFalse(summary["ok"])
        self.assertTrue(any("ad_name 'nasmedia2188_ad1'이 중복" in error for error in summary["errors"]))

    def test_inspect_workbook_warns_when_only_template_samples_exist(self) -> None:
        from io import BytesIO

        from openpyxl import Workbook

        wb = Workbook()
        wb.remove(wb.active)
        rows_by_sheet = {
            "campaigns": [
                ["campaign_name", "budget_max", "budget_type", "launch_date", "end_date", "objective", "target_countries"],
                ["oaitestcmp1234567", 5000, "Lifetime", "2026-02-09", "2026-03-31", "Views", '["US", "CA"]'],
            ],
            "adgroups": [
                ["campaign_name", "adgroup_name", "max_bid", "keywords"],
                ["oaitestcmp1234567", "oaitestadg9876543", 3, '["test1", "test2"]'],
            ],
            "ads": [
                ["adgroup_name", "ad_name", "title", "copy", "link", "image_link"],
                [
                    "oaitestadg9876543",
                    "oaitestad1234567",
                    "My website",
                    "Check out the latest on my website",
                    "https://www.example.com/?utm_source=chat",
                    "https://www.example.com/logo.png",
                ],
            ],
        }
        for sheet_name, rows in rows_by_sheet.items():
            ws = wb.create_sheet(sheet_name)
            for row in rows:
                ws.append(row)
        buffer = BytesIO()
        wb.save(buffer)

        summary = inspect_workbook_bytes(buffer.getvalue())

        self.assertFalse(summary["ok"])
        self.assertEqual(summary["sheets"]["campaigns"]["rows"], 0)
        self.assertEqual(summary["sheets"]["campaigns"]["sample_rows"], 1)
        self.assertTrue(any("oaitest" in warning for warning in summary["warnings"]))

    def test_inspect_workbook_rejects_views_max_bid_and_unknown_adgroup(self) -> None:
        from io import BytesIO

        from openpyxl import Workbook

        wb = Workbook()
        wb.remove(wb.active)
        rows_by_sheet = {
            "campaigns": [
                ["campaign_name", "budget_max", "budget_type", "launch_date", "end_date", "objective", "target_countries"],
                ["nasmedia2188", 5000, "lifetime", "2026-06-30", "2026-07-06", "views", '["KR"]'],
            ],
            "adgroups": [
                ["campaign_name", "adgroup_name", "max_bid", "keywords"],
                ["nasmedia2188", "nasmedia2188_ag1", 3, ""],
            ],
            "ads": [
                ["adgroup_name", "ad_name", "title", "copy", "link", "image_link"],
                ["nasmedia2188", "nasmedia2188_ad1", "내사이트야", "내사이트라고", "https://example.com", "https://example.com/logo.png"],
            ],
        }
        for sheet_name, rows in rows_by_sheet.items():
            ws = wb.create_sheet(sheet_name)
            for row in rows:
                ws.append(row)
        buffer = BytesIO()
        wb.save(buffer)

        summary = inspect_workbook_bytes(buffer.getvalue())

        self.assertFalse(summary["ok"])
        self.assertTrue(any("Views(CPM)" in error and "max_bid" in error for error in summary["errors"]))
        self.assertTrue(any("존재하지 않는 adgroup_name" in error for error in summary["errors"]))

    def test_inspect_workbook_rejects_kr_target_country_but_accepts_http_image_url_shape(self) -> None:
        from io import BytesIO

        from openpyxl import Workbook

        wb = Workbook()
        wb.remove(wb.active)
        rows_by_sheet = {
            "campaigns": [
                ["campaign_name", "budget_max", "budget_type", "launch_date", "end_date", "objective", "target_countries"],
                ["nasmedia2188", 5000, "lifetime", "2026-06-30", "2026-07-06", "views", '["KR"]'],
            ],
            "adgroups": [
                ["campaign_name", "adgroup_name", "max_bid", "keywords"],
                ["nasmedia2188", "nasmedia2188_ag1", "", ""],
            ],
            "ads": [
                ["adgroup_name", "ad_name", "title", "copy", "link", "image_link"],
                ["nasmedia2188_ag1", "nasmedia2188_ad1", "내사이트야", "내사이트라고", "https://example.com", "https://openads.admate.ai.kr/"],
            ],
        }
        for sheet_name, rows in rows_by_sheet.items():
            ws = wb.create_sheet(sheet_name)
            for row in rows:
                ws.append(row)
        buffer = BytesIO()
        wb.save(buffer)

        summary = inspect_workbook_bytes(buffer.getvalue())

        self.assertFalse(summary["ok"])
        self.assertTrue(any("target_countries에 KR" in error for error in summary["errors"]))
        self.assertFalse(any("image_link" in error for error in summary["errors"]))

    def test_inspect_workbook_rejects_malformed_landing_and_image_urls(self) -> None:
        from io import BytesIO

        from openpyxl import Workbook

        wb = Workbook()
        wb.remove(wb.active)
        rows_by_sheet = {
            "campaigns": [
                ["campaign_name", "budget_max", "budget_type", "launch_date", "end_date", "objective", "target_countries"],
                ["nasmedia2188", 5000, "lifetime", "2026-06-30", "2026-07-06", "views", ""],
            ],
            "adgroups": [
                ["campaign_name", "adgroup_name", "max_bid", "keywords"],
                ["nasmedia2188", "nasmedia2188_ag1", "", ""],
            ],
            "ads": [
                ["adgroup_name", "ad_name", "title", "copy", "link", "image_link"],
                ["nasmedia2188_ag1", "nasmedia2188_ad1", "내사이트야", "내사이트라고", "openads.admate.ai.kr/", "https://co.kr.."],
            ],
        }
        for sheet_name, rows in rows_by_sheet.items():
            ws = wb.create_sheet(sheet_name)
            for row in rows:
                ws.append(row)
        buffer = BytesIO()
        wb.save(buffer)

        summary = inspect_workbook_bytes(buffer.getvalue())

        self.assertFalse(summary["ok"])
        self.assertTrue(any("link는 http:// 또는 https://" in error for error in summary["errors"]))
        self.assertTrue(any("image_link는 http:// 또는 https://" in error for error in summary["errors"]))


if __name__ == "__main__":
    unittest.main()

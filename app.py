from __future__ import annotations

import asyncio
import csv
from datetime import datetime, timedelta, timezone
import ipaddress
from io import BytesIO, StringIO
import json
import os
from pathlib import Path
import re
import socket
from uuid import uuid4
from typing import Any
from urllib.parse import urljoin, urlparse

import httpx
from bs4 import BeautifulSoup
from fastapi import FastAPI, File, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, PlainTextResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field, ValidationError

from rag_chatbot.config import project_root


app = FastAPI(title="Nasmedia ChatGPT Ads RAG", version="0.1.0")
app.mount(
    "/images",
    StaticFiles(directory=project_root() / "public" / "images", check_dir=False),
    name="images",
)
app.mount(
    "/dev-assets",
    StaticFiles(directory=project_root() / "dev" / "assets", check_dir=False),
    name="dev-assets",
)
app.mount(
    "/assets",
    StaticFiles(directory=project_root() / "dev" / "assets", check_dir=False),
    name="assets",
)


IpNetwork = ipaddress.IPv4Network | ipaddress.IPv6Network
IpAddress = ipaddress.IPv4Address | ipaddress.IPv6Address
ADS_DASHBOARD_CACHE_KEY = "active_advertisers"
KST = timezone(timedelta(hours=9))


def _configured_access_networks() -> list[IpNetwork]:
    raw = os.getenv("ACCESS_ALLOWED_IPS", "").strip()
    if not raw:
        return []
    networks: list[IpNetwork] = []
    for item in re.split(r"[,\s]+", raw):
        value = item.strip()
        if not value:
            continue
        try:
            networks.append(ipaddress.ip_network(value, strict=False))
        except ValueError:
            continue
    return networks


def _request_client_ip(request: Request) -> IpAddress | None:
    header_value = (
        request.headers.get("x-vercel-forwarded-for")
        or request.headers.get("x-forwarded-for")
        or request.headers.get("x-real-ip")
        or ""
    )
    candidate = header_value.split(",", 1)[0].strip()
    if not candidate and request.client:
        candidate = request.client.host
    if not candidate:
        return None
    candidate = candidate.strip("[]")
    try:
        return ipaddress.ip_address(candidate)
    except ValueError:
        return None


@app.middleware("http")
async def restrict_access_by_ip(request: Request, call_next):
    networks = _configured_access_networks()
    if not networks:
        return await call_next(request)
    client_ip = _request_client_ip(request)
    if client_ip and any(client_ip in network for network in networks):
        return await call_next(request)
    return PlainTextResponse("허용된 사내 네트워크에서만 접속할 수 있습니다.", status_code=403)

class ChatRequest(BaseModel):
    question: str = Field(..., min_length=1)


class ChatResponse(BaseModel):
    answer: str
    sources: list[dict[str, Any]]


class CheckRequest(BaseModel):
    urls: list[str] = Field(..., min_length=1, max_length=100)


class CheckResultResponse(BaseModel):
    input_url: str
    normalized_url: str
    origin: str
    path: str
    robots_url: str
    verdict: str
    badge: str
    reason: str
    action: str
    http_status: int | None
    robots_txt: str
    firewall_hint: bool = False
    firewall_badge: str | None = None
    developer_message: str | None = None
    developer_checklist: str | None = None
    official_guide_url: str | None = None
    official_crawler_doc_url: str | None = None
    bot_details: list[dict[str, str]] = Field(default_factory=list)


class FaviconCheckRequest(BaseModel):
    urls: list[str] = Field(..., min_length=1, max_length=100)


class FaviconCheckResultResponse(BaseModel):
    input_url: str
    normalized_url: str
    verdict: str
    badge: str
    size: str
    width: int | None
    height: int | None
    format: str
    background: str
    reason: str
    action: str
    http_status: int | None
    preview_url: str | None


class IngestResponse(BaseModel):
    counts: dict[str, int]


class IntakeResponse(BaseModel):
    receipt_number: str
    submitted_at_kst: str
    mail_sent: bool | None = None
    mail_error: str | None = None
    mail_sender: str | None = None
    mail_recipient: str | None = None
    mail_cc: str | None = None
    mail_quota_remaining: int | None = None


class WorkbookInspectResponse(BaseModel):
    ok: bool
    sheets: dict[str, Any]
    errors: list[str]
    warnings: list[str] = Field(default_factory=list)
    data: dict[str, Any] | None = None


class ImageUploadResponse(BaseModel):
    image_url: str
    width: int
    height: int
    content_type: str


class NoticeConfigRequest(BaseModel):
    title: str = Field(..., min_length=1, max_length=120)
    body_html: str | None = Field(default=None, max_length=8000)
    modal_background: str | None = Field(default="#ffffff", max_length=20)
    bullets: list[str] = Field(default_factory=list, max_length=20)
    source_label: str = Field(..., min_length=1, max_length=200)
    source_url: str = Field(..., min_length=1, max_length=300)
    enabled: bool = True


class MenuSettingsRequest(BaseModel):
    menus: dict[str, bool] = Field(default_factory=dict)
    guide_decks: dict[str, bool] = Field(default_factory=dict)


class SlideContentItemRequest(BaseModel):
    key: str = Field(..., min_length=1, max_length=120)
    value: str = Field(default="", max_length=1200)
    deck: str | None = Field(default="", max_length=40)
    label: str | None = Field(default="", max_length=200)
    default: str | None = Field(default="", max_length=1200)
    multiline: bool | None = False
    alt: str | None = Field(default="", max_length=300)
    caption: str | None = Field(default="", max_length=300)


class SlideContentRequest(BaseModel):
    items: list[SlideContentItemRequest] = Field(default_factory=list, max_length=700)
    images: list[SlideContentItemRequest] = Field(default_factory=list, max_length=120)
    layout: dict[str, Any] | None = Field(default_factory=dict)


class MailReviewUpdateRequest(BaseModel):
    duplicate_hash: str = Field(..., min_length=8, max_length=128)
    review_status: str = Field(..., min_length=1, max_length=40)
    review_note: str | None = Field(default="", max_length=2000)
    approved_title: str | None = Field(default="", max_length=300)
    approved_summary: str | None = Field(default="", max_length=8000)
    approved_by: str | None = Field(default="", max_length=80)
    supersedes_duplicate_hash: str | None = Field(default="", max_length=128)


class ManualRagItemRequest(BaseModel):
    title: str = Field(..., min_length=2, max_length=300)
    category: str | None = Field(default="", max_length=120)
    source_note: str | None = Field(default="", max_length=1000)
    content: str = Field(..., min_length=1, max_length=20000)


class FaqCategoryRequest(BaseModel):
    id: str | None = Field(default="", max_length=40)
    label: str = Field(..., min_length=1, max_length=80)
    sort_order: int | None = Field(default=0)


class FaqItemRequest(BaseModel):
    category: str = Field(..., min_length=1, max_length=40)
    question: str = Field(..., min_length=1, max_length=500)
    answer: str = Field(..., min_length=1, max_length=3000)
    source_summary: str | None = Field(default="관리자 FAQ 관리", max_length=300)


class CampaignIntakeOpsRequest(BaseModel):
    receipt_number: str = Field(..., min_length=3, max_length=80)
    operator_name: str | None = Field(default="", max_length=80)
    status: str = Field(default="ready", min_length=1, max_length=40)
    memo: str | None = Field(default="", max_length=1000)


class VisitRequest(BaseModel):
    page: str = Field(..., min_length=1, max_length=80)
    label: str | None = Field(default=None, max_length=120)


class AdsApiKeyRequest(BaseModel):
    advertiser_name: str = Field(..., min_length=1, max_length=120)
    industry: str | None = Field(default="", max_length=40)
    ads_api_key: str | None = Field(default="", max_length=500)
    conversion_api_key: str | None = Field(default="", max_length=500)
    enabled: bool = True


class AdsApiKeyInspectRequest(BaseModel):
    ads_api_key: str = Field(..., min_length=1, max_length=500)


class AdsCampaignObjectiveRequest(BaseModel):
    advertiser_name: str = Field(..., min_length=1, max_length=120)
    campaign_id: str = Field(..., min_length=1, max_length=160)
    campaign_name: str | None = Field(default="", max_length=1000)
    objective: str | None = Field(default="", max_length=40)


class AdsAdcopyGenerateRequest(BaseModel):
    engine: str = Field(default="admate", min_length=1, max_length=30)
    advertiser_name: str = Field(..., min_length=1, max_length=120)
    industry: str | None = Field(default="", max_length=40)
    campaign_name: str = Field(..., min_length=1, max_length=160)
    objective: str = Field(default="Views", min_length=1, max_length=40)
    budget_max: float | None = Field(default=None, ge=0)
    budget_type: str = Field(default="daily", min_length=1, max_length=20)
    launch_date: str | None = Field(default="", max_length=20)
    end_date: str | None = Field(default="", max_length=20)
    target_countries: list[str] = Field(default_factory=lambda: ["KR"], max_length=20)
    product_name: str = Field(..., min_length=1, max_length=200)
    landing_url: str = Field(..., min_length=1, max_length=1000)
    image_link: str | None = Field(default="", max_length=1000)
    audience: str | None = Field(default="", max_length=1000)
    selling_points: str | None = Field(default="", max_length=2000)
    tone: str | None = Field(default="", max_length=300)
    banned_terms: str | None = Field(default="", max_length=1500)
    required_phrases: str | None = Field(default="", max_length=1500)
    required_exact_phrases: str | None = Field(default="", max_length=1500)
    adgroup_count: int = Field(default=3, ge=1, le=12)
    ads_per_adgroup: int = Field(default=3, ge=1, le=8)


class AdsAdcopyLandingInspectRequest(BaseModel):
    landing_url: str = Field(..., min_length=1, max_length=1000)


class AdsAdcopyImportRequest(BaseModel):
    generated: dict[str, Any] = Field(default_factory=dict)
    advertiser_name: str | None = Field(default="", max_length=120)
    industry: str | None = Field(default="", max_length=40)
    campaign_name: str | None = Field(default="", max_length=160)
    objective: str | None = Field(default="Views", max_length=40)
    budget_max: float | None = Field(default=None, ge=0)
    budget_type: str | None = Field(default="daily", max_length=20)
    launch_date: str | None = Field(default="", max_length=20)
    end_date: str | None = Field(default="", max_length=20)
    target_countries: list[str] = Field(default_factory=lambda: ["KR"], max_length=20)
    landing_url: str | None = Field(default="", max_length=1000)
    image_link: str | None = Field(default="", max_length=1000)
    banned_terms: str | None = Field(default="", max_length=1500)
    required_phrases: str | None = Field(default="", max_length=1500)
    required_exact_phrases: str | None = Field(default="", max_length=1500)
    source_label: str | None = Field(default="기존 작업 파일", max_length=120)


class AdsAdcopyReviewStateRequest(BaseModel):
    advertiser_name: str | None = Field(default="", max_length=120)
    campaign_name: str | None = Field(default="", max_length=160)
    source_label: str | None = Field(default="AI 광고 문안 생성기", max_length=120)
    generated: dict[str, Any] = Field(default_factory=dict)
    validation_report: dict[str, Any] = Field(default_factory=dict)


class AdsAdcopyDraftPlanRequest(BaseModel):
    advertiser_name: str = Field(..., min_length=1, max_length=120)
    generated: dict[str, Any] = Field(default_factory=dict)
    default_max_bid_krw: float | None = Field(default=7000, ge=0)
    location_ids: list[str] = Field(default_factory=list, max_length=50)


class AdsAdcopyDraftExecuteRequest(AdsAdcopyDraftPlanRequest):
    action: str = Field(..., min_length=1, max_length=40)
    state: dict[str, Any] = Field(default_factory=dict)
    confirm: bool = False


PAGE_LABELS = {
    "root": "메인",
    "chat": "광고 Q&A",
    "crawler": "랜딩 URL 검사",
    "favicon": "파비콘 검사",
    "intake": "소재 업로드",
    "slides": "광고주 안내 자료",
    "setupGuide": "캠페인 세팅 가이드",
    "pixelGuide": "픽셀 설치 가이드",
}

GUIDE_DECK_PANELS = {
    "advertiser": "slides-panel",
    "setup": "setup-guide-panel",
    "pixel": "pixel-guide-panel",
}


def _admin_password() -> str:
    return os.getenv("ADMIN_PASSWORD") or "nas2026@"


def _require_admin(request: Request) -> None:
    provided = request.headers.get("x-admin-password", "")
    if provided != _admin_password():
        raise HTTPException(status_code=403, detail="관리자 비밀번호가 올바르지 않습니다.")


def _require_cron(request: Request) -> None:
    cron_secret = os.getenv("CRON_SECRET", "").strip()
    if cron_secret:
        expected = f"Bearer {cron_secret}"
        if request.headers.get("authorization", "") != expected:
            raise HTTPException(status_code=401, detail="cron authorization failed")


def _join_unique_messages(*messages: str) -> str:
    return " · ".join(dict.fromkeys(str(message or "").strip() for message in messages if str(message or "").strip()))


def _query_int(request: Request, key: str, default: int, *, minimum: int = 1, maximum: int = 200) -> int:
    try:
        value = int(str(request.query_params.get(key) or default))
    except (TypeError, ValueError):
        value = default
    return max(minimum, min(maximum, value))


def _date_label(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=KST)
        return parsed.astimezone(KST).date().isoformat()
    except ValueError:
        return text[:10] if re.match(r"^\d{4}-\d{2}-\d{2}", text) else text


def _ads_dashboard_range_cache_key(start_date: str | None, end_date: str | None) -> str:
    if start_date and end_date:
        return f"{ADS_DASHBOARD_CACHE_KEY}:{start_date}:{end_date}"
    return ADS_DASHBOARD_CACHE_KEY


def _ads_dashboard_payload_matches_range(payload: dict[str, Any], start_date: str | None, end_date: str | None) -> bool:
    if not start_date or not end_date:
        return True
    range_payload = payload.get("range") if isinstance(payload.get("range"), dict) else {}
    return str(range_payload.get("start_date") or "") == start_date and str(range_payload.get("end_date") or "") == end_date


def _split_admin_list(value: str | None) -> list[str]:
    items = re.split(r"[\n,]+", str(value or ""))
    return [item.strip() for item in items if item.strip()]


ADCOPY_LANDING_MAX_BYTES = 300_000
ADCOPY_PURCHASE_JOURNEY_STAGES = (
    "문제정의",
    "제품발견",
    "비교검토",
    "단일제품평가",
    "신청전환",
    "사용도움",
)
ADCOPY_REVIEW_HOLD_TOKENS = (
    "의미 중복 후보",
    "경고",
    "광고주 확인 필요",
    "길이 초과",
    "근거 확인 필요",
    "정책 확인 필요",
)
ADCOPY_CONTEXTUAL_HINT_TOKENS = (
    "?",
    "어떻게",
    "어떤",
    "어디",
    "왜",
    "언제",
    "얼마",
    "무엇",
    "뭐가",
    "뭐부터",
    "가능",
    "궁금",
    "고민",
    "걱정",
    "고를",
    "선택",
    "비교",
    "차이",
    "좋을까",
    "할까",
    "때",
    "상황",
    "필요",
    "찾고",
    "찾는",
    "원할",
    "모르겠",
    "막혔",
    "어려",
    "힘들",
    "안 돼",
    "못 ",
    "않",
    "망설",
    "헷갈",
    "지쳐",
    "쌓여",
    "부담",
)
ADCOPY_CONTEXTUAL_HINT_ENDINGS = ("어요", "아요", "네요", "인데요", "더라고요", "거예요", "까요", "나요", "죠")
ADCOPY_CTA_ENDINGS = (
    "확인해보세요",
    "확인하세요",
    "신청하세요",
    "시작해보세요",
    "비교해보세요",
    "알아보세요",
    "체험해보세요",
    "선택해보세요",
    "해보세요",
    "하세요",
)

ADCOPY_TITLE_QUESTION_ENDINGS = (
    "까요",
    "나요",
    "일까요",
    "될까요",
    "할까요",
    "죠",
)

ADCOPY_TITLE_CONDITION_ENDINGS = (
    "다면",
    "라면",
    "때",
)

ADCOPY_COPY_ROLE_TOKENS = {
    "comparison": ("비교", "차이", "보다", "대신", "달리", "나란히", "중에", "어느 쪽"),
    "usage": ("시작", "이용", "사용", "켜면", "따라", "연습", "진행", "받아", "신청"),
    "result": ("달라", "변화", "남아요", "익숙", "자신감", "습관", "이어져", "줄어요", "잡혀요"),
    "benefit": ("수 있어요", "도와", "덜어", "부담", "편하게", "쉽게", "바로", "무료"),
}

ADCOPY_WARNING_SCORE_WEIGHTS = {
    "budget_empty": 1,
    "policy_risk_phrase": 2,
    "context_hint_question_ratio": 2,
    "title_form_repetition": 3,
    "copy_role_diversity": 3,
    "copy_ending_repetition": 3,
    "copy_opening_repetition": 3,
    "copy_repeats_title": 4,
}


def _adcopy_stage_from_adgroup_name(name: str) -> str:
    text = str(name or "").strip()
    if not text:
        return ""
    suffix = re.split(r"[_*]", text)[-1].strip()
    return suffix if suffix in ADCOPY_PURCHASE_JOURNEY_STAGES else ""


def _adcopy_canonical_adgroup_name(value: Any) -> str:
    text = str(value or "").strip()
    if not text or "*" in text:
        return text
    parts = text.rsplit("_", 3)
    if len(parts) == 4 and all(parts) and parts[-1] in ADCOPY_PURCHASE_JOURNEY_STAGES:
        return f"{parts[0]}*{parts[1]}*{parts[2]}_{parts[3]}"
    return text


def _adcopy_is_contextual_hint(value: Any) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    if any(token in text for token in ADCOPY_CONTEXTUAL_HINT_TOKENS):
        return True
    normalized = re.sub(r"[.!?。！？…~\s]+$", "", text)
    return len(normalized) >= 14 and any(normalized.endswith(ending) for ending in ADCOPY_CONTEXTUAL_HINT_ENDINGS)


def _adcopy_normalized_hint(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip().lower())


def _adcopy_core_terms(value: Any) -> list[str]:
    terms: list[str] = []
    for chunk in re.split(r"[/,;\n]+", str(value or "")):
        chunk = re.sub(r"\s+", " ", chunk).strip()
        if len(chunk) >= 2:
            terms.append(chunk)
        for token in re.split(r"\s+", chunk):
            token = token.strip()
            if len(token) >= 2:
                terms.append(token)
    return list(dict.fromkeys(terms))


def _adcopy_is_cta_like(value: Any) -> bool:
    text = re.sub(r"[.!?。！？…~\s]+$", "", str(value or "").strip())
    return bool(text and any(text.endswith(ending) for ending in ADCOPY_CTA_ENDINGS))


def _adcopy_ending_bucket(value: Any) -> str:
    text = re.sub(r"[.!?。！？…~\s]+$", "", str(value or "").strip())
    if not text:
        return ""
    for ending in ADCOPY_CTA_ENDINGS:
        if text.endswith(ending):
            return ending
    return text[-4:]


def _adcopy_title_form(value: Any) -> str:
    text = re.sub(r"[.!?。！？…~\s]+$", "", str(value or "").strip())
    if not text:
        return ""
    if any(text.endswith(ending) for ending in ADCOPY_TITLE_CONDITION_ENDINGS):
        return "condition"
    if "?" in str(value or "") or any(text.endswith(ending) for ending in ADCOPY_TITLE_QUESTION_ENDINGS):
        return "question"
    return "statement"


def _adcopy_opening_bucket(value: Any) -> str:
    tokens = [token for token in re.split(r"\s+", re.sub(r"[^0-9A-Za-z가-힣\s]", " ", str(value or "").strip())) if token]
    return " ".join(tokens[:2]).lower()


def _adcopy_normalized_creative_text(value: Any) -> str:
    return re.sub(r"[^0-9a-z가-힣]", "", str(value or "").strip().lower())


def _adcopy_copy_role(value: Any) -> str:
    text = str(value or "").strip()
    if _adcopy_is_cta_like(text):
        return "cta"
    for role, tokens in ADCOPY_COPY_ROLE_TOKENS.items():
        if any(token in text for token in tokens):
            return role
    return "information"


def _clean_meta_text(value: str | None, max_length: int = 500) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    return text[:max_length]


def _landing_url_parts(url: str) -> tuple[str, str, int | None]:
    parsed = urlparse(str(url or "").strip())
    if parsed.scheme not in {"http", "https"} or not parsed.hostname:
        raise HTTPException(status_code=400, detail="랜딩 URL은 http 또는 https 주소여야 합니다.")
    if parsed.username or parsed.password:
        raise HTTPException(status_code=400, detail="사용자 정보가 포함된 URL은 지원하지 않습니다.")
    return parsed.scheme, parsed.hostname, parsed.port


def _assert_public_ip_address(address: str) -> None:
    try:
        ip = ipaddress.ip_address(address)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="랜딩 URL 호스트를 확인하지 못했습니다.") from exc
    if not ip.is_global:
        raise HTTPException(status_code=400, detail="내부망 또는 비공개 IP로 연결되는 랜딩 URL은 읽을 수 없습니다.")


async def _assert_public_landing_url(url: str) -> None:
    _, host, port = _landing_url_parts(url)
    try:
        literal_ip = ipaddress.ip_address(host.strip("[]"))
    except ValueError:
        literal_ip = None
    if literal_ip is not None:
        if not literal_ip.is_global:
            raise HTTPException(status_code=400, detail="내부망 또는 비공개 IP로 연결되는 랜딩 URL은 읽을 수 없습니다.")
        return
    if host.lower() in {"localhost", "localhost.localdomain"} or host.lower().endswith(".local"):
        raise HTTPException(status_code=400, detail="내부 호스트로 보이는 랜딩 URL은 읽을 수 없습니다.")
    try:
        addr_info = await asyncio.to_thread(socket.getaddrinfo, host, port or 443, type=socket.SOCK_STREAM)
    except socket.gaierror as exc:
        raise HTTPException(status_code=400, detail="랜딩 URL 호스트를 확인하지 못했습니다.") from exc
    for item in addr_info:
        sockaddr = item[4]
        if sockaddr:
            _assert_public_ip_address(str(sockaddr[0]))


async def _fetch_landing_html(url: str) -> tuple[str, str]:
    current_url = str(url or "").strip()
    headers = {
        "User-Agent": "NasmediaOpenAIAdsAdmin/1.0 (+https://openads.admate.ai.kr)",
        "Accept": "text/html,application/xhtml+xml;q=0.9,*/*;q=0.5",
    }
    async with httpx.AsyncClient(timeout=httpx.Timeout(12.0, connect=5.0), follow_redirects=False) as client:
        for _ in range(4):
            await _assert_public_landing_url(current_url)
            async with client.stream("GET", current_url, headers=headers) as response:
                if response.status_code in {301, 302, 303, 307, 308} and response.headers.get("location"):
                    current_url = urljoin(current_url, response.headers["location"])
                    continue
                if response.status_code >= 400:
                    raise HTTPException(status_code=502, detail=f"랜딩 URL 응답 오류: HTTP {response.status_code}")
                content_type = response.headers.get("content-type", "")
                if content_type and "html" not in content_type.lower() and "text/" not in content_type.lower():
                    raise HTTPException(status_code=400, detail="랜딩 URL이 HTML 문서로 응답하지 않았습니다.")
                chunks: list[bytes] = []
                total = 0
                async for chunk in response.aiter_bytes():
                    chunks.append(chunk)
                    total += len(chunk)
                    if total >= ADCOPY_LANDING_MAX_BYTES:
                        break
            raw = b"".join(chunks)[:ADCOPY_LANDING_MAX_BYTES]
            return raw.decode("utf-8", errors="ignore"), str(response.url)
    raise HTTPException(status_code=400, detail="랜딩 URL 리다이렉트가 너무 많습니다.")


def _meta_content(soup: BeautifulSoup, *names: str) -> str:
    for name in names:
        tag = soup.find("meta", attrs={"property": name}) or soup.find("meta", attrs={"name": name})
        if tag and tag.get("content"):
            return _clean_meta_text(tag.get("content"))
    return ""


def _link_rel_contains(value: Any, token: str) -> bool:
    items = value if isinstance(value, list) else str(value or "").split()
    return token.lower() in {str(item).lower() for item in items}


def _inspect_landing_html(html: str, final_url: str) -> dict[str, Any]:
    soup = BeautifulSoup(html or "", "html.parser")
    title = _clean_meta_text(_meta_content(soup, "og:title", "twitter:title") or (soup.title.string if soup.title else ""), 180)
    description = _clean_meta_text(_meta_content(soup, "og:description", "twitter:description", "description"), 600)
    image_url = _clean_meta_text(_meta_content(soup, "og:image", "twitter:image"), 1000)
    if image_url:
        image_url = urljoin(final_url, image_url)
    canonical_tag = soup.find("link", attrs={"rel": lambda value: _link_rel_contains(value, "canonical")})
    canonical_url = urljoin(final_url, canonical_tag.get("href")) if canonical_tag and canonical_tag.get("href") else final_url
    suggestions = [item for item in [title, description] if item]
    selling_points = "\n".join(suggestions)
    return {
        "title": title,
        "description": description,
        "image_url": image_url,
        "canonical_url": canonical_url,
        "selling_points": selling_points[:1200],
    }


def _adcopy_trace(
    source_type: str = "AI 생성",
    source_url: str = "",
    source_excerpt: str = "",
    generation_basis: str = "",
    confidence_score: float = 0.7,
    validation_status: str = "담당자 확인 필요",
) -> dict[str, Any]:
    return {
        "source_type": source_type,
        "source_url": source_url,
        "source_excerpt": source_excerpt,
        "generation_basis": generation_basis,
        "confidence_score": confidence_score,
        "validation_status": validation_status,
        "review_comment": "",
        "exclusion_reason": "",
    }


def _adcopy_response_schema() -> dict[str, Any]:
    trace_schema = {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "source_type": {"type": "string"},
            "source_url": {"type": "string"},
            "source_excerpt": {"type": "string"},
            "generation_basis": {"type": "string"},
            "confidence_score": {"type": "number"},
            "validation_status": {"type": "string"},
            "review_comment": {"type": "string"},
            "exclusion_reason": {"type": "string"},
        },
        "required": [
            "source_type",
            "source_url",
            "source_excerpt",
            "generation_basis",
            "confidence_score",
            "validation_status",
            "review_comment",
            "exclusion_reason",
        ],
    }
    keyword_schema = {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "text": {"type": "string"},
            "origin": {"type": "string", "enum": ["customer_data", "ai_inferred"]},
        },
        "required": ["text", "origin"],
    }
    return {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "policy": {
                "type": "object",
                "additionalProperties": False,
                "properties": {
                    "banned_terms": {"type": "array", "items": {"type": "string"}},
                },
                "required": ["banned_terms"],
            },
            "campaigns": {
                "type": "array",
                "items": {
                    "type": "object",
                    "additionalProperties": False,
                    "properties": {
                        "campaign_name": {"type": "string"},
                        "budget_max": {"type": "number"},
                        "budget_type": {"type": "string"},
                        "launch_date": {"type": "string"},
                        "end_date": {"type": "string"},
                        "objective": {"type": "string"},
                        "target_countries": {"type": "array", "items": {"type": "string"}},
                    },
                    "required": [
                        "campaign_name",
                        "budget_max",
                        "budget_type",
                        "launch_date",
                        "end_date",
                        "objective",
                        "target_countries",
                    ],
                },
            },
            "adgroups": {
                "type": "array",
                "items": {
                    "type": "object",
                    "additionalProperties": False,
                    "properties": {
                        "campaign_name": {"type": "string"},
                        "adgroup_name": {"type": "string"},
                        "keywords": {"type": "array", "items": keyword_schema},
                        "required_phrases": {"type": "array", "items": {"type": "string"}},
                        "required_exact_phrases": {"type": "array", "items": {"type": "string"}},
                        "trace": trace_schema,
                    },
                    "required": [
                        "campaign_name",
                        "adgroup_name",
                        "keywords",
                        "required_phrases",
                        "required_exact_phrases",
                        "trace",
                    ],
                },
            },
            "ads": {
                "type": "array",
                "items": {
                    "type": "object",
                    "additionalProperties": False,
                    "properties": {
                        "ad_name": {"type": "string"},
                        "adgroup_name": {"type": "string"},
                        "title": {"type": "string"},
                        "copy": {"type": "string"},
                        "link": {"type": "string"},
                        "image_link": {"type": "string"},
                        "trace": trace_schema,
                    },
                    "required": ["ad_name", "adgroup_name", "title", "copy", "link", "image_link", "trace"],
                },
            },
        },
        "required": ["policy", "campaigns", "adgroups", "ads"],
    }


def _adcopy_system_prompt() -> str:
    return (
        "You are an expert Korean OpenAI Ads copy strategist. "
        "Generate upload-ready draft data for the exact JSON schema only. "
        "Do not create or activate campaigns. The output is only a human-review draft. "
        "Keep campaigns from the user brief unchanged. Generate only adgroups and ads. "
        "Never include max_bid. Titles must be 24 Korean characters or fewer. "
        "Prefer Korean titles around 16 to 18 characters when possible. "
        "Ad copy must be 48 Korean characters or fewer and should usually be 18 to 42 characters. "
        "Write natural Korean that a Korean ads operator would approve; avoid stiff phrases like '고려한', "
        "generic filler, awkward noun stacks, unverifiable superlatives, guarantees, sensitive targeting language, "
        "and banned terms. Each ad must carry one concrete product benefit or usage situation from the brief, "
        "and ads inside the same adgroup must not repeat the same title/copy structure. Purchase journey stage names "
        "are fixed to 문제정의, 제품발견, 비교검토, 단일제품평가, 신청전환, 사용도움 only. Adgroup names must follow "
        "상품/SKU*타깃*세부의도_구매여정. Context hints must mostly be Korean question or situation sentences, not "
        "plain search keywords, and each keyword origin must be customer_data or ai_inferred."
    )


def _adcopy_user_prompt(payload: AdsAdcopyGenerateRequest) -> str:
    banned_terms = _split_admin_list(payload.banned_terms)
    required_phrases = _split_admin_list(payload.required_phrases)
    required_exact_phrases = _split_admin_list(payload.required_exact_phrases)
    brief = {
        "advertiser_name": payload.advertiser_name.strip(),
        "industry": (payload.industry or "").strip(),
        "campaign": {
            "campaign_name": payload.campaign_name.strip(),
            "budget_max": float(payload.budget_max or 0),
            "budget_type": payload.budget_type.strip() or "daily",
            "launch_date": (payload.launch_date or "").strip(),
            "end_date": (payload.end_date or "").strip(),
            "objective": payload.objective.strip() or "Views",
            "target_countries": payload.target_countries or ["KR"],
        },
        "product_name": payload.product_name.strip(),
        "landing_url": payload.landing_url.strip(),
        "image_link": (payload.image_link or "").strip(),
        "audience": (payload.audience or "").strip(),
        "selling_points": (payload.selling_points or "").strip(),
        "tone": (payload.tone or "").strip(),
        "banned_terms": banned_terms,
        "copy_core_elements": required_phrases,
        "required_exact_phrases": required_exact_phrases,
        "fixed_purchase_journey_stages": list(ADCOPY_PURCHASE_JOURNEY_STAGES),
        "generation_quantity": {
            "adgroups": payload.adgroup_count,
            "ads_per_adgroup": payload.ads_per_adgroup,
        },
    }
    return (
        "Create Korean OpenAI Ads draft copy from this brief.\n"
        "- Return exactly one campaign, copied from brief.campaign.\n"
        "- Create the requested number of adgroups and ads.\n"
        "- Set adgroups.required_phrases to the brief copy_core_elements array. These are core elements to reflect naturally, not verbatim mandatory phrases.\n"
        "- Set adgroups.required_exact_phrases to the brief required_exact_phrases array. These exact phrases must appear verbatim only when provided.\n"
        "- Use landing_url and image_link for every ad.\n"
        "- Make adgroup themes distinct by campaign_name, product/SKU, detailed intent, and one fixed purchase journey stage.\n"
        "- The final part of every adgroup_name must be one of: 문제정의, 제품발견, 비교검토, 단일제품평가, 신청전환, 사용도움.\n"
        "- Format adgroup_name as 상품/SKU*타깃*세부의도_구매여정, for example 영어학습*초등학부모*학습시작고민_문제정의.\n"
        "- Context hints must match only that adgroup's campaign, product/SKU, detailed intent, and purchase journey stage.\n"
        "- Prefer question/situation context hints such as '초등 아이 영어를 집에서 어떻게 시작하면 좋을까?' over plain keywords.\n"
        "- Avoid repeated context hints across adgroups; if a shared idea is needed, make it specific to that adgroup's intent.\n"
        "- Prefer title length 16~18 Korean characters; never exceed 24 characters.\n"
        "- Prefer copy length 18~42 Korean characters; never exceed 48 characters.\n"
        "- Reflect copy_core_elements as benefits, conditions, or product traits without forcing awkward exact wording.\n"
        "- Titles must lead with a user's tension, question, choice, or concrete situation. Avoid feature-list, report-heading, and purely descriptive titles.\n"
        "- Copy must continue the title instead of paraphrasing it, then add a concrete solution direction and a benefit, change, or decision point supported by the brief.\n"
        "- Build each title/copy pair in this order: user tension or curiosity -> solution direction -> benefit or judgment point. Keep the flow natural rather than forcing every clause into one sentence.\n"
        "- Distribute title/copy forms inside each adgroup across question, empathy/condition, information, benefit, comparison, usage, result, and CTA styles. Assign distinct roles before writing; do not reuse the same opening, ending, or sentence skeleton across most ads in a group.\n"
        "- Across the campaign, avoid overusing title frames such as '~다면', '~까요?', and '~죠?' even when each individual title sounds natural.\n"
        "- Do not overuse endings like ~하세요, ~해보세요, ~확인하세요, ~합니다, ~있습니다. Direct CTA endings must be under 30% of ads in the same adgroup.\n"
        "- Keep Korean phrasing crisp, specific, and natural enough to be used after a human review.\n"
        "- Set trace.review_comment and trace.exclusion_reason to empty strings.\n"
        "- Set trace.validation_status to '담당자 확인 필요' unless a concrete warning is needed.\n"
        f"\nBRIEF_JSON:\n{json.dumps(brief, ensure_ascii=False, indent=2)}"
    )


def _extract_openai_response_text(data: dict[str, Any]) -> str:
    if isinstance(data.get("output_text"), str):
        return data["output_text"]
    for item in data.get("output") or []:
        if not isinstance(item, dict):
            continue
        for content in item.get("content") or []:
            if isinstance(content, dict) and isinstance(content.get("text"), str):
                return content["text"]
    return ""


async def _call_openai_adcopy(payload: AdsAdcopyGenerateRequest) -> dict[str, Any]:
    api_key = os.getenv("OPENAI_ADCOPY_API_KEY") or os.getenv("OPENAI_API_KEY")
    if not api_key:
        raise HTTPException(status_code=503, detail="OPENAI_API_KEY 또는 OPENAI_ADCOPY_API_KEY가 설정되어 있지 않습니다.")
    model = os.getenv("OPENAI_ADCOPY_MODEL") or os.getenv("OPENAI_MODEL") or os.getenv("OPENAI_CHAT_MODEL") or "gpt-5"
    base_url = (os.getenv("OPENAI_BASE_URL") or "https://api.openai.com/v1").rstrip("/")
    try:
        timeout_seconds = float(os.getenv("OPENAI_ADCOPY_TIMEOUT_SECONDS") or 55)
    except (TypeError, ValueError):
        timeout_seconds = 55
    timeout_seconds = max(15.0, min(timeout_seconds, 85.0))
    request_body = {
        "model": model,
        "input": [
            {"role": "system", "content": _adcopy_system_prompt()},
            {"role": "user", "content": _adcopy_user_prompt(payload)},
        ],
        "text": {
            "format": {
                "type": "json_schema",
                "name": "openai_ads_adcopy_generated",
                "strict": True,
                "schema": _adcopy_response_schema(),
            }
        },
        "max_output_tokens": 12000,
        "store": False,
    }
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    try:
        async with httpx.AsyncClient(timeout=httpx.Timeout(timeout_seconds, connect=10.0)) as client:
            response = await client.post(f"{base_url}/responses", headers=headers, json=request_body)
    except httpx.TimeoutException as exc:
        raise HTTPException(
            status_code=504,
            detail="AI 문안 생성 API 응답 시간이 초과되었습니다. 생성 수량을 줄이거나 잠시 후 다시 시도해 주세요.",
        ) from exc
    except httpx.RequestError as exc:
        raise HTTPException(
            status_code=502,
            detail=f"AI 문안 생성 API에 연결하지 못했습니다: {exc.__class__.__name__}",
        ) from exc
    if response.status_code >= 400:
        raise HTTPException(status_code=502, detail=f"OpenAI 광고 문안 생성 API 오류: HTTP {response.status_code} · {response.text[:600]}")
    try:
        data = response.json()
    except ValueError as exc:
        raise HTTPException(status_code=502, detail="OpenAI 광고 문안 생성 API가 JSON이 아닌 응답을 반환했습니다.") from exc
    raw_text = _extract_openai_response_text(data)
    if not raw_text.strip():
        raise HTTPException(status_code=502, detail="OpenAI 광고 문안 생성 응답에서 JSON 텍스트를 찾지 못했습니다.")
    try:
        generated = json.loads(raw_text)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=502, detail="OpenAI 광고 문안 생성 응답 JSON 파싱에 실패했습니다.") from exc
    return {"model": model, "raw_response_id": data.get("id"), "generated": generated}


ADCOPY_ENGINE_LABELS = {
    "ai_team_plugin": "AI부서 플러그인",
    "admate": "AdMate 엔진",
}


def _configured_adcopy_default_engine() -> str:
    configured = os.getenv("ADCOPY_DEFAULT_ENGINE", "ai_team_plugin").strip().lower()
    return configured if configured in ADCOPY_ENGINE_LABELS else "ai_team_plugin"


def _adcopy_engine_settings() -> dict[str, Any]:
    configured_default = _configured_adcopy_default_engine()
    engines = {
        "ai_team_plugin": {
            "label": "AI부서 플러그인",
            "available": True,
            "execution_mode": "external_plugin",
            "description": "Claude Code에서 생성한 review.xlsx를 웹 검수 화면으로 가져옵니다.",
        },
        "admate": {
            "label": "AdMate 엔진",
            "available": bool((os.getenv("OPENAI_ADCOPY_API_KEY") or os.getenv("OPENAI_API_KEY") or "").strip()),
            "execution_mode": "server",
            "description": "현재 서비스에서 광고 문안을 직접 생성합니다.",
        },
    }
    return {
        "default_engine": configured_default,
        "effective_default_engine": configured_default,
        "default_fallback_used": False,
        "engines": [
            {
                "id": engine_id,
                **engine,
                "is_configured_default": engine_id == configured_default,
            }
            for engine_id, engine in engines.items()
        ],
    }


def _resolve_adcopy_engine(requested_engine: str) -> str:
    requested = (requested_engine or "admate").strip().lower()
    if requested not in ADCOPY_ENGINE_LABELS:
        raise HTTPException(status_code=400, detail="지원하지 않는 광고 문안 생성 엔진입니다.")
    if requested == "ai_team_plugin":
        raise HTTPException(
            status_code=409,
            detail="AI부서 엔진은 Claude Code 플러그인에서 실행합니다. 생성된 review.xlsx를 관리자 화면에서 불러와 주세요.",
        )
    return requested


def _normalize_generated_adcopy(data: dict[str, Any], payload: AdsAdcopyGenerateRequest) -> dict[str, Any]:
    campaign = {
        "campaign_name": payload.campaign_name.strip(),
        "budget_max": float(payload.budget_max or 0),
        "budget_type": payload.budget_type.strip() or "daily",
        "launch_date": (payload.launch_date or "").strip(),
        "end_date": (payload.end_date or "").strip(),
        "objective": payload.objective.strip() or "Views",
        "target_countries": [country.strip().upper() for country in payload.target_countries if str(country or "").strip()] or ["KR"],
    }
    banned_terms = _split_admin_list(payload.banned_terms)
    model_policy = data.get("policy") if isinstance(data.get("policy"), dict) else {}
    merged_banned_terms = list(dict.fromkeys([*banned_terms, *[str(item).strip() for item in model_policy.get("banned_terms") or [] if str(item).strip()]]))
    required_phrases = _split_admin_list(payload.required_phrases)
    required_exact_phrases = _split_admin_list(payload.required_exact_phrases)
    adgroups: list[dict[str, Any]] = []
    seen_adgroups: set[str] = set()
    for index, item in enumerate(data.get("adgroups") or [], start=1):
        if not isinstance(item, dict):
            continue
        name = _adcopy_canonical_adgroup_name(item.get("adgroup_name") or f"{index:02d}_광고그룹")
        if not name:
            name = f"{index:02d}_광고그룹"
        if name in seen_adgroups:
            name = f"{name}_{index:02d}"
        seen_adgroups.add(name)
        keywords = []
        for keyword in item.get("keywords") or []:
            if isinstance(keyword, dict):
                text = str(keyword.get("text") or "").strip()
                origin = str(keyword.get("origin") or "ai_inferred").strip()
            else:
                text = str(keyword or "").strip()
                origin = "ai_inferred"
            if text:
                keywords.append({"text": text, "origin": origin if origin in {"customer_data", "ai_inferred"} else "ai_inferred"})
        trace = item.get("trace") if isinstance(item.get("trace"), dict) else _adcopy_trace()
        adgroups.append({
            "campaign_name": campaign["campaign_name"],
            "adgroup_name": name,
            "keywords": keywords,
            "required_phrases": [str(value).strip() for value in (item.get("required_phrases") or required_phrases) if str(value).strip()],
            "required_exact_phrases": [
                str(value).strip()
                for value in (item.get("required_exact_phrases") or required_exact_phrases)
                if str(value).strip()
            ],
            "trace": {**_adcopy_trace(), **trace, "review_comment": "", "exclusion_reason": ""},
        })
    valid_adgroup_names = {item["adgroup_name"] for item in adgroups}
    ads: list[dict[str, Any]] = []
    for index, item in enumerate(data.get("ads") or [], start=1):
        if not isinstance(item, dict):
            continue
        adgroup_name = _adcopy_canonical_adgroup_name(item.get("adgroup_name"))
        if adgroup_name not in valid_adgroup_names and adgroups:
            adgroup_name = adgroups[min(index - 1, len(adgroups) - 1)]["adgroup_name"]
        trace = item.get("trace") if isinstance(item.get("trace"), dict) else _adcopy_trace()
        ads.append({
            "ad_name": str(item.get("ad_name") or f"AD_{index:03d}").strip(),
            "adgroup_name": adgroup_name,
            "title": str(item.get("title") or "").strip(),
            "copy": str(item.get("copy") or "").strip(),
            "link": payload.landing_url.strip(),
            "image_link": (payload.image_link or "").strip(),
            "trace": {**_adcopy_trace(), **trace, "review_comment": "", "exclusion_reason": ""},
        })
    return {"policy": {"banned_terms": merged_banned_terms}, "campaigns": [campaign], "adgroups": adgroups, "ads": ads}


ADCOPY_IMPORT_CAMPAIGN_LIST_KEYS = (
    "campaigns",
    "campaigns_검수",
    "campaign_rows",
    "campaign_list",
    "campaign_review",
    "campaign_review_rows",
    "캠페인",
    "캠페인 목록",
)
ADCOPY_IMPORT_CAMPAIGN_OBJECT_KEYS = ("campaign", "campaign_meta", "campaign_info", "캠페인정보", "캠페인 정보")
ADCOPY_IMPORT_ADGROUP_LIST_KEYS = (
    "adgroups",
    "adgroups_검수",
    "ad_groups",
    "adgroup_rows",
    "ad_group_rows",
    "adgroup_review",
    "adgroup_review_rows",
    "groups",
    "광고그룹",
    "광고그룹 목록",
)
ADCOPY_IMPORT_AD_LIST_KEYS = (
    "ads",
    "ads_검수",
    "ad_rows",
    "creatives",
    "creative_rows",
    "ads_review",
    "ads_review_rows",
    "copies",
    "copy_rows",
    "소재",
    "소재 목록",
    "카피",
    "카피 목록",
)


def _adcopy_import_text(source: dict[str, Any], keys: tuple[str, ...] | list[str], fallback: str = "") -> str:
    if not isinstance(source, dict):
        return fallback
    for key in keys:
        if key not in source:
            continue
        value = source.get(key)
        if value is None:
            continue
        if isinstance(value, str):
            text = value.strip()
        elif isinstance(value, (int, float)) and not isinstance(value, bool):
            text = str(value).strip()
        else:
            text = ""
        if text:
            return text
    return fallback


def _adcopy_import_first_value(source: dict[str, Any], keys: tuple[str, ...] | list[str]) -> Any:
    if not isinstance(source, dict):
        return None
    for key in keys:
        value = source.get(key)
        if value not in (None, ""):
            return value
    return None


def _adcopy_import_number(value: Any, fallback: float = 0) -> float:
    if value in (None, ""):
        return fallback
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = re.sub(r"[^\d.\-]", "", str(value))
    if not text:
        return fallback
    try:
        return float(text)
    except ValueError:
        return fallback


def _adcopy_try_json(value: str) -> Any:
    text = str(value or "").strip()
    if not text or text[0] not in "[{":
        return None
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return None


def _adcopy_import_items(value: Any) -> list[Any]:
    if value in (None, ""):
        return []
    if isinstance(value, list):
        return value
    if isinstance(value, dict):
        if any(key in value for key in ("text", "keyword", "name", "키워드")):
            return [value]
        return list(value.values())
    if isinstance(value, str):
        parsed = _adcopy_try_json(value)
        if parsed is not None:
            return _adcopy_import_items(parsed)
    return [item.strip() for item in re.split(r"[\n,;|]+", str(value)) if item.strip()]


def _adcopy_rows_from_value(value: Any) -> list[dict[str, Any]]:
    if isinstance(value, str):
        parsed = _adcopy_try_json(value)
        if parsed is not None:
            value = parsed
    rows: list[dict[str, Any]] = []
    if isinstance(value, list):
        for item in value:
            if isinstance(item, dict):
                rows.append(dict(item))
        return rows
    if isinstance(value, dict):
        nested_rows = value.get("rows")
        if isinstance(nested_rows, list):
            return _adcopy_rows_from_value(nested_rows)
        if any(isinstance(item, dict) for item in value.values()):
            for row_key, item in value.items():
                if not isinstance(item, dict):
                    continue
                row = dict(item)
                row.setdefault("name", row_key)
                rows.append(row)
            return rows
        return [dict(value)] if value else []
    return []


def _adcopy_import_source_containers(source: dict[str, Any]) -> list[dict[str, Any]]:
    containers: list[dict[str, Any]] = [source]
    for key in ("generated", "result", "data", "review", "workbook"):
        value = source.get(key)
        if isinstance(value, dict):
            containers.append(value)
    return containers


def _adcopy_import_sheet_rows(source: dict[str, Any], keys: tuple[str, ...]) -> list[dict[str, Any]]:
    sheets = source.get("sheets")
    if not isinstance(sheets, list):
        return []
    normalized_keys = {str(key).strip().lower() for key in keys}
    for sheet in sheets:
        if not isinstance(sheet, dict):
            continue
        name = str(sheet.get("name") or sheet.get("sheet") or "").strip().lower()
        if name in normalized_keys:
            rows = _adcopy_rows_from_value(sheet.get("rows"))
            if rows:
                return rows
    return []


def _adcopy_import_rows(source: dict[str, Any], keys: tuple[str, ...]) -> list[dict[str, Any]]:
    if not isinstance(source, dict):
        return []
    for container in _adcopy_import_source_containers(source):
        sheet_rows = _adcopy_import_sheet_rows(container, keys)
        if sheet_rows:
            return sheet_rows
        for key in keys:
            rows = _adcopy_rows_from_value(container.get(key))
            if rows:
                return rows
    return []


def _adcopy_is_instruction_workbook_row(headers: list[str], values: list[str]) -> bool:
    cells = [str(value or "").strip() for value in values if str(value or "").strip()]
    if not cells:
        return True
    lowered = [cell.lower() for cell in cells]
    if all(cell in {"required", "optional"} for cell in lowered):
        return True
    first = lowered[0]
    instruction_prefixes = (
        "how we will",
        "maximum you will",
        "when this campaign",
        "when this",
        "lifetime or daily",
        "views or clicks",
        "list of countries",
    )
    if any(first.startswith(prefix) for prefix in instruction_prefixes):
        return True
    header_set = {str(header or "").strip().lower() for header in headers}
    if {"campaign_name", "budget_max", "budget_type"}.issubset(header_set) and first in {"required", "how we will identify each campaign"}:
        return True
    return False


def _adcopy_workbook_dump(content: bytes, filename: str) -> dict[str, Any]:
    if content.startswith(b"SCDSA"):
        raise HTTPException(
            status_code=400,
            detail=(
                "Microsoft 365 민감도 레이블 또는 조직 보안으로 보호된 파일은 서버에서 읽을 수 없습니다. "
                "Excel에서 보호를 해제한 표준 .xlsx 복사본으로 저장한 뒤 다시 업로드해 주세요."
            ),
        )
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(content), data_only=True, read_only=False)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"워크북을 읽을 수 없습니다: {exc}") from exc

    sheets: list[dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            headers: list[str] = []
            rows: list[dict[str, Any]] = []
            preamble: list[str] = []
            for raw_row in sheet.iter_rows(values_only=True):
                values = ["" if value is None else str(value).strip() for value in raw_row]
                while values and values[-1] == "":
                    values.pop()
                if not values or not any(values):
                    continue
                if not headers:
                    non_empty = [value for value in values if value]
                    if len(non_empty) == 1:
                        preamble.append(non_empty[0])
                        continue
                    headers = [value or f"COL_{index + 1}" for index, value in enumerate(values)]
                    continue
                row: dict[str, Any] = {}
                for index, header in enumerate(headers):
                    row[header] = values[index] if index < len(values) else ""
                if _adcopy_is_instruction_workbook_row(headers, values):
                    continue
                if any(str(value or "").strip() for value in row.values()):
                    rows.append(row)
            sheets.append({"name": sheet.title, "preamble": preamble, "headers": headers, "rows": rows})
    finally:
        workbook.close()
    return {"file": filename, "sheets": sheets}


def _adcopy_sample_workbook_bytes() -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    campaign_sheet = workbook.active
    campaign_sheet.title = "campaigns_검수"
    campaign_sheet.append(["OpenAI Ads AI 광고 문안 생성기 샘플 워크북"])
    campaign_sheet.append(["campaign_name", "advertiser_name", "budget_max", "budget_type", "launch_date", "end_date", "objective", "target_countries"])
    campaign_sheet.append(["샘플_초안_캠페인", "캐츠잉글리시", "150000", "total", "2026-07-20", "2026-08-02", "Views", "KR"])

    adgroup_sheet = workbook.create_sheet("adgroups_검수")
    adgroup_sheet.append(["광고그룹 검수 목록"])
    adgroup_sheet.append(["adgroup_name", "keywords", "카피 핵심 반영 요소", "필수 고정 문구", "검수상태"])
    adgroup_sheet.append([
        "영어학습*초등학부모*학습루틴고민_문제정의",
        "초등 아이 영어를 집에서 어떻게 시작하면 좋을까?; 파닉스와 문법을 집에서 같이 챙길 수 있을까?; 영어 학습을 매일 이어가려면 어떤 루틴이 좋을까?; 학원 없이 아이 영어 습관을 만들 수 있을까?; 집 공부가 자꾸 끊길 때 무엇부터 바꾸면 좋을까?",
        "학습 루틴",
        "",
        "확인 완료 검수",
    ])
    adgroup_sheet.append([
        "영어학습*초등학부모*레벨테스트검토_단일제품평가",
        "아이 영어 수준을 먼저 확인하려면 무엇부터 보면 좋을까?; 레벨테스트 결과로 어떤 학습 방향을 알 수 있을까?; 초등 영어 레벨테스트는 집에서 진행해도 괜찮을까?; 수준별 학습 리포트는 어떤 기준으로 봐야 할까?; 학습 상담 전에 아이 수준을 어떻게 점검하면 좋을까?",
        "레벨테스트, 학습 리포트",
        "",
        "확인 완료 검수",
    ])

    ads_sheet = workbook.create_sheet("ads_검수")
    ads_sheet.append(["소재 검수 목록"])
    ads_sheet.append(["ad_name", "adgroup_name", "title", "copy", "link", "image_link", "검수상태"])
    ads_sheet.append([
        "AD_001",
        "영어학습*초등학부모*학습루틴고민_문제정의",
        "집에서도 이어지는 초등 영어 루틴",
        "학습 루틴과 반복 훈련 흐름 안내",
        "https://example.com/cats/landing",
        "https://example.com/cats/image.png",
        "확인 완료 검수",
    ])
    ads_sheet.append([
        "AD_002",
        "영어학습*초등학부모*학습루틴고민_문제정의",
        "보호자가 보기 쉬운 학습 리포트",
        "보호자 관리에 맞춘 반복 학습 루틴",
        "https://example.com/cats/landing",
        "https://example.com/cats/image-2.png",
        "확인 완료 검수",
    ])
    ads_sheet.append([
        "AD_003",
        "영어학습*초등학부모*레벨테스트검토_단일제품평가",
        "레벨테스트 후 학습 방향 안내",
        "수준별 학습 방향을 리포트로 확인",
        "https://example.com/cats/landing",
        "https://example.com/cats/image.png",
        "확인 완료 검수",
    ])

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _adcopy_brief_sample_workbook_bytes() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    campaign_sheet = workbook.active
    campaign_sheet.title = "campaigns"
    campaign_sheet.append([
        "campaign_name", "advertiser_name", "industry", "budget_max", "budget_type",
        "launch_date", "end_date", "objective", "target_countries", "adgroup_count", "ads_per_adgroup",
    ])
    campaign_sheet.append([
        "SAMPLE_영어학습", "캐츠잉글리시", "교육", 50000, "daily",
        "2026-07-22", "2026-08-31", "Views", "KR", 3, 3,
    ])

    brief_sheet = workbook.create_sheet("상품.브리프")
    brief_sheet.append([
        "campaign_name", "상품·SKU명", "대표 랜딩 URL", "이미지 URL", "타깃 고객",
        "핵심 문제·니즈", "상품 특징·혜택", "카피 핵심 반영 요소", "톤앤매너", "금지 표현", "필수 고정 문구",
    ])
    brief_sheet.append([
        "SAMPLE_영어학습", "초등 영어 학습 서비스", "https://www.catsenglish.net", "https://example.com/image.png",
        "초등 자녀의 영어 학습을 고민하는 학부모", "아이에게 맞는 학습 시작점을 찾기 어려움",
        "집에서 이용하는 온라인 영어 학습 콘텐츠", "영어 학습 시작; 아이에게 맞는 학습 확인; 먼저 경험 후 결정",
        "담백하고 신뢰감 있는 교육 정보형", "무조건; 1위; 최고", "",
    ])

    header_fill = PatternFill("solid", fgColor="EAF0F8")
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="1E293B")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_index, column_cells in enumerate(sheet.columns, start=1):
            longest = max((len(str(cell.value or "")) for cell in column_cells), default=0)
            sheet.column_dimensions[get_column_letter(column_index)].width = min(max(longest + 2, 14), 38)

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _adcopy_brief_from_workbook_dump(workbook_dump: dict[str, Any]) -> dict[str, Any]:
    sheets = workbook_dump.get("sheets") if isinstance(workbook_dump.get("sheets"), list) else []

    def sheet_rows(*tokens: str) -> list[dict[str, Any]]:
        for sheet in sheets:
            if not isinstance(sheet, dict):
                continue
            name = str(sheet.get("name") or "").strip().lower()
            if any(token in name for token in tokens):
                rows = sheet.get("rows") if isinstance(sheet.get("rows"), list) else []
                return [row for row in rows if isinstance(row, dict)]
        return []

    campaign_row = (sheet_rows("campaign", "캠페인") or [{}])[0]
    brief_row = (sheet_rows("brief", "브리프", "상품") or [{}])[0]
    source = {**campaign_row, **brief_row}

    def text(*keys: str, fallback: str = "") -> str:
        return _adcopy_import_text(source, keys, fallback)

    audience_parts = [
        text("audience", "target_customer", "target_audience", "타깃 고객", "타겟 고객", "타깃", "타겟"),
        text("problem_need", "core_need", "핵심 문제·니즈", "핵심 문제", "문제·니즈", "니즈"),
    ]
    target_countries = [
        str(item).strip().upper()
        for item in _adcopy_import_items(
            _adcopy_import_first_value(source, ("target_countries", "countries", "target_country", "타겟 국가", "타깃 국가", "국가"))
        )
        if str(item).strip()
    ] or ["KR"]
    budget_type = text("budget_type", "budgetType", "예산유형", "예산 유형", fallback="daily")
    if budget_type in {"일 예산", "일예산"}:
        budget_type = "daily"
    elif budget_type in {"총 예산", "총예산"}:
        budget_type = "total"

    brief = {
        "advertiser_name": text("advertiser_name", "advertiser", "광고주명", "광고주"),
        "industry": text("industry", "category", "업종", "산업"),
        "campaign_name": text("campaign_name", "campaign", "캠페인명", "캠페인"),
        "objective": _adcopy_import_objective(text("objective", "goal", "목표", fallback="Views")),
        "budget_max": _adcopy_import_number(
            _adcopy_import_first_value(source, ("budget_max", "budget", "예산", "일예산", "총예산")),
            0,
        ),
        "budget_type": budget_type or "daily",
        "launch_date": text("launch_date", "start_date", "시작일"),
        "end_date": text("end_date", "종료일"),
        "target_countries": target_countries,
        "product_name": text(
            "product_name", "product", "service_name", "sku", "제품·서비스명", "상품·SKU명", "상품/SKU명", "상품명", "서비스명",
        ),
        "landing_url": text("landing_url", "representative_landing_url", "url", "대표 랜딩 URL", "랜딩 URL", "랜딩URL"),
        "image_link": text("image_link", "image_url", "이미지 URL", "이미지URL"),
        "audience": "\n".join(dict.fromkeys(part for part in audience_parts if part)),
        "selling_points": text("selling_points", "features_benefits", "상품 특징·혜택", "상품 특장점", "상품 특징", "혜택"),
        "tone": text("tone", "tone_and_manner", "톤앤매너", "톤앤 매너"),
        "banned_terms": text("banned_terms", "금지 표현", "금지어"),
        "required_phrases": text("required_phrases", "copy_core_elements", "카피 핵심 반영 요소", "핵심 반영 요소"),
        "required_exact_phrases": text("required_exact_phrases", "필수 고정 문구", "고정 문구", "필수 고지"),
        "adgroup_count": int(_adcopy_import_number(_adcopy_import_first_value(source, ("adgroup_count", "광고그룹 수", "광고그룹수")), 3)),
        "ads_per_adgroup": int(_adcopy_import_number(_adcopy_import_first_value(source, ("ads_per_adgroup", "소재/그룹", "그룹별 소재 수")), 3)),
    }
    required = {
        "advertiser_name": "광고주명",
        "campaign_name": "캠페인명",
        "product_name": "제품·서비스명",
        "landing_url": "랜딩 URL",
        "image_link": "이미지 URL",
    }
    missing_fields = [label for key, label in required.items() if not str(brief.get(key) or "").strip()]
    return {"brief": brief, "missing_fields": missing_fields, "ready": not missing_fields}


def _adcopy_review_workbook_bytes(generated: dict[str, Any], validation_report: dict[str, Any] | None = None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    report = validation_report if isinstance(validation_report, dict) else {}
    checks_by_ad = {
        str(item.get("ad_name") or ""): item
        for item in report.get("creative_checks", [])
        if isinstance(item, dict)
    }
    workbook = Workbook()
    campaign_sheet = workbook.active
    campaign_sheet.title = "campaigns_검수"
    campaign_headers = [
        "campaign_name", "advertiser_name", "budget_max", "budget_type", "launch_date",
        "end_date", "objective", "target_countries",
    ]
    campaign_sheet.append(["광고 문안 검토 작업본 · 서버 저장과 별개인 다운로드 파일입니다."])
    campaign_sheet.append(campaign_headers)
    for campaign in generated.get("campaigns", []):
        if not isinstance(campaign, dict):
            continue
        campaign_sheet.append([
            campaign.get("campaign_name", ""),
            campaign.get("advertiser_name", ""),
            campaign.get("budget_max", ""),
            campaign.get("budget_type", ""),
            campaign.get("launch_date", ""),
            campaign.get("end_date", ""),
            campaign.get("objective", ""),
            json.dumps(campaign.get("target_countries", []), ensure_ascii=False),
        ])

    adgroup_sheet = workbook.create_sheet("adgroups_검수")
    adgroup_headers = [
        "campaign_name", "adgroup_name", "max_bid", "keywords", "required_phrases",
        "required_exact_phrases", "validation_status", "검수상태", "검수 메모", "제외 사유",
    ]
    adgroup_sheet.append(["광고그룹 검토 작업본"])
    adgroup_sheet.append(adgroup_headers)
    for group in generated.get("adgroups", []):
        if not isinstance(group, dict):
            continue
        trace = group.get("trace") if isinstance(group.get("trace"), dict) else {}
        adgroup_sheet.append([
            group.get("campaign_name", ""),
            group.get("adgroup_name", ""),
            group.get("max_bid", ""),
            json.dumps(group.get("keywords", []), ensure_ascii=False),
            json.dumps(group.get("required_phrases", []), ensure_ascii=False),
            json.dumps(group.get("required_exact_phrases", []), ensure_ascii=False),
            trace.get("validation_status", ""),
            trace.get("review_status", "담당자 확인 필요"),
            trace.get("review_comment", ""),
            trace.get("exclusion_reason", ""),
        ])

    ads_sheet = workbook.create_sheet("ads_검수")
    ads_headers = [
        "ad_name", "adgroup_name", "title", "copy", "link", "image_link",
        "validation_status", "플랫폼 자동 점검", "확인 필요 항목", "검수상태", "검수 메모", "제외 사유",
    ]
    ads_sheet.append(["소재 검토 작업본 · 자동 점검은 참고 정보이며 검수상태는 담당자가 결정합니다."])
    ads_sheet.append(ads_headers)
    for ad in generated.get("ads", []):
        if not isinstance(ad, dict):
            continue
        trace = ad.get("trace") if isinstance(ad.get("trace"), dict) else {}
        check = checks_by_ad.get(str(ad.get("ad_name") or ""), {})
        check_status = str(check.get("status") or "pass")
        check_label = {"error": "오류", "warning": "주의", "pass": "양호"}.get(check_status, check_status)
        issues = " · ".join(
            str(item.get("message") or "").strip()
            for item in check.get("issues", [])
            if isinstance(item, dict) and str(item.get("message") or "").strip()
        )
        ads_sheet.append([
            ad.get("ad_name", ""),
            ad.get("adgroup_name", ""),
            ad.get("title", ""),
            ad.get("copy", ""),
            ad.get("link", ""),
            ad.get("image_link", ""),
            trace.get("validation_status", ""),
            check_label,
            issues,
            trace.get("review_status", "담당자 확인 필요"),
            trace.get("review_comment", ""),
            trace.get("exclusion_reason", ""),
        ])

    status_validation = DataValidation(
        type="list",
        formula1='"담당자 확인 필요,광고주 확인 필요,승인,수정 필요,제외"',
        allow_blank=False,
    )
    ads_sheet.add_data_validation(status_validation)
    if ads_sheet.max_row >= 3:
        status_validation.add(f"J3:J{ads_sheet.max_row}")

    header_fill = PatternFill("solid", fgColor="EAF0F8")
    note_fill = PatternFill("solid", fgColor="FFF4F4")
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A3"
        sheet.auto_filter.ref = f"A2:{sheet.cell(2, sheet.max_column).coordinate}"
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=sheet.max_column)
        note_cell = sheet.cell(1, 1)
        note_cell.fill = note_fill
        note_cell.font = Font(bold=True, color="9F1239")
        note_cell.alignment = Alignment(vertical="center")
        for cell in sheet[2]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="1E293B")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in sheet.iter_rows(min_row=3):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.lstrip().startswith(("=", "+", "-", "@")):
                    cell.value = f"'{cell.value}"
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_index, column_cells in enumerate(sheet.columns, start=1):
            letter = get_column_letter(column_index)
            longest = max((len(str(cell.value or "")) for cell in column_cells[:40]), default=0)
            sheet.column_dimensions[letter].width = min(max(longest + 2, 12), 42)

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _adcopy_import_campaign_rows(source: dict[str, Any]) -> list[dict[str, Any]]:
    rows = _adcopy_import_rows(source, ADCOPY_IMPORT_CAMPAIGN_LIST_KEYS)
    if rows:
        return rows
    for key in ADCOPY_IMPORT_CAMPAIGN_OBJECT_KEYS:
        value = source.get(key) if isinstance(source, dict) else None
        if isinstance(value, dict):
            return [dict(value)]
    return []


def _adcopy_import_nested_ads(group_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for group in group_rows:
        adgroup_name = _adcopy_import_text(
            group,
            ("adgroup_name", "ad_group_name", "adgroup", "ad_group", "name", "광고그룹명", "광고그룹"),
        )
        for key in ADCOPY_IMPORT_AD_LIST_KEYS:
            for item in _adcopy_import_items(group.get(key)):
                if not isinstance(item, dict):
                    continue
                row = dict(item)
                if adgroup_name:
                    row.setdefault("adgroup_name", adgroup_name)
                    row.setdefault("광고그룹명", adgroup_name)
                rows.append(row)
    return rows


def _adcopy_import_trace(raw: Any, status_source: Any = None) -> dict[str, Any]:
    source = raw if isinstance(raw, dict) else {}
    nested_trace = source.get("trace") if isinstance(source.get("trace"), dict) else {}
    trace = nested_trace if isinstance(nested_trace, dict) else {}
    ai_validation_status = str(
        source.get("validation_status")
        or source.get("자동 판정")
        or source.get("검수 결과")
        or source.get("검수결과")
        or trace.get("validation_status")
        or ""
    ).strip()
    explicit_review_keys = ("human_check", "review_status", "검수상태", "휴먼 체크", "검수")
    has_explicit_review_field = any(key in source for key in explicit_review_keys) or any(key in trace for key in explicit_review_keys)
    review_status_candidates = [
        status_source,
        trace.get("review_status"),
        trace.get("검수상태"),
        trace.get("human_check"),
        trace.get("휴먼 체크"),
    ]
    if not has_explicit_review_field:
        review_status_candidates.extend([source.get("status"), ai_validation_status, trace.get("status")])
    status_text_raw = " ".join(str(item or "").strip() for item in review_status_candidates if str(item or "").strip())
    status_text = status_text_raw.lower()
    review_status = "담당자 확인 필요"
    if any(token in status_text for token in ("사용 불가", "사용불가", "불가", "제외", "excluded", "exclude", "drop", "reject")):
        review_status = "제외"
    elif any(token.lower() in status_text for token in ADCOPY_REVIEW_HOLD_TOKENS):
        review_status = "광고주 확인 필요"
    elif any(token in status_text for token in ("무수정 승인", "수정 후 승인", "확인 완료", "승인", "approved", "approve", "ok", "pass", "통과")):
        review_status = "승인"
    elif any(token in status_text for token in ("부분 재생성", "확인 필요", "수정", "보류", "재생성", "revise", "hold", "needs")):
        review_status = "수정 필요"
    validation_status = ai_validation_status if has_explicit_review_field and ai_validation_status else review_status
    review_comment = str(
        source.get("review_comment")
        or source.get("검수 메모")
        or source.get("review_note")
        or source.get("운영 메모")
        or trace.get("review_comment")
        or trace.get("memo")
        or trace.get("검수 메모")
        or trace.get("review_note")
        or trace.get("운영 메모")
        or ""
    ).strip()
    if review_status == "광고주 확인 필요" and not review_comment and status_text_raw:
        review_comment = status_text_raw[:500]
    return {
        **_adcopy_trace(source_type="기존 작업 파일", generation_basis="기존 작업 파일 변환", confidence_score=0.62),
        **trace,
        "source_type": str(source.get("source_type") or trace.get("source_type") or "기존 작업 파일").strip(),
        "source_url": str(source.get("source_url") or trace.get("source_url") or "").strip(),
        "source_excerpt": str(source.get("source_excerpt") or trace.get("source_excerpt") or "").strip(),
        "generation_basis": str(source.get("generation_basis") or trace.get("generation_basis") or "기존 작업 파일 변환").strip(),
        "confidence_score": _safe_float(source.get("confidence_score") or trace.get("confidence_score") or 0.62),
        "validation_status": validation_status,
        "review_status": review_status,
        "review_comment": review_comment,
        "exclusion_reason": str(source.get("exclusion_reason") or source.get("제외 사유") or trace.get("exclusion_reason") or trace.get("제외 사유") or "").strip(),
    }


def _adcopy_import_objective(value: Any, fallback: str = "Views") -> str:
    text = str(value or "").strip()
    normalized = text.lower()
    if normalized in {"click", "clicks", "클릭"}:
        return "Clicks"
    if normalized in {"conversion", "conversions", "전환"}:
        return "Conversions"
    if normalized in {"view", "views", "reach", "impression", "impressions", "노출"}:
        return "Views"
    return text or fallback or "Views"


def _normalize_external_generated_adcopy(payload: AdsAdcopyImportRequest) -> dict[str, Any]:
    source = payload.generated if isinstance(payload.generated, dict) else {}
    if not source:
        raise HTTPException(status_code=400, detail="불러올 광고 작업 파일 내용을 입력해 주세요.")

    campaign_rows = _adcopy_import_campaign_rows(source)
    group_rows = _adcopy_import_rows(source, ADCOPY_IMPORT_ADGROUP_LIST_KEYS)
    ad_rows = [
        *_adcopy_import_rows(source, ADCOPY_IMPORT_AD_LIST_KEYS),
        *_adcopy_import_nested_ads(group_rows),
    ]
    first_campaign = campaign_rows[0] if campaign_rows else {}
    first_group = group_rows[0] if group_rows else {}
    first_ad = ad_rows[0] if ad_rows else {}
    source_label = str(payload.source_label or "기존 작업 파일").strip()

    fallback_campaign_name = (
        str(payload.campaign_name or "").strip()
        or _adcopy_import_text(first_group, ("campaign_name", "campaign", "캠페인명", "캠페인"))
        or _adcopy_import_text(first_ad, ("campaign_name", "campaign", "캠페인명", "캠페인"))
        or "외부_캠페인"
    )

    def build_campaign(row: dict[str, Any], index: int) -> dict[str, Any]:
        name = _adcopy_import_text(
            row,
            ("campaign_name", "campaign", "name", "Campaign", "캠페인명", "캠페인"),
            fallback_campaign_name if index == 1 else f"{fallback_campaign_name}_{index:02d}",
        )
        budget_value = _adcopy_import_first_value(row, ("budget_max", "budget", "daily_budget", "total_budget", "예산", "일예산", "총예산"))
        budget_max = _adcopy_import_number(budget_value, float(payload.budget_max or 0))
        budget_type = _adcopy_import_text(row, ("budget_type", "budgetType", "예산유형", "예산 유형"), str(payload.budget_type or "daily").strip() or "daily")
        if budget_type in {"일 예산", "일예산"}:
            budget_type = "daily"
        elif budget_type in {"총 예산", "총예산"}:
            budget_type = "total"
        target_countries = [
            str(item).strip().upper()
            for item in _adcopy_import_items(_adcopy_import_first_value(row, ("target_countries", "countries", "국가", "타겟 국가")) or payload.target_countries)
            if str(item).strip()
        ] or ["KR"]
        return {
            "campaign_name": name,
            "budget_max": budget_max,
            "budget_type": budget_type or "daily",
            "launch_date": _adcopy_import_text(row, ("launch_date", "start_date", "start", "시작일"), str(payload.launch_date or "").strip()),
            "end_date": _adcopy_import_text(row, ("end_date", "end", "종료일"), str(payload.end_date or "").strip()),
            "objective": _adcopy_import_objective(
                _adcopy_import_text(row, ("objective", "goal", "campaign_objective", "목표", "유형"), str(payload.objective or "Views").strip()),
                str(payload.objective or "Views").strip(),
            ),
            "target_countries": target_countries,
        }

    derived_campaign_rows: list[dict[str, Any]] = []
    if not campaign_rows:
        seen_campaign_names: set[str] = set()
        for group_row in group_rows:
            group_campaign_name = _adcopy_import_text(
                group_row,
                ("campaign_name", "campaign", "Campaign", "캠페인명", "캠페인"),
            )
            if group_campaign_name and group_campaign_name not in seen_campaign_names:
                seen_campaign_names.add(group_campaign_name)
                derived_campaign_rows.append({"campaign_name": group_campaign_name})
    banned_terms = list(dict.fromkeys([
        *_split_admin_list(payload.banned_terms),
        *[
            str(item).strip()
            for item in _adcopy_import_items((source.get("policy") if isinstance(source.get("policy"), dict) else {}).get("banned_terms"))
            if str(item).strip()
        ],
        *[
            str(item).strip()
            for item in _adcopy_import_items(_adcopy_import_first_value(source, ("banned_terms", "금지 표현", "금지어")))
            if str(item).strip()
        ],
    ]))
    default_required_phrases = _split_admin_list(payload.required_phrases)
    default_required_exact_phrases = _split_admin_list(payload.required_exact_phrases)
    default_link = str(payload.landing_url or "").strip()
    default_image = str(payload.image_link or "").strip()

    effective_campaign_rows = campaign_rows or derived_campaign_rows
    campaigns = [build_campaign(row, index) for index, row in enumerate(effective_campaign_rows, start=1)] or [build_campaign(first_campaign, 1)]
    campaign = campaigns[0]
    campaign_name = campaign["campaign_name"]
    campaign_names = {item["campaign_name"] for item in campaigns}

    adgroups: list[dict[str, Any]] = []
    seen_adgroups: set[str] = set()
    for index, item in enumerate(group_rows, start=1):
        name = _adcopy_canonical_adgroup_name(
            _adcopy_import_text(item, ("adgroup_name", "ad_group_name", "adgroup", "ad_group", "name", "광고그룹명", "광고그룹"), f"{index:02d}_광고그룹")
        )
        group_campaign_name = _adcopy_import_text(
            item,
            ("campaign_name", "campaign", "Campaign", "캠페인명", "캠페인"),
            campaign_name,
        )
        if group_campaign_name not in campaign_names:
            group_campaign_name = campaign_name
        if name in seen_adgroups:
            name = f"{name}_{index:02d}"
        seen_adgroups.add(name)
        keywords = []
        raw_keywords = _adcopy_import_first_value(item, ("keywords", "context_hints", "expanded_keywords", "키워드", "확장 키워드", "컨텍스트 힌트", "Context Hints"))
        keyword_origins = [
            str(origin).strip()
            for origin in _adcopy_import_items(_adcopy_import_first_value(item, ("keywords_origin", "keyword_origins", "키워드 출처", "출처")))
            if str(origin).strip()
        ]
        for keyword_index, keyword in enumerate(_adcopy_import_items(raw_keywords)):
            if isinstance(keyword, dict):
                text = _adcopy_import_text(keyword, ("text", "keyword", "name", "키워드"))
                origin = _adcopy_import_text(keyword, ("origin", "source", "출처"), "ai_inferred")
            else:
                text = str(keyword or "").strip()
                origin = keyword_origins[keyword_index] if keyword_index < len(keyword_origins) else ("customer_data" if source_label else "ai_inferred")
            if text:
                keywords.append({"text": text, "origin": origin if origin in {"customer_data", "ai_inferred"} else "ai_inferred"})
        required = [
            str(value).strip()
            for value in _adcopy_import_items(
                _adcopy_import_first_value(
                    item,
                    (
                        "required_phrases",
                        "copy_core_elements",
                        "core_elements",
                        "required_terms",
                        "카피 핵심 반영 요소",
                        "핵심 반영 요소",
                        "필수 포함 문구",
                        "필수 문구",
                    ),
                )
                or default_required_phrases
            )
            if str(value).strip()
        ]
        required_exact = [
            str(value).strip()
            for value in _adcopy_import_items(
                _adcopy_import_first_value(
                    item,
                    ("required_exact_phrases", "exact_required_phrases", "필수 고정 문구", "고정 문구", "필수 고지", "필수 고지·주의사항"),
                )
                or default_required_exact_phrases
            )
            if str(value).strip()
        ]
        adgroups.append(
            {
                "campaign_name": group_campaign_name,
                "adgroup_name": name,
                "keywords": keywords,
                "required_phrases": list(dict.fromkeys(required)),
                "required_exact_phrases": list(dict.fromkeys(required_exact)),
                "trace": _adcopy_import_trace(item, _adcopy_import_first_value(item, ("human_check", "review_status", "검수상태", "휴먼 체크", "검수"))),
            }
        )

    valid_group_names = {item["adgroup_name"] for item in adgroups}
    inferred_group_names: list[str] = []
    for item in ad_rows:
        name = _adcopy_canonical_adgroup_name(
            _adcopy_import_text(item, ("adgroup_name", "ad_group_name", "adgroup", "ad_group", "광고그룹명", "광고그룹"))
        )
        if name and name not in valid_group_names and name not in inferred_group_names:
            inferred_group_names.append(name)
    for name in inferred_group_names:
        adgroups.append(
            {
                "campaign_name": campaign_name,
                "adgroup_name": name,
                "keywords": [],
                "required_phrases": default_required_phrases,
                "required_exact_phrases": default_required_exact_phrases,
                "trace": _adcopy_import_trace(None),
            }
        )
    if not adgroups and ad_rows:
        adgroups.append(
            {
                "campaign_name": campaign_name,
                "adgroup_name": "01_광고그룹",
                "keywords": [],
                "required_phrases": default_required_phrases,
                "required_exact_phrases": default_required_exact_phrases,
                "trace": _adcopy_import_trace(None),
            }
        )
    valid_group_names = {item["adgroup_name"] for item in adgroups}
    fallback_group = adgroups[0]["adgroup_name"] if adgroups else ""

    ads: list[dict[str, Any]] = []
    for index, item in enumerate(ad_rows, start=1):
        adgroup_name = _adcopy_canonical_adgroup_name(
            _adcopy_import_text(item, ("adgroup_name", "ad_group_name", "adgroup", "ad_group", "광고그룹명", "광고그룹"), fallback_group)
        )
        if adgroup_name not in valid_group_names:
            adgroup_name = fallback_group
        ad_name = _adcopy_import_text(item, ("ad_name", "creative_name", "copy_name", "name", "소재명", "광고명"), f"AD_{index:03d}")
        ads.append(
            {
                "ad_name": ad_name,
                "adgroup_name": adgroup_name,
                "title": _adcopy_import_text(item, ("title", "headline", "heading", "제목", "타이틀")),
                "copy": _adcopy_import_text(item, ("copy", "body", "description", "text", "카피", "본문", "설명")),
                "link": _adcopy_import_text(item, ("link", "landing_url", "target_url", "url", "랜딩 URL", "랜딩URL"), default_link),
                "image_link": _adcopy_import_text(item, ("image_link", "image_url", "image", "이미지 URL", "이미지URL"), default_image),
                "trace": _adcopy_import_trace(item, _adcopy_import_first_value(item, ("human_check", "review_status", "검수상태", "휴먼 체크", "검수"))),
            }
        )

    generated = {
        "policy": {"banned_terms": banned_terms},
        "campaigns": campaigns,
        "adgroups": adgroups,
        "ads": ads,
    }
    source_format = "native" if all(isinstance(source.get(key), list) for key in ("campaigns", "adgroups", "ads")) else "compat"
    return {
        "generated": generated,
        "import_report": {
            "source_format": source_format,
            "source_label": source_label,
            "campaigns": len(generated["campaigns"]),
            "adgroups": len(adgroups),
            "ads": len(ads),
            "warnings": [
                message
                for message in [
                    ""
                    if campaign_rows
                    else (
                        f"별도 캠페인 시트가 없어 광고그룹 campaign_name에서 캠페인 {len(derived_campaign_rows)}개를 복원했습니다."
                        if derived_campaign_rows
                        else "캠페인 행이 없어 현재 입력값으로 campaign을 보강했습니다."
                    ),
                    "" if group_rows else "광고그룹 행이 없어 소재의 광고그룹명 또는 기본 광고그룹으로 보강했습니다.",
                    "" if ad_rows else "소재 행을 찾지 못했습니다.",
                ]
                if message
            ],
        },
    }


def _adcopy_finding(level: str, scope: str, item: str, field: str, rule: str, message: str) -> dict[str, str]:
    return {"level": level, "scope": scope, "item": item, "field": field, "rule": rule, "message": message}


ADCOPY_AWKWARD_PATTERNS = (
    "고려한",
    "누려보세요",
    "만나보세요",
    "도와드립니다",
    "최적의",
    "완벽한",
    "획기적인",
)


ADCOPY_POLICY_RISK_PATTERNS = (
    "무조건",
    "누구나",
    "최고",
    "최저가",
    "1위",
    "100%",
    "보장",
    "완치",
    "무료",
    "공짜",
    "당첨",
    "수익",
    "환불 보장",
)


def _adcopy_hostname(value: Any) -> str:
    try:
        parsed = urlparse(str(value or "").strip())
    except ValueError:
        return ""
    return parsed.netloc.lower().replace("www.", "")


def _adcopy_recommendation(kind: str, title: str, message: str) -> dict[str, str]:
    return {"kind": kind, "title": title, "message": message}


def _adcopy_quality_grade(score: int) -> str:
    if score >= 90:
        return "A"
    if score >= 80:
        return "B"
    if score >= 70:
        return "C"
    if score >= 55:
        return "D"
    return "F"


def _adcopy_warning_penalty(warnings: list[dict[str, str]]) -> int:
    counts: dict[str, int] = {}
    for warning in warnings:
        rule = str(warning.get("rule") or "warning")
        counts[rule] = counts.get(rule, 0) + 1
    penalty = 0
    for rule, count in counts.items():
        weight = ADCOPY_WARNING_SCORE_WEIGHTS.get(rule, 3)
        penalty += min(count * weight, weight * 4)
    return penalty


def _validate_generated_adcopy(data: dict[str, Any]) -> dict[str, Any]:
    errors: list[dict[str, str]] = []
    warnings: list[dict[str, str]] = []
    creative_checks: list[dict[str, Any]] = []
    campaigns = data.get("campaigns") if isinstance(data.get("campaigns"), list) else []
    adgroups = data.get("adgroups") if isinstance(data.get("adgroups"), list) else []
    ads = data.get("ads") if isinstance(data.get("ads"), list) else []
    policy = data.get("policy") if isinstance(data.get("policy"), dict) else {}
    banned_terms = [str(term).strip() for term in policy.get("banned_terms") or [] if str(term).strip()]
    if not campaigns:
        errors.append(_adcopy_finding("error", "campaigns", "-", "campaign_name", "campaign_required", "캠페인 정보가 비어 있습니다."))
    for campaign in campaigns:
        name = str(campaign.get("campaign_name") or "").strip()
        if not name:
            errors.append(_adcopy_finding("error", "campaigns", "-", "campaign_name", "campaign_name_required", "캠페인명이 비어 있습니다."))
        if not campaign.get("target_countries"):
            errors.append(_adcopy_finding("error", "campaigns", name or "-", "target_countries", "countries_required", "target_countries가 비어 있습니다."))
        if float(campaign.get("budget_max") or 0) <= 0:
            warnings.append(_adcopy_finding("warning", "campaigns", name or "-", "budget_max", "budget_empty", "예산이 0 또는 미입력입니다. 업로드 전 확인하세요."))
    adgroup_by_name = {str(item.get("adgroup_name") or "").strip(): item for item in adgroups}
    adgroup_names = set(adgroup_by_name)
    context_hint_owner: dict[str, str] = {}
    if not adgroups:
        errors.append(_adcopy_finding("error", "adgroups", "-", "adgroup_name", "adgroup_required", "광고그룹이 생성되지 않았습니다."))
    for adgroup in adgroups:
        name = str(adgroup.get("adgroup_name") or "").strip()
        if not name:
            errors.append(_adcopy_finding("error", "adgroups", "-", "adgroup_name", "adgroup_name_required", "광고그룹명이 비어 있습니다."))
        elif not _adcopy_stage_from_adgroup_name(name):
            warnings.append(
                _adcopy_finding(
                    "warning",
                    "adgroups",
                    name or "-",
                    "adgroup_name",
                    "purchase_journey_stage",
                    "광고그룹명 끝의 구매여정 단계는 문제정의, 제품발견, 비교검토, 단일제품평가, 신청전환, 사용도움 중 하나여야 합니다.",
                )
            )
        if name and "*" not in name:
            warnings.append(
                _adcopy_finding(
                    "warning",
                    "adgroups",
                    name,
                    "adgroup_name",
                    "adgroup_name_structure",
                    "광고그룹명은 상품/SKU*타깃*세부의도_구매여정 구조를 권장합니다.",
                )
            )
        if adgroup.get("max_bid") not in (None, "", []):
            errors.append(_adcopy_finding("error", "adgroups", name or "-", "max_bid", "max_bid_must_be_empty", "max_bid는 항상 빈칸이어야 합니다."))
        keywords = adgroup.get("keywords") or []
        if len(keywords) < 5:
            warnings.append(_adcopy_finding("warning", "adgroups", name or "-", "keywords", "context_hint_min_count", "검색 맥락이 5개 미만입니다."))
        keyword_texts = [str(keyword.get("text") if isinstance(keyword, dict) else keyword).strip() for keyword in keywords]
        if keyword_texts:
            contextual_count = sum(1 for text in keyword_texts if _adcopy_is_contextual_hint(text))
            if contextual_count / len(keyword_texts) < 0.5:
                warnings.append(
                    _adcopy_finding(
                        "warning",
                        "adgroups",
                        name or "-",
                        "keywords",
                        "context_hint_question_ratio",
                        "Context Hints는 검색어형보다 질문형·상황형 문장 비중을 높이는 편이 좋습니다.",
                    )
                )
        for keyword_text in keyword_texts:
            normalized_hint = _adcopy_normalized_hint(keyword_text)
            if not normalized_hint:
                continue
            owner = context_hint_owner.get(normalized_hint)
            if owner and owner != name:
                warnings.append(
                    _adcopy_finding(
                        "warning",
                        "adgroups",
                        name or "-",
                        "keywords",
                        "context_hint_duplicate",
                        f"'{keyword_text}' 문맥 힌트가 {owner} 광고그룹과 중복됩니다.",
                    )
                )
            else:
                context_hint_owner[normalized_hint] = name
        required_phrases = [str(value).strip() for value in adgroup.get("required_phrases") or [] if str(value).strip()]
        if required_phrases and not any(str(ad.get("adgroup_name") or "").strip() == name for ad in ads):
            warnings.append(_adcopy_finding("warning", "adgroups", name or "-", "required_phrases", "core_element_no_ads", "카피 핵심 반영 요소를 확인할 대상 광고가 없습니다."))
        for keyword in keywords:
            keyword_text = str(keyword.get("text") if isinstance(keyword, dict) else keyword).strip()
            if any(term in keyword_text for term in banned_terms):
                errors.append(_adcopy_finding("error", "adgroups", name or "-", "keywords", "banned_term", f"금지어 포함: {keyword_text}"))
    seen_creatives: dict[str, str] = {}
    seen_titles: dict[str, str] = {}
    landing_hosts: list[str] = []
    ads_by_group: dict[str, list[dict[str, str]]] = {}
    risk_finding_count = 0
    if not ads:
        errors.append(_adcopy_finding("error", "ads", "-", "title", "ads_required", "광고 소재가 생성되지 않았습니다."))
    for ad in ads:
        ad_name = str(ad.get("ad_name") or "").strip() or "-"
        adgroup_name = str(ad.get("adgroup_name") or "").strip()
        title = str(ad.get("title") or "").strip()
        copy = str(ad.get("copy") or "").strip()
        combined = f"{title} {copy}"
        ads_by_group.setdefault(adgroup_name, []).append({"ad_name": ad_name, "title": title, "copy": copy})
        ad_issues: list[dict[str, str]] = []
        if adgroup_name not in adgroup_names:
            errors.append(_adcopy_finding("error", "ads", ad_name, "adgroup_name", "adgroup_missing", "존재하지 않는 광고그룹을 참조합니다."))
            ad_issues.append({"level": "error", "rule": "adgroup_missing", "message": "존재하지 않는 광고그룹을 참조합니다."})
        if not title:
            errors.append(_adcopy_finding("error", "ads", ad_name, "title", "title_required", "제목이 비어 있습니다."))
            ad_issues.append({"level": "error", "rule": "title_required", "message": "제목이 비어 있습니다."})
        elif len(title) > 24:
            errors.append(_adcopy_finding("error", "ads", ad_name, "title", "title_max_24", f"제목 {len(title)}자 — 최대 24자"))
            ad_issues.append({"level": "error", "rule": "title_max_24", "message": f"제목 {len(title)}자 — 최대 24자"})
        elif len(title) < 16 or len(title) > 18:
            warnings.append(_adcopy_finding("warning", "ads", ad_name, "title", "title_len_recommended", f"제목 {len(title)}자 — 권장 16~18자"))
            ad_issues.append({"level": "warning", "rule": "title_len_recommended", "message": f"제목 {len(title)}자 — 권장 16~18자"})
        if not copy:
            errors.append(_adcopy_finding("error", "ads", ad_name, "copy", "copy_required", "문구가 비어 있습니다."))
            ad_issues.append({"level": "error", "rule": "copy_required", "message": "문구가 비어 있습니다."})
        elif len(copy) > 48:
            errors.append(_adcopy_finding("error", "ads", ad_name, "copy", "copy_max_48", f"문구 {len(copy)}자 — 최대 48자"))
            ad_issues.append({"level": "error", "rule": "copy_max_48", "message": f"문구 {len(copy)}자 — 최대 48자"})
        elif len(copy) < 18 or len(copy) > 42:
            warnings.append(_adcopy_finding("warning", "ads", ad_name, "copy", "copy_len_recommended", f"문구 {len(copy)}자 — 권장 18~42자"))
            ad_issues.append({"level": "warning", "rule": "copy_len_recommended", "message": f"문구 {len(copy)}자 — 권장 18~42자"})
        awkward_hits = [pattern for pattern in ADCOPY_AWKWARD_PATTERNS if pattern and pattern in combined]
        if awkward_hits:
            message = f"담당자 확인이 필요한 어색한 표현: {', '.join(awkward_hits)}"
            warnings.append(_adcopy_finding("warning", "ads", ad_name, "title/copy", "awkward_phrase", message))
            ad_issues.append({"level": "warning", "rule": "awkward_phrase", "message": message})
        risk_hits = [pattern for pattern in ADCOPY_POLICY_RISK_PATTERNS if pattern and pattern in combined]
        if risk_hits:
            risk_finding_count += 1
            message = f"광고 기준 또는 광고주 확인이 필요한 표현: {', '.join(risk_hits)}"
            warnings.append(_adcopy_finding("warning", "ads", ad_name, "title/copy", "policy_risk_phrase", message))
            ad_issues.append({"level": "warning", "rule": "policy_risk_phrase", "message": message})
        if title and title == copy:
            errors.append(_adcopy_finding("error", "ads", ad_name, "copy", "copy_equals_title", "문구가 제목과 동일합니다."))
            ad_issues.append({"level": "error", "rule": "copy_equals_title", "message": "문구가 제목과 동일합니다."})
        normalized_title = _adcopy_normalized_creative_text(title)
        normalized_copy = _adcopy_normalized_creative_text(copy)
        if (
            normalized_title
            and normalized_copy
            and title != copy
            and len(normalized_title) >= 8
            and (normalized_title in normalized_copy or normalized_copy in normalized_title)
        ):
            message = "문구가 제목을 거의 그대로 반복합니다. 제목의 맥락을 이어 해결 방향이나 혜택을 추가하세요."
            warnings.append(_adcopy_finding("warning", "ads", ad_name, "title/copy", "copy_repeats_title", message))
            ad_issues.append({"level": "warning", "rule": "copy_repeats_title", "message": message})
        creative_key = title + "\x00" + copy
        exact_duplicate = bool(title and creative_key in seen_creatives)
        if exact_duplicate:
            errors.append(_adcopy_finding("error", "ads", ad_name, "title", "creative_duplicate", f"제목·문구가 {seen_creatives[creative_key]}와 동일합니다."))
            ad_issues.append({"level": "error", "rule": "creative_duplicate", "message": f"제목·문구가 {seen_creatives[creative_key]}와 동일합니다."})
        seen_creatives[creative_key] = ad_name
        title_key = re.sub(r"\s+", "", title).lower()
        if title_key and title_key in seen_titles and not exact_duplicate:
            warnings.append(_adcopy_finding("warning", "ads", ad_name, "title", "title_duplicate", f"제목이 {seen_titles[title_key]}와 거의 같습니다."))
            ad_issues.append({"level": "warning", "rule": "title_duplicate", "message": f"제목이 {seen_titles[title_key]}와 거의 같습니다."})
        if title_key:
            seen_titles.setdefault(title_key, ad_name)
        for field in ("link", "image_link"):
            url = str(ad.get(field) or "").strip()
            if not re.match(r"^https?://", url):
                errors.append(_adcopy_finding("error", "ads", ad_name, field, "url_required", f"{field}는 http 또는 https URL이어야 합니다."))
                ad_issues.append({"level": "error", "rule": "url_required", "message": f"{field}는 http 또는 https URL이어야 합니다."})
            elif field == "link":
                host = _adcopy_hostname(url)
                if host:
                    landing_hosts.append(host)
        for term in banned_terms:
            if term and term in combined:
                errors.append(_adcopy_finding("error", "ads", ad_name, "title/copy", "banned_term", f"금지어 포함: {term}"))
                ad_issues.append({"level": "error", "rule": "banned_term", "message": f"금지어 포함: {term}"})
        group = adgroup_by_name.get(adgroup_name)
        for phrase in [str(value).strip() for value in (group or {}).get("required_phrases") or [] if str(value).strip()]:
            terms = _adcopy_core_terms(phrase)
            if terms and not any(term in combined for term in terms):
                warnings.append(_adcopy_finding("warning", "ads", ad_name, "title/copy", "core_element_weak", f"카피 핵심 반영 요소 확인 필요: {phrase}"))
                ad_issues.append({"level": "warning", "rule": "core_element_weak", "message": f"카피 핵심 반영 요소 확인 필요: {phrase}"})
        for phrase in [str(value).strip() for value in (group or {}).get("required_exact_phrases") or [] if str(value).strip()]:
            if phrase not in combined:
                errors.append(_adcopy_finding("error", "ads", ad_name, "title/copy", "required_exact_phrase_missing", f"필수 고정 문구 누락: {phrase}"))
                ad_issues.append({"level": "error", "rule": "required_exact_phrase_missing", "message": f"필수 고정 문구 누락: {phrase}"})
        status = "error" if any(issue["level"] == "error" for issue in ad_issues) else ("warning" if ad_issues else "pass")
        creative_checks.append(
            {
                "ad_name": ad_name,
                "adgroup_name": adgroup_name,
                "title_length": len(title),
                "copy_length": len(copy),
                "status": status,
                "issues": ad_issues[:5],
            }
        )
    for group_name, group_ads in ads_by_group.items():
        if len(group_ads) < 3:
            continue
        cta_count = sum(1 for item in group_ads if _adcopy_is_cta_like(item.get("copy")))
        if cta_count / len(group_ads) > 0.3:
            warnings.append(
                _adcopy_finding(
                    "warning",
                    "ads",
                    group_name or "-",
                    "copy",
                    "cta_ending_ratio",
                    "동일 광고그룹 내 직접 행동 유도형 종결은 전체 카피의 30% 이하로 분산하는 편이 좋습니다.",
                )
            )
        ending_counts: dict[str, int] = {}
        for item in group_ads:
            bucket = _adcopy_ending_bucket(item.get("copy"))
            if bucket:
                ending_counts[bucket] = ending_counts.get(bucket, 0) + 1
        repeated_ending = next((ending for ending, count in ending_counts.items() if count >= 3 and count / len(group_ads) >= 0.6), "")
        if repeated_ending:
            warnings.append(
                _adcopy_finding(
                    "warning",
                    "ads",
                    group_name or "-",
                    "copy",
                    "copy_ending_repetition",
                    f"동일 광고그룹 내 카피 종결 '{repeated_ending}' 반복이 많습니다.",
                )
            )
        title_form_counts: dict[str, int] = {}
        opening_counts: dict[str, int] = {}
        copy_roles: set[str] = set()
        for item in group_ads:
            title_form = _adcopy_title_form(item.get("title"))
            if title_form:
                title_form_counts[title_form] = title_form_counts.get(title_form, 0) + 1
            opening = _adcopy_opening_bucket(item.get("copy"))
            if opening:
                opening_counts[opening] = opening_counts.get(opening, 0) + 1
            role = _adcopy_copy_role(item.get("copy"))
            if role:
                copy_roles.add(role)
        repeated_title_form = next(
            (form for form, count in title_form_counts.items() if count >= 3 and count / len(group_ads) >= 0.6),
            "",
        )
        if repeated_title_form:
            warnings.append(
                _adcopy_finding(
                    "warning",
                    "ads",
                    group_name or "-",
                    "title",
                    "title_form_repetition",
                    "동일 광고그룹 내 제목 형식이 편중되었습니다. 질문형·공감/조건형·진술형을 섞어 주세요.",
                )
            )
        repeated_opening = next(
            (opening for opening, count in opening_counts.items() if count >= 3 and count / len(group_ads) >= 0.6),
            "",
        )
        if repeated_opening:
            warnings.append(
                _adcopy_finding(
                    "warning",
                    "ads",
                    group_name or "-",
                    "copy",
                    "copy_opening_repetition",
                    f"동일 광고그룹 내 카피 도입부 '{repeated_opening}' 반복이 많습니다.",
                )
            )
        if len(group_ads) >= 3 and len(copy_roles) < 2:
            warnings.append(
                _adcopy_finding(
                    "warning",
                    "ads",
                    group_name or "-",
                    "copy",
                    "copy_role_diversity",
                    "동일 광고그룹의 카피 역할이 한 유형에 치우쳤습니다. 해결·혜택·비교·이용·결과·행동 유도형을 섞어 주세요.",
                )
            )
    unique_landing_hosts = sorted(set(landing_hosts))
    if len(unique_landing_hosts) > 1:
        message = f"소재 랜딩 도메인이 {len(unique_landing_hosts)}개입니다: {', '.join(unique_landing_hosts[:4])}"
        warnings.append(_adcopy_finding("warning", "ads", "-", "link", "landing_domain_mismatch", message))
    warning_rules = [item.get("rule", "") for item in warnings]
    total_keywords = 0
    customer_keywords = 0
    confidence_values: list[float] = []
    for adgroup in adgroups:
        for keyword in adgroup.get("keywords") or []:
            total_keywords += 1
            if isinstance(keyword, dict) and keyword.get("origin") == "customer_data":
                customer_keywords += 1
        trace = adgroup.get("trace") if isinstance(adgroup.get("trace"), dict) else {}
        try:
            confidence = float(trace.get("confidence_score"))
            if 0 <= confidence <= 1:
                confidence_values.append(confidence)
        except (TypeError, ValueError):
            pass
    for ad in ads:
        trace = ad.get("trace") if isinstance(ad.get("trace"), dict) else {}
        try:
            confidence = float(trace.get("confidence_score"))
            if 0 <= confidence <= 1:
                confidence_values.append(confidence)
        except (TypeError, ValueError):
            pass
    average_confidence = round(sum(confidence_values) / len(confidence_values), 2) if confidence_values else None
    quality_score = 0 if not ads else max(0, min(100, 100 - (len(errors) * 24) - _adcopy_warning_penalty(warnings)))
    grade = _adcopy_quality_grade(quality_score)
    if errors:
        readiness = "업로드 보류"
    elif quality_score >= 90 and not warnings:
        readiness = "초안 양호"
    elif quality_score >= 75:
        readiness = "담당자 확인 권장"
    else:
        readiness = "수정 권장"
    recommendations: list[dict[str, str]] = []
    if errors:
        recommendations.append(_adcopy_recommendation("error", "오류 우선 수정", "필수 고정 문구 누락, URL, 길이 초과 같은 오류를 먼저 해결해야 합니다."))
    if any(rule in warning_rules for rule in ("purchase_journey_stage", "adgroup_name_structure")):
        recommendations.append(_adcopy_recommendation("brief", "구매여정 구조 확인", "광고그룹명은 상품/SKU*타깃*세부의도_구매여정 구조와 6개 고정 단계로 맞춰 주세요."))
    if any(rule in warning_rules for rule in ("context_hint_question_ratio", "context_hint_duplicate")):
        recommendations.append(_adcopy_recommendation("brief", "대화 맥락 보강", "Context Hints는 질문형·상황형 문장 중심으로, 광고그룹 간 중복을 줄여 주세요."))
    if any(rule in warning_rules for rule in ("cta_ending_ratio", "copy_ending_repetition")):
        recommendations.append(_adcopy_recommendation("copy", "카피 종결 분산", "같은 광고그룹 안에서는 질문형, 공감형, 정보형, 조건형, 결과형 등으로 문장 끝을 분산하는 편이 좋습니다."))
    if "title_len_recommended" in warning_rules:
        recommendations.append(_adcopy_recommendation("copy", "제목 길이 보정", "제목은 가능하면 16~18자로 맞추면 노출 영역에서 더 안정적으로 보입니다."))
    if "copy_len_recommended" in warning_rules:
        recommendations.append(_adcopy_recommendation("copy", "문구 밀도 조정", "문구는 18~42자 안에서 구체적 혜택이나 사용 상황을 한 가지 넣는 편이 좋습니다."))
    if "awkward_phrase" in warning_rules:
        recommendations.append(_adcopy_recommendation("tone", "자연어 표현 수정", "'고려한', '누려보세요' 같은 문어체 표현은 더 구체적인 사용 장면으로 바꾸는 편이 좋습니다."))
    if "policy_risk_phrase" in warning_rules:
        recommendations.append(_adcopy_recommendation("policy", "민감 표현 재검토", "'최고', '보장', '무료' 같은 표현은 근거와 광고주 확인 없이 그대로 쓰지 않는 편이 안전합니다."))
    if "landing_domain_mismatch" in warning_rules:
        recommendations.append(_adcopy_recommendation("landing", "랜딩 도메인 통일 확인", "소재별 랜딩 도메인이 다릅니다. 의도된 분기인지, 잘못 붙은 URL인지 확인해 주세요."))
    if total_keywords and customer_keywords == 0:
        recommendations.append(_adcopy_recommendation("brief", "브리프 기반 검색 맥락 보강", "모든 검색 맥락이 AI 추론입니다. 광고주가 제공한 검색어 또는 제품 표현을 1개 이상 넣어 주세요."))
    if average_confidence is not None and average_confidence < 0.75:
        recommendations.append(_adcopy_recommendation("review", "확인 강도 상향", "모델 신뢰도가 낮습니다. 랜딩 페이지와 소재 이미지의 실제 표현을 한 번 더 대조해 주세요."))
    if not recommendations and not errors:
        recommendations.append(_adcopy_recommendation("ok", "담당자 확인 후 사용 가능", "자동 점검에서 큰 문제는 없습니다. 광고주명, 랜딩, 이미지, 광고 기준 표현만 최종 확인하세요."))
    return {
        "ok": not errors,
        "errors": errors,
        "warnings": warnings,
        "creative_checks": creative_checks,
        "quality": {
            "score": quality_score,
            "grade": grade,
            "readiness": readiness,
            "average_confidence": average_confidence,
            "customer_keyword_count": customer_keywords,
            "keyword_count": total_keywords,
            "policy_risk_count": risk_finding_count,
            "landing_domain_count": len(unique_landing_hosts),
            "landing_domains": unique_landing_hosts[:8],
            "recommendations": recommendations[:5],
        },
        "summary": {
            "campaigns": len(campaigns),
            "adgroups": len(adgroups),
            "ads": len(ads),
            "errors": len(errors),
            "warnings": len(warnings),
            "quality_score": quality_score,
            "quality_grade": grade,
            "readiness": readiness,
            "policy_risk_count": risk_finding_count,
            "landing_domain_count": len(unique_landing_hosts),
        },
    }


def _safe_float(value: Any) -> float:
    try:
        return float(str(value or 0).replace(",", "").strip())
    except (TypeError, ValueError):
        return 0.0


def _money_to_micros(value: Any) -> int:
    return max(0, int(round(_safe_float(value) * 1_000_000)))


def _cpm_krw_to_max_bid_micros(value: Any) -> int:
    # OpenAI Ads max_bid_micros is per impression. Operators enter CPM in KRW.
    return max(0, int(round(_safe_float(value) * 1000)))


def _date_to_kst_timestamp(value: Any, *, end_of_day: bool = False) -> int | None:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        parsed = datetime.fromisoformat(text[:10])
    except ValueError:
        return None
    hour, minute, second = (23, 59, 59) if end_of_day else (0, 0, 0)
    parsed = parsed.replace(hour=hour, minute=minute, second=second, microsecond=0, tzinfo=KST)
    return int(parsed.timestamp())


def _campaign_day_count(start_date: Any, end_date: Any) -> int:
    try:
        start = datetime.fromisoformat(str(start_date or "")[:10]).date()
        end = datetime.fromisoformat(str(end_date or "")[:10]).date()
    except ValueError:
        return 1
    return max(1, (end - start).days + 1)


def _adcopy_trace_status(item: dict[str, Any]) -> str:
    trace = item.get("trace") if isinstance(item.get("trace"), dict) else {}
    return str(trace.get("review_status") or trace.get("validation_status") or "").strip()


def _adcopy_is_excluded(item: dict[str, Any]) -> bool:
    return _adcopy_trace_status(item) in {"제외", "excluded", "exclude"}


def _adcopy_keyword_texts(group: dict[str, Any]) -> list[str]:
    values: list[str] = []
    for keyword in group.get("keywords") or []:
        if isinstance(keyword, dict):
            text = str(keyword.get("text") or "").strip()
        else:
            text = str(keyword or "").strip()
        if text:
            values.append(text)
    return list(dict.fromkeys(values))


def _first_campaign(data: dict[str, Any]) -> dict[str, Any]:
    campaigns = data.get("campaigns") if isinstance(data.get("campaigns"), list) else []
    for campaign in campaigns:
        if isinstance(campaign, dict):
            return campaign
    return {}


def _build_adcopy_draft_plan(payload: AdsAdcopyDraftPlanRequest) -> dict[str, Any]:
    generated = payload.generated if isinstance(payload.generated, dict) else {}
    validation_report = _validate_generated_adcopy(generated)
    campaign = _first_campaign(generated)
    campaign_name = str(campaign.get("campaign_name") or "").strip()
    if not campaign_name:
        raise HTTPException(status_code=400, detail="임시 등록할 캠페인명이 없습니다. 먼저 광고 문안을 생성하거나 업로드 파일을 확인해 주세요.")
    adgroups = [item for item in (generated.get("adgroups") or []) if isinstance(item, dict)]
    ads_all = [item for item in (generated.get("ads") or []) if isinstance(item, dict)]
    ads = [item for item in ads_all if not _adcopy_is_excluded(item)]
    if not ads:
        raise HTTPException(status_code=400, detail="임시 등록에 포함할 소재가 없습니다. 제외 상태가 아닌 소재를 1개 이상 남겨 주세요.")
    adgroup_names = {str(ad.get("adgroup_name") or "").strip() for ad in ads if str(ad.get("adgroup_name") or "").strip()}
    included_adgroups = [group for group in adgroups if str(group.get("adgroup_name") or "").strip() in adgroup_names]
    if not included_adgroups and adgroups:
        included_adgroups = adgroups

    warnings: list[str] = []
    budget_type = str(campaign.get("budget_type") or "daily").strip().lower()
    budget_max = _safe_float(campaign.get("budget_max"))
    days = _campaign_day_count(campaign.get("launch_date"), campaign.get("end_date"))
    lifetime_budget = budget_max * days if budget_type in {"daily", "일 예산", "일예산"} else budget_max
    if budget_type in {"daily", "일 예산", "일예산"}:
        warnings.append(f"일 예산 {int(budget_max):,}원을 캠페인 기간 {days}일 기준 총 예산 {int(lifetime_budget):,}원으로 변환합니다.")
    if lifetime_budget <= 0:
        warnings.append("캠페인 예산이 0원입니다. 실제 생성 전 예산을 확인해야 합니다.")

    default_bid_krw = _safe_float(payload.default_max_bid_krw)
    default_bid_micros = _cpm_krw_to_max_bid_micros(default_bid_krw)
    if default_bid_micros <= 0:
        warnings.append("기본 CPM 입찰가가 0원입니다. 광고그룹 생성 단계는 입찰가를 입력해야 진행할 수 있습니다.")
    if not payload.location_ids:
        warnings.append("Location ID가 비어 있어 위치 타깃 payload를 보내지 않습니다. 필요 시 국가/지역 location id를 입력해 주세요.")

    objective = str(campaign.get("objective") or "").strip()
    if objective and objective.lower() not in {"views", "view", "reach", "impression", "impressions", "노출"}:
        warnings.append("캠페인 목표값은 화면 표시와 점검 기준으로만 사용합니다. OpenAI Ads 등록 요청에는 별도 목표 필드를 보내지 않습니다.")

    campaign_payload: dict[str, Any] = {
        "name": campaign_name,
        "status": "paused",
        "budget": {"lifetime_spend_limit_micros": _money_to_micros(lifetime_budget)},
    }
    start_time = _date_to_kst_timestamp(campaign.get("launch_date"))
    end_time = _date_to_kst_timestamp(campaign.get("end_date"), end_of_day=True)
    if start_time:
        campaign_payload["start_time"] = start_time
    if end_time:
        campaign_payload["end_time"] = end_time
    locations = [str(item).strip() for item in payload.location_ids if str(item or "").strip()]
    if locations:
        campaign_payload["targeting"] = {"locations": {"include": [{"id": item} for item in locations]}}

    adgroup_payloads: list[dict[str, Any]] = []
    for index, group in enumerate(included_adgroups, start=1):
        name = str(group.get("adgroup_name") or f"{index:02d}_광고그룹").strip()
        context_hints = _adcopy_keyword_texts(group)
        if not context_hints:
            warnings.append(f"{name} 광고그룹의 검색 맥락이 비어 있습니다.")
        adgroup_payloads.append(
            {
                "name": name,
                "ad_count": sum(1 for ad in ads if str(ad.get("adgroup_name") or "").strip() == name),
                "context_hints": context_hints,
                "api_payload": {
                    "name": name,
                    "status": "paused",
                    "context_hints": context_hints,
                    "bidding_config": {
                        "billing_event_type": "impression",
                        "max_bid_micros": default_bid_micros,
                    },
                },
            }
        )

    image_urls = list(dict.fromkeys(str(ad.get("image_link") or "").strip() for ad in ads if re.match(r"^https?://", str(ad.get("image_link") or "").strip(), re.I)))
    assets = [{"image_url": url} for url in image_urls]
    if not assets:
        warnings.append("업로드할 이미지 URL이 없습니다. 소재 생성 전 이미지 URL을 확인해야 합니다.")

    ad_payloads: list[dict[str, Any]] = []
    for index, ad in enumerate(ads, start=1):
        ad_name = str(ad.get("ad_name") or f"AD_{index:03d}").strip()
        adgroup_name = str(ad.get("adgroup_name") or "").strip()
        target_url = str(ad.get("link") or "").strip()
        image_url = str(ad.get("image_link") or "").strip()
        if not re.match(r"^https?://", target_url, re.I):
            warnings.append(f"{ad_name} 소재의 랜딩 URL이 올바르지 않습니다.")
        if not re.match(r"^https?://", image_url, re.I):
            warnings.append(f"{ad_name} 소재의 이미지 URL이 올바르지 않습니다.")
        ad_payloads.append(
            {
                "key": f"{adgroup_name}::{ad_name}",
                "name": ad_name,
                "adgroup_name": adgroup_name,
                "image_url": image_url,
                "api_payload": {
                    "name": ad_name,
                    "status": "paused",
                    "creative": {
                        "type": "chat_card",
                        "title": str(ad.get("title") or "").strip(),
                        "body": str(ad.get("copy") or "").strip(),
                        "target_url": target_url,
                    },
                },
            }
        )

    return {
        "ok": True,
        "advertiser_name": payload.advertiser_name,
        "mode": "draft_paused",
        "safety": {
            "default_status": "paused",
            "activation_requires_explicit_confirm": True,
            "activation_disabled": os.getenv("ADS_DRAFT_ALLOW_ACTIVATION", "").strip().lower() not in {"1", "true", "yes"},
            "operator_step_required": True,
        },
        "summary": {
            "campaigns": 1,
            "adgroups": len(adgroup_payloads),
            "ads": len(ad_payloads),
            "assets": len(assets),
            "excluded_ads": len(ads_all) - len(ads),
            "validation_errors": len(validation_report.get("errors") or []),
            "validation_warnings": len(validation_report.get("warnings") or []),
        },
        "warnings": list(dict.fromkeys(warnings)),
        "campaign": {
            "name": campaign_name,
            "objective": objective,
            "budget_type": budget_type or "-",
            "budget_krw": lifetime_budget,
            "start_date": campaign.get("launch_date") or "",
            "end_date": campaign.get("end_date") or "",
            "api_payload": campaign_payload,
        },
        "ad_groups": adgroup_payloads,
        "assets": assets,
        "ads": ad_payloads,
        "next_steps": [
            "광고주 계정 확인",
            "비활성 캠페인 임시 등록",
            "이미지 파일 업로드",
            "비활성 광고그룹 임시 등록",
            "비활성 소재 임시 등록",
            "담당자 최종 확인 후 게시 전환은 현재 정책상 비활성화",
        ],
        "validation_report": validation_report,
    }


async def _ads_draft_api_request(
    api_key: str,
    method: str,
    path: str,
    json_body: dict[str, Any] | None = None,
) -> dict[str, Any]:
    from rag_chatbot.ads_api import load_ads_api_settings

    settings = load_ads_api_settings(api_key=api_key)
    url = f"{settings.base_url}{path}"
    headers = {
        "Authorization": f"Bearer {settings.api_key}",
        "Accept": "application/json",
        "Content-Type": "application/json",
    }
    async with httpx.AsyncClient(timeout=30) as client:
        response = await client.request(method.upper(), url, headers=headers, json=json_body)
    if response.status_code >= 400:
        raise httpx.HTTPStatusError(
            f"HTTP {response.status_code}",
            request=response.request,
            response=response,
        )
    if not response.text.strip():
        return {}
    return response.json()


def _state_mapping(state: dict[str, Any], key: str) -> dict[str, str]:
    value = state.get(key)
    return dict(value) if isinstance(value, dict) else {}


def _ads_response_id(data: dict[str, Any]) -> str:
    for key in ("id", "campaign_id", "ad_group_id", "ad_id", "file_id"):
        value = str(data.get(key) or "").strip()
        if value:
            return value
    nested = data.get("data")
    if isinstance(nested, dict):
        return _ads_response_id(nested)
    return ""


def _adcopy_draft_check(level: str, title: str, message: str, *, code: str = "") -> dict[str, str]:
    return {"level": level, "title": title, "message": message, "code": code}


def _adcopy_draft_preflight_result(
    plan: dict[str, Any],
    credential: dict[str, str],
    account: dict[str, Any],
) -> dict[str, Any]:
    checks: list[dict[str, str]] = []
    summary = plan.get("summary") if isinstance(plan.get("summary"), dict) else {}
    campaign_payload = (plan.get("campaign") or {}).get("api_payload") or {}
    campaign_budget = ((campaign_payload.get("budget") or {}).get("lifetime_spend_limit_micros") or 0)
    ad_groups = plan.get("ad_groups") if isinstance(plan.get("ad_groups"), list) else []
    ads = plan.get("ads") if isinstance(plan.get("ads"), list) else []
    assets = plan.get("assets") if isinstance(plan.get("assets"), list) else []
    validation_errors = int(summary.get("validation_errors") or 0)
    validation_warnings = int(summary.get("validation_warnings") or 0)

    if credential.get("api_key"):
        if account.get("ok") is False:
            checks.append(
                _adcopy_draft_check(
                    "error",
                    "계정 확인 실패",
                    str(account.get("error") or "Ads API 키로 광고 계정을 확인하지 못했습니다."),
                    code="account_error",
                )
            )
        else:
            ad_account = account.get("ad_account") if isinstance(account.get("ad_account"), dict) else {}
            account_label = " · ".join(str(value) for value in [ad_account.get("name"), ad_account.get("currency_code"), ad_account.get("timezone")] if value)
            checks.append(_adcopy_draft_check("success", "계정 확인", account_label or "Ads API 키가 등록되어 있고 계정 조회가 가능합니다.", code="account_ok"))
    else:
        checks.append(_adcopy_draft_check("error", "Ads API 키 필요", "활성화된 광고주 API 키가 있어야 임시 등록을 진행할 수 있습니다.", code="api_key_missing"))

    if (plan.get("safety") or {}).get("default_status") == "paused":
        checks.append(_adcopy_draft_check("success", "비활성 생성", "캠페인, 광고그룹, 소재는 비활성 상태로 생성됩니다.", code="paused_default"))
    else:
        checks.append(_adcopy_draft_check("error", "상태 확인 필요", "생성 요청의 기본 상태가 비활성인지 확인해야 합니다.", code="paused_missing"))

    if (plan.get("safety") or {}).get("activation_disabled", True):
        checks.append(_adcopy_draft_check("success", "게시 전환 차단", "현재 환경에서는 게시 전환 단계가 서버에서 차단됩니다.", code="activation_blocked"))
    else:
        checks.append(_adcopy_draft_check("warning", "게시 전환 가능 환경", "환경변수상 게시 전환이 허용되어 있습니다. 운영 승인 전 실행하지 마세요.", code="activation_enabled"))

    if validation_errors:
        checks.append(_adcopy_draft_check("error", "점검 오류", f"자동 점검 오류 {validation_errors}개를 먼저 수정해야 합니다.", code="validation_error"))
    elif validation_warnings:
        checks.append(_adcopy_draft_check("warning", "점검 주의", f"자동 점검 주의 {validation_warnings}개가 있습니다. 담당자 확인 후 진행하세요.", code="validation_warning"))
    else:
        checks.append(_adcopy_draft_check("success", "점검 상태", "자동 점검에서 차단 오류가 없습니다.", code="validation_ok"))

    if campaign_budget > 0:
        checks.append(_adcopy_draft_check("success", "예산", f"총 예산 {int(campaign_budget / 1_000_000):,}원 기준으로 임시 등록 계획을 만들었습니다.", code="budget_ok"))
    else:
        checks.append(_adcopy_draft_check("warning", "예산 확인", "예산이 0원입니다. 생성 전 캠페인 예산을 확인하세요.", code="budget_empty"))

    if ad_groups:
        max_bid_values = [int(((item.get("api_payload") or {}).get("bidding_config") or {}).get("max_bid_micros") or 0) for item in ad_groups]
        if all(value > 0 for value in max_bid_values):
            checks.append(_adcopy_draft_check("success", "입찰가", f"광고그룹 {len(ad_groups)}개의 CPM 입찰가가 설정되어 있습니다.", code="bid_ok"))
        else:
            checks.append(_adcopy_draft_check("error", "입찰가 필요", "광고그룹 생성 전 기본 CPM 입찰가를 0보다 크게 입력해야 합니다.", code="bid_missing"))
    else:
        checks.append(_adcopy_draft_check("error", "광고그룹 없음", "임시 등록할 광고그룹이 없습니다.", code="adgroup_missing"))

    if assets:
        checks.append(_adcopy_draft_check("success", "이미지 URL", f"업로드 대상 이미지 {len(assets)}개가 있습니다.", code="asset_ok"))
    else:
        checks.append(_adcopy_draft_check("error", "이미지 URL 필요", "소재 임시 등록까지 진행하려면 http/https 이미지 URL이 필요합니다.", code="asset_missing"))

    if ads:
        checks.append(_adcopy_draft_check("success", "소재", f"생성 대상 소재 {len(ads)}개를 확인했습니다.", code="ads_ok"))
    else:
        checks.append(_adcopy_draft_check("error", "소재 없음", "임시 등록할 소재가 없습니다.", code="ads_missing"))

    plan_warnings = plan.get("warnings") if isinstance(plan.get("warnings"), list) else []
    for warning in plan_warnings[:5]:
        checks.append(_adcopy_draft_check("warning", "계획 확인", str(warning), code="plan_warning"))

    blocking = [check for check in checks if check["level"] == "error"]
    ready = not blocking
    state_patch: dict[str, str] = {}
    if ready and account.get("ok") is not False and credential.get("api_key"):
        state_patch["account_verified_at"] = datetime.now(KST).isoformat()
    return {
        "ok": ready,
        "readiness": "draft_ready" if ready else "needs_fix",
        "checks": checks,
        "state_patch": state_patch,
        "blocking_count": len(blocking),
        "warning_count": sum(1 for check in checks if check["level"] == "warning"),
    }


def _ads_api_error_detail(exc: httpx.HTTPStatusError) -> str:
    response = exc.response
    detail = response.text[:1000] if response is not None else str(exc)
    status = response.status_code if response is not None else "?"
    return f"OpenAI Ads API 임시 등록 실패: HTTP {status} · {detail}"


def _ads_dashboard_cached_payload(cache_row: dict[str, Any]) -> dict[str, Any]:
    payload = dict(cache_row.get("payload") or {})
    refreshed_at = str(cache_row.get("refreshed_at") or "")
    advertiser_count = int(payload.get("advertiser_count") or payload.get("live_status_total_advertisers") or 0)
    advertiser_label = f"전체 활성 광고주({advertiser_count}개)" if advertiser_count else "전체 활성 광고주"
    refreshed_label = _date_label(refreshed_at)
    cache_note = "전체 활성 광고주 지표는 백그라운드 캐시 기준입니다. 개별 광고주 실시간 조회와 수치가 조금 다를 수 있습니다."
    if refreshed_label:
        cache_note = f"{advertiser_label} 지표는 {refreshed_label} 캐시 기준입니다. 개별 광고주 실시간 조회와 수치가 조금 다를 수 있습니다."
    payload.update(
        {
            "cache_status": "hit",
            "cache_source": "background",
            "cache_refreshed_at": refreshed_at,
            "cache_note": cache_note,
            "key_source": "advertiser_collection",
            "advertiser_name": "전체 활성 광고주",
            "advertiser_label": advertiser_label,
        }
    )
    payload["warning"] = _join_unique_messages(cache_note, str(payload.get("warning") or ""))
    return payload


def _validation_error_details(exc: ValidationError) -> list[dict[str, Any]]:
    return exc.errors(include_url=False, include_input=False, include_context=False)


async def _validated_admin_payload(request: Request, model: type[BaseModel], invalid_message: str) -> Any:
    try:
        body = await request.json()
    except Exception as exc:
        raise HTTPException(status_code=400, detail=invalid_message) from exc
    try:
        return model.model_validate(body)
    except ValidationError as exc:
        raise HTTPException(status_code=422, detail=_validation_error_details(exc)) from exc


def _production_guide_decks() -> dict[str, str]:
    root = project_root()
    source_candidates = [
        root / "backups" / "production_pages_20260621_pre_dev_cutover" / "index.html",
        root / "public" / "index.html",
        root / "templates" / "index.html",
    ]
    missing_panels: list[str] = []
    for source in source_candidates:
        if not source.exists():
            continue
        soup = BeautifulSoup(source.read_text(encoding="utf-8"), "html.parser")
        decks: dict[str, str] = {}
        missing_panels = []
        for deck_key, panel_id in GUIDE_DECK_PANELS.items():
            deck = soup.select_one(f"#{panel_id} .slide-deck")
            if deck is None:
                missing_panels.append(panel_id)
                break
            decks[deck_key] = deck.decode_contents()
        if not missing_panels:
            return decks
    missing = ", ".join(missing_panels or GUIDE_DECK_PANELS.values())
    raise HTTPException(status_code=500, detail=f"{missing} 안내자료 슬라이드를 찾을 수 없습니다.")


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.get("/favicon.ico", include_in_schema=False)
def favicon() -> Response:
    return Response(status_code=204)


def _index_file() -> FileResponse:
    return FileResponse(project_root() / "templates" / "index.html")


@app.get("/", include_in_schema=False)
def index() -> FileResponse:
    return _index_file()


@app.get("/checker", include_in_schema=False)
def checker_page() -> FileResponse:
    return _index_file()


@app.get("/rag", include_in_schema=False)
def rag_page() -> FileResponse:
    return _index_file()


@app.get("/intake", include_in_schema=False)
def intake_page() -> FileResponse:
    return _index_file()


@app.get("/creative-upload-draft", include_in_schema=False)
def creative_upload_draft_page() -> FileResponse:
    return FileResponse(project_root() / "templates" / "creative_upload_draft.html")


@app.get("/ads-api-draft", include_in_schema=False)
def ads_api_draft_page() -> FileResponse:
    raise HTTPException(status_code=404, detail="공개 접근이 제한된 검토 페이지입니다.")


@app.get("/slides", include_in_schema=False)
def slides_page() -> FileResponse:
    return _index_file()


@app.get("/admin", include_in_schema=False)
def admin_page() -> FileResponse:
    return FileResponse(project_root() / "dev" / "admin.html")


@app.get("/dev", include_in_schema=False)
def dev_index_page() -> FileResponse:
    return FileResponse(project_root() / "dev" / "index.html")


@app.get("/dev/admin", include_in_schema=False)
def dev_admin_page() -> FileResponse:
    return FileResponse(project_root() / "dev" / "admin.html")


@app.get("/dev/creative-upload-draft", include_in_schema=False)
def dev_creative_upload_draft_page() -> FileResponse:
    return FileResponse(project_root() / "dev" / "creative_upload_draft.html")


@app.get("/api/notice", include_in_schema=False)
def public_notice() -> dict[str, Any]:
    from admin_store import get_notice_config

    return get_notice_config()


@app.get("/api/menu-settings", include_in_schema=False)
def public_menu_settings() -> dict[str, Any]:
    from admin_store import get_menu_settings

    return get_menu_settings()


@app.post("/api/analytics/visit", include_in_schema=False)
def analytics_visit(request: VisitRequest) -> dict[str, Any]:
    from admin_store import record_page_visit

    raw_page = (request.page or "root").strip()
    page = raw_page[:80] or "root"
    label = (request.label or "").strip()[:120] or PAGE_LABELS.get(page, page)
    return record_page_visit(page, label)


@app.get("/api/faqs", include_in_schema=False)
def public_operating_faqs() -> dict[str, Any]:
    from admin_store import list_static_operating_faqs

    return list_static_operating_faqs()


@app.post("/api/admin/faqs/refresh", include_in_schema=False)
def admin_refresh_operating_faqs(request: Request) -> dict[str, Any]:
    _require_admin(request)
    return {
        "ok": True,
        "disabled": True,
        "message": "FAQ 자동 업데이트는 비활성화되어 있습니다. FAQ 관리 메뉴에서 저장한 항목만 사용자 화면에 노출됩니다.",
    }


@app.get("/api/admin/faqs", include_in_schema=False)
def admin_operating_faqs(request: Request) -> dict[str, Any]:
    from admin_store import list_admin_operating_faqs

    _require_admin(request)
    include_deleted = str(request.query_params.get("include_deleted") or "1").lower() in {"1", "true", "yes"}
    return list_admin_operating_faqs(include_deleted=include_deleted)


@app.post("/api/admin/faqs/categories", include_in_schema=False)
def admin_create_faq_category(request: Request, item: FaqCategoryRequest) -> dict[str, Any]:
    from admin_store import upsert_operating_faq_category

    _require_admin(request)
    try:
        return upsert_operating_faq_category(item.model_dump())
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/api/admin/faqs/categories/{category_id}/update", include_in_schema=False)
def admin_update_faq_category(request: Request, category_id: str, item: FaqCategoryRequest) -> dict[str, Any]:
    from admin_store import upsert_operating_faq_category

    _require_admin(request)
    try:
        return upsert_operating_faq_category(item.model_dump(), category_id=category_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/api/admin/faqs/categories/{category_id}/delete", include_in_schema=False)
def admin_delete_faq_category(request: Request, category_id: str) -> dict[str, Any]:
    from admin_store import delete_operating_faq_category

    _require_admin(request)
    try:
        return delete_operating_faq_category(category_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/api/admin/faqs/items", include_in_schema=False)
def admin_create_faq_item(request: Request, item: FaqItemRequest) -> dict[str, Any]:
    from admin_store import upsert_operating_faq_item

    _require_admin(request)
    try:
        return upsert_operating_faq_item(item.model_dump())
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/api/admin/faqs/items/{item_id}/update", include_in_schema=False)
def admin_update_faq_item(request: Request, item_id: str, item: FaqItemRequest) -> dict[str, Any]:
    from admin_store import upsert_operating_faq_item

    _require_admin(request)
    try:
        return upsert_operating_faq_item(item.model_dump(), item_id=item_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/api/admin/faqs/items/{item_id}/delete", include_in_schema=False)
def admin_delete_faq_item(request: Request, item_id: str) -> dict[str, Any]:
    from admin_store import delete_operating_faq_item

    _require_admin(request)
    try:
        return delete_operating_faq_item(item_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/api/cron/faqs", include_in_schema=False)
def cron_refresh_operating_faqs(request: Request) -> dict[str, Any]:
    _require_cron(request)
    return {
        "ok": True,
        "disabled": True,
        "message": "FAQ automatic refresh is disabled.",
    }


@app.get("/api/cron/ads-dashboard-cache", include_in_schema=False)
async def cron_refresh_ads_dashboard_cache(request: Request) -> dict[str, Any]:
    _require_cron(request)
    from admin_store import (
        get_ads_campaign_objective_override_map,
        list_active_ads_api_key_credentials,
        save_ads_dashboard_cache,
    )
    from rag_chatbot.ads_api import apply_campaign_objective_overrides, fetch_ads_dashboard_for_advertisers

    credentials = list_active_ads_api_key_credentials()
    if not credentials:
        payload = {
            "ok": False,
            "configured": False,
            "advertiser_name": "전체 활성 광고주",
            "key_source": "advertiser_collection",
            "error": "활성화된 광고주 Ads API 키가 없습니다.",
        }
    else:
        payload = await fetch_ads_dashboard_for_advertisers(
            credentials,
            include_aggregate_extensions=True,
        )
    payload = apply_campaign_objective_overrides(payload, get_ads_campaign_objective_override_map())
    saved = save_ads_dashboard_cache(payload, ADS_DASHBOARD_CACHE_KEY)
    return {
        "ok": True,
        "refreshed": True,
        "cache_key": ADS_DASHBOARD_CACHE_KEY,
        "refreshed_at": saved.get("refreshed_at", ""),
        "active_advertiser_count": len(credentials),
        "payload_ok": bool(payload.get("ok")),
        "campaign_count": len(payload.get("campaigns") or []),
        "trend_count": len(payload.get("trend") or []),
        "device_count": len(payload.get("device_breakdown") or []),
        "storage": saved.get("storage", ""),
    }


@app.get("/api/admin/notice", include_in_schema=False)
def admin_notice(request: Request) -> dict[str, Any]:
    from admin_store import get_notice_config

    _require_admin(request)
    return get_notice_config()


@app.post("/api/admin/notice", include_in_schema=False)
def update_admin_notice(request: Request, notice: NoticeConfigRequest) -> dict[str, Any]:
    from admin_store import save_notice_config

    _require_admin(request)
    return save_notice_config(notice.model_dump())


@app.get("/api/admin/menu-settings", include_in_schema=False)
def admin_menu_settings(request: Request) -> dict[str, Any]:
    from admin_store import get_menu_settings

    _require_admin(request)
    return get_menu_settings()


@app.post("/api/admin/menu-settings", include_in_schema=False)
def update_admin_menu_settings(request: Request, settings: MenuSettingsRequest) -> dict[str, Any]:
    from admin_store import save_menu_settings

    _require_admin(request)
    return save_menu_settings(settings.model_dump())


@app.get("/api/guide-slides", include_in_schema=False)
def public_guide_slides() -> dict[str, Any]:
    from admin_store import get_slide_content

    return get_slide_content()


@app.get("/api/guide-deck-html", include_in_schema=False)
def public_guide_deck_html() -> dict[str, Any]:
    return {"decks": _production_guide_decks()}


@app.get("/api/admin/guide-slides", include_in_schema=False)
def admin_guide_slides(request: Request) -> dict[str, Any]:
    from admin_store import get_slide_content

    _require_admin(request)
    return get_slide_content()


@app.post("/api/admin/guide-slides", include_in_schema=False)
def update_admin_guide_slides(request: Request, slides: SlideContentRequest) -> dict[str, Any]:
    from admin_store import save_slide_content

    _require_admin(request)
    return save_slide_content(slides.model_dump())


@app.post("/api/admin/guide-image", response_model=ImageUploadResponse, include_in_schema=False)
async def upload_admin_guide_image(request: Request, file: UploadFile = File(...)) -> ImageUploadResponse:
    _require_admin(request)

    supabase_url = os.getenv("SUPABASE_URL", "").strip().rstrip("/")
    service_role_key = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "").strip()
    bucket = os.getenv("SUPABASE_STORAGE_BUCKET", "openai-ad-assets").strip()
    if not supabase_url or not service_role_key:
        raise HTTPException(
            status_code=503,
            detail="이미지 업로드 저장소가 아직 설정되지 않았습니다. 공개 직접 이미지 URL을 입력해 주세요.",
        )

    suffix_by_type = {
        "image/png": ".png",
        "image/jpeg": ".jpg",
        "image/webp": ".webp",
    }
    content_type = file.content_type or ""
    suffix = suffix_by_type.get(content_type)
    if not suffix:
        raise HTTPException(status_code=400, detail="PNG, JPG, WEBP 이미지만 업로드할 수 있습니다.")

    data = await file.read()
    if len(data) > 8 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="안내자료 이미지는 8MB 이하만 업로드할 수 있습니다.")

    try:
        from PIL import Image

        image = Image.open(BytesIO(data))
        width, height = image.size
    except Exception as exc:
        raise HTTPException(status_code=400, detail="이미지 파일을 열 수 없습니다.") from exc
    if width < 240 or height < 160 or width > 4096 or height > 4096:
        raise HTTPException(status_code=400, detail="안내자료 이미지는 240×160 이상, 4096×4096 이하만 업로드할 수 있습니다.")

    safe_stem = re.sub(r"[^0-9A-Za-z._-]+", "-", Path(file.filename or "guide-image").stem).strip("-")
    object_name = f"guide-slides/{uuid4().hex}-{safe_stem or 'guide-image'}{suffix}"
    upload_url = f"{supabase_url}/storage/v1/object/{bucket}/{object_name}"
    headers = {
        "apikey": service_role_key,
        "Authorization": f"Bearer {service_role_key}",
        "Content-Type": content_type,
        "x-upsert": "false",
    }
    async with httpx.AsyncClient(timeout=30, follow_redirects=True) as client:
        response = await client.post(upload_url, content=data, headers=headers)
    if response.status_code >= 400:
        raise HTTPException(status_code=502, detail=f"Supabase Storage 업로드 실패(HTTP {response.status_code})")

    public_url = f"{supabase_url}/storage/v1/object/public/{bucket}/{object_name}"
    return ImageUploadResponse(
        image_url=public_url,
        width=width,
        height=height,
        content_type=content_type,
    )


@app.get("/api/admin/analytics", include_in_schema=False)
def admin_analytics(request: Request) -> dict[str, Any]:
    from admin_store import get_visit_analytics

    _require_admin(request)
    return get_visit_analytics(request.query_params.get("period", "30"))


@app.get("/api/admin/ads-api-keys", include_in_schema=False)
def admin_ads_api_keys(request: Request) -> dict[str, Any]:
    from admin_store import list_ads_api_keys

    _require_admin(request)
    return list_ads_api_keys()


@app.post("/api/admin/ads-api-keys", include_in_schema=False)
def admin_save_ads_api_key(request: Request, payload: AdsApiKeyRequest) -> dict[str, Any]:
    from admin_store import upsert_ads_api_key

    _require_admin(request)
    try:
        return upsert_ads_api_key(payload.model_dump(exclude_unset=True))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/api/admin/ads-api-keys/inspect", include_in_schema=False)
async def admin_inspect_ads_api_key(request: Request, payload: AdsApiKeyInspectRequest) -> dict[str, Any]:
    from rag_chatbot.ads_api import fetch_ad_account_metadata

    _require_admin(request)
    try:
        return await fetch_ad_account_metadata(payload.ads_api_key)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.get("/api/admin/ads-campaign-objectives", include_in_schema=False)
def admin_ads_campaign_objectives(request: Request) -> dict[str, Any]:
    from admin_store import list_ads_campaign_objective_overrides

    _require_admin(request)
    return list_ads_campaign_objective_overrides()


@app.post("/api/admin/ads-campaign-objectives", include_in_schema=False)
def admin_save_ads_campaign_objective(
    request: Request,
    payload: AdsCampaignObjectiveRequest,
) -> dict[str, Any]:
    from admin_store import upsert_ads_campaign_objective_override

    _require_admin(request)
    try:
        return upsert_ads_campaign_objective_override(payload.model_dump())
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/api/admin/adcopy/generate", include_in_schema=False)
async def admin_generate_adcopy(request: Request) -> dict[str, Any]:
    _require_admin(request)
    payload = await _validated_admin_payload(request, AdsAdcopyGenerateRequest, "생성할 카피 브리프 본문이 올바르지 않습니다.")
    engine = _resolve_adcopy_engine(payload.engine)
    generated_payload = await _call_openai_adcopy(payload)
    generated = _normalize_generated_adcopy(generated_payload.get("generated") or {}, payload)
    validation_report = _validate_generated_adcopy(generated)
    return {
        "ok": validation_report["ok"],
        "engine": engine,
        "engine_label": ADCOPY_ENGINE_LABELS[engine],
        "requested_engine": payload.engine,
        "fallback_used": False,
        "model": generated_payload.get("model"),
        "response_id": generated_payload.get("raw_response_id"),
        "generated": generated,
        "validation_report": validation_report,
        "summary": validation_report["summary"],
        "notice": "광고 문안 생성 결과입니다. 업로드 또는 임시 등록 전 담당자와 광고주 확인이 필요합니다.",
    }


@app.get("/api/admin/adcopy/engines", include_in_schema=False)
def admin_adcopy_engines(request: Request) -> dict[str, Any]:
    _require_admin(request)
    return _adcopy_engine_settings()


@app.post("/api/admin/adcopy/validate", include_in_schema=False)
async def admin_validate_adcopy(request: Request) -> dict[str, Any]:
    _require_admin(request)
    try:
        body = await request.json()
    except Exception as exc:
        raise HTTPException(status_code=400, detail="점검할 광고 업로드 파일 내용이 올바르지 않습니다.") from exc
    generated = body.get("generated") if isinstance(body, dict) and isinstance(body.get("generated"), dict) else body
    if not isinstance(generated, dict):
        raise HTTPException(status_code=400, detail="광고 업로드 파일 내용을 전달해 주세요.")
    validation_report = _validate_generated_adcopy(generated)
    return {
        "ok": validation_report["ok"],
        "generated": generated,
        "validation_report": validation_report,
        "summary": validation_report["summary"],
        "notice": "수정된 광고 업로드 파일을 다시 점검했습니다. 업로드 전 담당자 최종 확인이 필요합니다.",
    }


@app.post("/api/admin/adcopy/import", include_in_schema=False)
async def admin_import_adcopy(request: Request) -> dict[str, Any]:
    _require_admin(request)
    payload = await _validated_admin_payload(request, AdsAdcopyImportRequest, "불러올 기존 작업 파일 내용이 올바르지 않습니다.")
    imported = _normalize_external_generated_adcopy(payload)
    generated = imported["generated"]
    validation_report = _validate_generated_adcopy(generated)
    return {
        "ok": validation_report["ok"],
        "model": "기존 작업 불러오기",
        "generated": generated,
        "validation_report": validation_report,
        "summary": validation_report["summary"],
        "import_report": imported["import_report"],
        "notice": "기존 작업 파일을 광고 업로드 형식으로 변환했습니다. 자동 점검과 임시 등록에 사용할 수 있습니다.",
    }


@app.post("/api/admin/adcopy/review-state", include_in_schema=False)
async def admin_save_adcopy_review_state(request: Request) -> dict[str, Any]:
    _require_admin(request)
    payload = await _validated_admin_payload(request, AdsAdcopyReviewStateRequest, "저장할 문안 검토 상태 내용이 올바르지 않습니다.")
    generated = payload.generated if isinstance(payload.generated, dict) else {}
    if not generated:
        raise HTTPException(status_code=400, detail="저장할 광고 업로드 파일 내용이 없습니다.")
    validation_report = _validate_generated_adcopy(generated)
    from admin_store import save_adcopy_review_snapshot

    snapshot = save_adcopy_review_snapshot(
        {
            "advertiser_name": payload.advertiser_name,
            "campaign_name": payload.campaign_name,
            "source_label": payload.source_label,
            "generated": generated,
            "validation_report": validation_report or payload.validation_report,
        }
    )
    return {
        "ok": True,
        "snapshot": snapshot,
        "validation_report": validation_report,
        "notice": "현재 문안 검토 상태를 저장했습니다. 저장은 임시 등록이나 게시 전환과 무관합니다.",
    }


@app.get("/api/admin/adcopy/review-state", include_in_schema=False)
def admin_list_adcopy_review_state(request: Request) -> dict[str, Any]:
    from admin_store import list_adcopy_review_snapshots

    _require_admin(request)
    limit = _query_int(request, "limit", 20, maximum=50)
    return list_adcopy_review_snapshots(limit=limit)


@app.post("/api/admin/adcopy/review-workbook", include_in_schema=False)
async def admin_download_adcopy_review_workbook(request: Request) -> StreamingResponse:
    _require_admin(request)
    payload = await _validated_admin_payload(request, AdsAdcopyReviewStateRequest, "다운로드할 문안 검토 내용이 올바르지 않습니다.")
    generated = payload.generated if isinstance(payload.generated, dict) else {}
    if not generated:
        raise HTTPException(status_code=400, detail="다운로드할 문안 검토 내용이 없습니다.")
    validation_report = _validate_generated_adcopy(generated)
    return StreamingResponse(
        BytesIO(_adcopy_review_workbook_bytes(generated, validation_report)),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="adcopy_review_working_copy.xlsx"'},
    )


@app.get("/api/admin/adcopy/review-state/{snapshot_id}", include_in_schema=False)
def admin_get_adcopy_review_state(snapshot_id: str, request: Request) -> dict[str, Any]:
    from admin_store import get_adcopy_review_snapshot

    _require_admin(request)
    try:
        return get_adcopy_review_snapshot(snapshot_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc


@app.delete("/api/admin/adcopy/review-state/{snapshot_id}", include_in_schema=False)
def admin_delete_adcopy_review_state(snapshot_id: str, request: Request) -> dict[str, Any]:
    from admin_store import delete_adcopy_review_snapshot

    _require_admin(request)
    try:
        return delete_adcopy_review_snapshot(snapshot_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc


@app.get("/api/admin/adcopy/sample-workbook", include_in_schema=False)
def admin_download_adcopy_sample_workbook(request: Request) -> StreamingResponse:
    _require_admin(request)
    filename = "openai_ads_adcopy_review_sample.xlsx"
    return StreamingResponse(
        BytesIO(_adcopy_sample_workbook_bytes()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/admin/adcopy/brief-sample-workbook", include_in_schema=False)
def admin_download_adcopy_brief_sample_workbook(request: Request) -> StreamingResponse:
    _require_admin(request)
    filename = "openai_ads_adcopy_brief_sample.xlsx"
    return StreamingResponse(
        BytesIO(_adcopy_brief_sample_workbook_bytes()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/api/admin/adcopy/brief-workbook", include_in_schema=False)
async def admin_import_adcopy_brief_workbook(request: Request, file: UploadFile = File(...)) -> dict[str, Any]:
    _require_admin(request)
    filename = file.filename or "brief.xlsx"
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="xlsx 파일만 업로드할 수 있습니다.")
    content = await file.read()
    if len(content) > 8 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="파일은 8MB 이하만 업로드할 수 있습니다.")
    workbook_dump = _adcopy_workbook_dump(content, filename)
    parsed = _adcopy_brief_from_workbook_dump(workbook_dump)
    return {
        "ok": True,
        **parsed,
        "workbook": {"file": filename, "sheets": len(workbook_dump.get("sheets") or [])},
        "notice": "브리프 파일의 내용을 입력 화면에 채웠습니다. 생성 전에 누락 항목을 확인해 주세요.",
    }


@app.post("/api/admin/adcopy/import-workbook", include_in_schema=False)
async def admin_import_adcopy_workbook(request: Request, file: UploadFile = File(...)) -> dict[str, Any]:
    _require_admin(request)
    filename = file.filename or "review.xlsx"
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="xlsx 파일만 업로드할 수 있습니다.")
    content = await file.read()
    if len(content) > 8 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="파일은 8MB 이하만 업로드할 수 있습니다.")
    workbook_dump = _adcopy_workbook_dump(content, filename)
    payload = AdsAdcopyImportRequest(
        generated=workbook_dump,
        source_label=f"엑셀 불러오기: {Path(filename).name}",
    )
    imported = _normalize_external_generated_adcopy(payload)
    generated = imported["generated"]
    validation_report = _validate_generated_adcopy(generated)
    return {
        "ok": validation_report["ok"],
        "model": "엑셀 불러오기",
        "generated": generated,
        "validation_report": validation_report,
        "summary": validation_report["summary"],
        "import_report": imported["import_report"],
        "workbook": {"file": filename, "sheets": len(workbook_dump.get("sheets") or [])},
        "notice": "업로드한 엑셀 작업 파일을 광고 업로드 형식으로 변환했습니다.",
    }


@app.post("/api/admin/adcopy/inspect-landing", include_in_schema=False)
async def admin_inspect_adcopy_landing(request: Request) -> dict[str, Any]:
    _require_admin(request)
    payload = await _validated_admin_payload(request, AdsAdcopyLandingInspectRequest, "읽을 랜딩 URL 본문이 올바르지 않습니다.")
    html, final_url = await _fetch_landing_html(payload.landing_url)
    metadata = _inspect_landing_html(html, final_url)
    return {
        "ok": True,
        "source_url": payload.landing_url,
        "final_url": final_url,
        "metadata": metadata,
        "notice": "랜딩 페이지의 공개 메타 정보를 읽었습니다. 자동 입력 전 담당자가 내용을 확인해야 합니다.",
    }


@app.post("/api/admin/adcopy/draft-plan", include_in_schema=False)
async def admin_adcopy_draft_plan(request: Request) -> dict[str, Any]:
    _require_admin(request)
    payload = await _validated_admin_payload(request, AdsAdcopyDraftPlanRequest, "임시 등록 요청 내용이 올바르지 않습니다.")
    return _build_adcopy_draft_plan(payload)


@app.post("/api/admin/adcopy/draft-preflight", include_in_schema=False)
async def admin_adcopy_draft_preflight(request: Request) -> dict[str, Any]:
    from admin_store import get_ads_api_key_credential
    from rag_chatbot.ads_api import fetch_ad_account_metadata

    _require_admin(request)
    payload = await _validated_admin_payload(request, AdsAdcopyDraftPlanRequest, "사전 점검 요청 내용이 올바르지 않습니다.")
    plan = _build_adcopy_draft_plan(payload)
    credential = get_ads_api_key_credential(payload.advertiser_name)
    api_key = credential.get("api_key", "")
    account = await fetch_ad_account_metadata(api_key) if api_key else {"ok": False, "error": "활성 Ads API 키가 없습니다."}
    preflight = _adcopy_draft_preflight_result(plan, credential, account)
    return {
        **preflight,
        "plan": plan,
        "account": account,
        "credential": {
            "advertiser_name": credential.get("advertiser_name") or payload.advertiser_name,
            "industry": credential.get("industry") or "",
            "has_ads_api_key": bool(api_key),
        },
        "notice": "읽기 전용 사전 점검입니다. 캠페인, 광고그룹, 소재를 생성하지 않았고 게시 전환도 하지 않았습니다.",
    }


@app.post("/api/admin/adcopy/draft-execute", include_in_schema=False)
async def admin_adcopy_draft_execute(request: Request) -> dict[str, Any]:
    from admin_store import get_ads_api_key_credential, save_adcopy_draft_audit_log
    from rag_chatbot.ads_api import fetch_ad_account_metadata

    _require_admin(request)
    payload = await _validated_admin_payload(request, AdsAdcopyDraftExecuteRequest, "임시 등록 실행 내용이 올바르지 않습니다.")
    action = payload.action.strip()
    state = dict(payload.state or {})
    if action not in {"verify_account", "create_campaign", "upload_assets", "create_ad_groups", "create_ads", "activate_all"}:
        raise HTTPException(status_code=400, detail="지원하지 않는 임시 등록 단계입니다.")
    if action == "activate_all" and os.getenv("ADS_DRAFT_ALLOW_ACTIVATION", "").strip().lower() not in {"1", "true", "yes"}:
        raise HTTPException(
            status_code=400,
            detail="현재 운영 정책상 캠페인 게시 전환은 비활성화되어 있습니다. 비활성 임시 등록까지만 진행할 수 있습니다.",
        )
    credential = get_ads_api_key_credential(payload.advertiser_name)
    api_key = credential.get("api_key", "")
    if not api_key:
        raise HTTPException(status_code=400, detail=f"{payload.advertiser_name}의 활성 Ads API 키가 필요합니다.")
    plan = _build_adcopy_draft_plan(payload)

    def record_draft_audit(status: str, message: str, audit_logs: list[dict[str, Any]]) -> dict[str, Any] | None:
        audit = save_adcopy_draft_audit_log(
            {
                "advertiser_name": payload.advertiser_name,
                "campaign_name": str((plan.get("campaign") or {}).get("name") or ""),
                "action": action,
                "status": status,
                "message": message,
                "state": state,
                "logs": audit_logs,
            }
        )
        return audit.get("item")

    if action == "verify_account":
        try:
            account = await fetch_ad_account_metadata(api_key)
        except httpx.HTTPStatusError as exc:
            message = _ads_api_error_detail(exc)
            record_draft_audit("error", message, [{"level": "error", "message": message}])
            raise HTTPException(status_code=502, detail=message) from exc
        state["account_verified_at"] = datetime.now(KST).isoformat()
        audit_log = record_draft_audit(
            "success" if account.get("ok") is not False else "warning",
            "광고주 계정 확인을 완료했습니다.",
            [{"level": "success" if account.get("ok") is not False else "warning", "message": "광고주 계정 확인"}],
        )
        return {"ok": True, "action": action, "plan": plan, "state": state, "account": account, "audit_log": audit_log}

    if not payload.confirm:
        record_draft_audit(
            "error",
            "실행 확인이 없어 임시 등록 단계를 진행하지 않았습니다.",
            [{"level": "error", "message": "실행 확인 누락"}],
        )
        raise HTTPException(status_code=400, detail="실행 확인이 필요합니다. 각 단계는 운영자 확인 후에만 진행됩니다.")

    logs: list[dict[str, Any]] = []
    try:
        if action == "create_campaign":
            if state.get("campaign_id"):
                logs.append({"level": "info", "message": "이미 생성된 캠페인 ID를 사용합니다.", "id": state.get("campaign_id")})
            else:
                data = await _ads_draft_api_request(api_key, "POST", "/v1/campaigns", plan["campaign"]["api_payload"])
                campaign_id = _ads_response_id(data)
                if not campaign_id:
                    raise HTTPException(status_code=502, detail="캠페인 생성 응답에서 campaign id를 찾지 못했습니다.")
                state["campaign_id"] = campaign_id
                logs.append({"level": "success", "message": "비활성 캠페인을 임시 등록했습니다.", "id": campaign_id})

        elif action == "upload_assets":
            file_ids = _state_mapping(state, "file_ids")
            for asset in plan["assets"]:
                image_url = asset["image_url"]
                if file_ids.get(image_url):
                    logs.append({"level": "info", "message": "이미 업로드된 이미지를 재사용합니다.", "url": image_url, "id": file_ids[image_url]})
                    continue
                data = await _ads_draft_api_request(api_key, "POST", "/v1/upload", {"image_url": image_url})
                file_id = str(data.get("file_id") or _ads_response_id(data)).strip()
                if not file_id:
                    raise HTTPException(status_code=502, detail=f"이미지 업로드 응답에서 file_id를 찾지 못했습니다: {image_url}")
                file_ids[image_url] = file_id
                logs.append({"level": "success", "message": "소재 이미지를 업로드했습니다.", "url": image_url, "id": file_id})
            state["file_ids"] = file_ids

        elif action == "create_ad_groups":
            campaign_id = str(state.get("campaign_id") or "").strip()
            if not campaign_id:
                raise HTTPException(status_code=400, detail="광고그룹 생성 전 캠페인을 먼저 임시 등록해야 합니다.")
            ad_group_ids = _state_mapping(state, "ad_group_ids")
            for group in plan["ad_groups"]:
                name = group["name"]
                if ad_group_ids.get(name):
                    logs.append({"level": "info", "message": "이미 생성된 광고그룹을 재사용합니다.", "name": name, "id": ad_group_ids[name]})
                    continue
                api_payload = dict(group["api_payload"])
                if _safe_float(api_payload.get("bidding_config", {}).get("max_bid_micros")) <= 0:
                    raise HTTPException(status_code=400, detail="광고그룹 생성 전 기본 CPM 입찰가를 0보다 크게 입력해 주세요.")
                api_payload["campaign_id"] = campaign_id
                data = await _ads_draft_api_request(api_key, "POST", "/v1/ad_groups", api_payload)
                ad_group_id = _ads_response_id(data)
                if not ad_group_id:
                    raise HTTPException(status_code=502, detail=f"광고그룹 생성 응답에서 ad group id를 찾지 못했습니다: {name}")
                ad_group_ids[name] = ad_group_id
                logs.append({"level": "success", "message": "비활성 광고그룹을 임시 등록했습니다.", "name": name, "id": ad_group_id})
            state["ad_group_ids"] = ad_group_ids

        elif action == "create_ads":
            ad_group_ids = _state_mapping(state, "ad_group_ids")
            file_ids = _state_mapping(state, "file_ids")
            if not ad_group_ids:
                raise HTTPException(status_code=400, detail="소재 생성 전 광고그룹을 먼저 임시 등록해야 합니다.")
            ad_ids = _state_mapping(state, "ad_ids")
            for ad in plan["ads"]:
                ad_name = ad["name"]
                ad_key = ad.get("key") or ad_name
                if ad_ids.get(ad_key):
                    logs.append({"level": "info", "message": "이미 생성된 소재를 재사용합니다.", "name": ad_name, "id": ad_ids[ad_key]})
                    continue
                ad_group_id = ad_group_ids.get(ad["adgroup_name"])
                file_id = file_ids.get(ad["image_url"])
                if not ad_group_id:
                    raise HTTPException(status_code=400, detail=f"{ad['adgroup_name']} 광고그룹 ID가 없습니다. 광고그룹 생성 단계를 먼저 완료해 주세요.")
                if not file_id:
                    raise HTTPException(status_code=400, detail=f"{ad_name} 소재 이미지 file_id가 없습니다. 이미지 업로드 단계를 먼저 완료해 주세요.")
                api_payload = json.loads(json.dumps(ad["api_payload"]))
                api_payload["ad_group_id"] = ad_group_id
                api_payload["creative"]["file_id"] = file_id
                data = await _ads_draft_api_request(api_key, "POST", "/v1/ads", api_payload)
                ad_id = _ads_response_id(data)
                if not ad_id:
                    raise HTTPException(status_code=502, detail=f"소재 생성 응답에서 ad id를 찾지 못했습니다: {ad_name}")
                ad_ids[ad_key] = ad_id
                logs.append({"level": "success", "message": "비활성 소재를 임시 등록했습니다.", "name": ad_name, "id": ad_id})
            state["ad_ids"] = ad_ids

        elif action == "activate_all":
            campaign_id = str(state.get("campaign_id") or "").strip()
            ad_group_ids = _state_mapping(state, "ad_group_ids")
            ad_ids = _state_mapping(state, "ad_ids")
            if not campaign_id or not ad_group_ids or not ad_ids:
                raise HTTPException(status_code=400, detail="게시 전환 전 캠페인, 광고그룹, 소재 임시 등록이 모두 완료되어야 합니다.")
            for name, ad_id in ad_ids.items():
                await _ads_draft_api_request(api_key, "POST", f"/v1/ads/{ad_id}/activate", {})
                logs.append({"level": "success", "message": "소재를 활성화했습니다.", "name": name.rsplit("::", 1)[-1], "id": ad_id})
            for name, ad_group_id in ad_group_ids.items():
                await _ads_draft_api_request(api_key, "POST", f"/v1/ad_groups/{ad_group_id}/activate", {})
                logs.append({"level": "success", "message": "광고그룹을 활성화했습니다.", "name": name, "id": ad_group_id})
            await _ads_draft_api_request(api_key, "POST", f"/v1/campaigns/{campaign_id}/activate", {})
            logs.append({"level": "success", "message": "캠페인을 활성화했습니다.", "id": campaign_id})
            state["activated_at"] = datetime.now(KST).isoformat()

    except httpx.HTTPStatusError as exc:
        message = _ads_api_error_detail(exc)
        record_draft_audit(
            "error",
            message,
            logs + [{"level": "error", "message": message}],
        )
        raise HTTPException(status_code=502, detail=message) from exc
    except HTTPException as exc:
        message = str(exc.detail or "임시 등록 실행 중 오류가 발생했습니다.")
        record_draft_audit(
            "error",
            message,
            logs + [{"level": "error", "message": message}],
        )
        raise

    state["last_action"] = action
    state["updated_at"] = datetime.now(KST).isoformat()
    audit_log = record_draft_audit("success", f"{action} 단계를 완료했습니다.", logs)
    return {"ok": True, "action": action, "plan": plan, "state": state, "logs": logs, "audit_log": audit_log}


@app.get("/api/admin/adcopy/draft-audit", include_in_schema=False)
def admin_list_adcopy_draft_audit(request: Request) -> dict[str, Any]:
    from admin_store import list_adcopy_draft_audit_logs

    _require_admin(request)
    limit = _query_int(request, "limit", 20, maximum=50)
    return list_adcopy_draft_audit_logs(limit=limit)


@app.get("/api/admin/adcopy/draft-audit.csv", include_in_schema=False)
def admin_download_adcopy_draft_audit_csv(request: Request) -> Response:
    from admin_store import list_adcopy_draft_audit_logs

    _require_admin(request)
    limit = _query_int(request, "limit", 50, maximum=50)
    result = list_adcopy_draft_audit_logs(limit=limit)
    output = StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(["생성일", "광고주", "캠페인", "단계", "상태", "메시지", "캠페인 ID", "이미지 수", "광고그룹 수", "소재 수", "로그 요약"])
    for item in result.get("items") or []:
        state = item.get("state") or {}
        log_summary = " | ".join(
            str(log.get("message") or "").strip()
            for log in (item.get("logs") or [])
            if isinstance(log, dict) and str(log.get("message") or "").strip()
        )
        writer.writerow(
            [
                item.get("created_at") or "",
                item.get("advertiser_name") or "",
                item.get("campaign_name") or "",
                item.get("action") or "",
                item.get("status") or "",
                item.get("message") or "",
                state.get("campaign_id") or "",
                state.get("file_count") or 0,
                state.get("ad_group_count") or 0,
                state.get("ad_count") or 0,
                log_summary,
            ]
        )
    filename = f"openai_ads_adcopy_draft_audit_{datetime.now(KST):%Y%m%d_%H%M}.csv"
    return Response(
        content="\ufeff" + output.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/admin/ads-api-keys/{advertiser_name}/reveal", include_in_schema=False)
def admin_reveal_ads_api_key(advertiser_name: str, request: Request) -> dict[str, Any]:
    from admin_store import get_ads_api_key_secrets

    _require_admin(request)
    try:
        return get_ads_api_key_secrets(advertiser_name)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.delete("/api/admin/ads-api-keys/{advertiser_name}", include_in_schema=False)
def admin_delete_ads_api_key(advertiser_name: str, request: Request) -> dict[str, Any]:
    from admin_store import delete_ads_api_key

    _require_admin(request)
    try:
        return delete_ads_api_key(advertiser_name)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.get("/api/admin/ads-dashboard", include_in_schema=False)
async def admin_ads_dashboard(request: Request) -> dict[str, Any]:
    from admin_store import get_ads_campaign_objective_override_map
    from rag_chatbot.ads_api import (
        apply_campaign_objective_overrides,
        fetch_ads_dashboard,
        fetch_ads_dashboard_for_advertisers,
    )

    _require_admin(request)
    start_date = str(request.query_params.get("start_date") or "") or None
    end_date = str(request.query_params.get("end_date") or "") or None
    detail_scope = str(request.query_params.get("detail_scope") or "") or None
    detail_id = str(request.query_params.get("detail_id") or "") or None
    advertiser_name = str(request.query_params.get("advertiser_name") or "").strip()
    detail_advertiser_name = str(request.query_params.get("detail_advertiser_name") or "").strip()
    api_key = None
    advertiser_industry = ""
    if advertiser_name:
        from admin_store import get_ads_api_key_credential

        credential = get_ads_api_key_credential(advertiser_name)
        api_key = credential.get("api_key", "")
        advertiser_industry = credential.get("industry", "")
        if not api_key:
            return {
                "ok": False,
                "configured": False,
                "advertiser_name": advertiser_name,
                "error": f"{advertiser_name} Ads API 키가 등록되어 있지 않거나 비활성 상태입니다.",
            }
    else:
        from admin_store import get_ads_dashboard_cache, list_active_ads_api_key_credentials, save_ads_dashboard_cache

        credentials = list_active_ads_api_key_credentials()
        if credentials:
            if not detail_scope and not detail_id:
                cache_key = _ads_dashboard_range_cache_key(start_date, end_date)
                cache_row = get_ads_dashboard_cache(cache_key)
                if cache_row and not _ads_dashboard_payload_matches_range(
                    dict(cache_row.get("payload") or {}),
                    start_date,
                    end_date,
                ):
                    cache_row = None
                if not cache_row and cache_key != ADS_DASHBOARD_CACHE_KEY:
                    legacy_cache_row = get_ads_dashboard_cache(ADS_DASHBOARD_CACHE_KEY)
                    if legacy_cache_row and _ads_dashboard_payload_matches_range(
                        dict(legacy_cache_row.get("payload") or {}),
                        start_date,
                        end_date,
                    ):
                        cache_row = legacy_cache_row
                if cache_row and cache_row.get("payload"):
                    return apply_campaign_objective_overrides(
                        _ads_dashboard_cached_payload(cache_row),
                        get_ads_campaign_objective_override_map(),
                    )
            data = await fetch_ads_dashboard_for_advertisers(
                credentials,
                start_date=start_date,
                end_date=end_date,
                detail_scope=detail_scope,
                detail_id=detail_id,
                detail_advertiser_name=detail_advertiser_name,
                include_aggregate_extensions=not (detail_scope or detail_id),
            )
            if not detail_scope and not detail_id:
                if data.get("ok"):
                    cache_key = _ads_dashboard_range_cache_key(
                        str(data.get("range", {}).get("start_date") or start_date or ""),
                        str(data.get("range", {}).get("end_date") or end_date or ""),
                    )
                    saved = save_ads_dashboard_cache(data, cache_key)
                    advertiser_count = int(data.get("advertiser_count") or len(credentials))
                    advertiser_label = f"전체 활성 광고주({advertiser_count}개)"
                    range_payload = data.get("range") if isinstance(data.get("range"), dict) else {}
                    range_label = f"{range_payload.get('start_date') or start_date or ''} ~ {range_payload.get('end_date') or end_date or ''}".strip()
                    cache_note = f"{advertiser_label} 지표는 {range_label} 선택 기간 기준으로 새로 계산했습니다. 다음 조회부터 같은 기간은 캐시를 사용합니다."
                    data["cache_status"] = "refreshed"
                    data["cache_source"] = "request"
                    data["cache_refreshed_at"] = saved.get("refreshed_at", "")
                    data["cache_note"] = cache_note
                    data["advertiser_label"] = advertiser_label
                    data["warning"] = _join_unique_messages(cache_note, str(data.get("warning") or ""))
                else:
                    data["cache_status"] = "miss"
                    data["cache_note"] = "전체 활성 광고주 캐시가 아직 준비되지 않아 기본 성과만 실시간 조회했습니다. 백그라운드 캐시 갱신 후 디바이스/추이 차트가 표시됩니다."
                    data["warning"] = _join_unique_messages(data["cache_note"], str(data.get("warning") or ""))
            return apply_campaign_objective_overrides(data, get_ads_campaign_objective_override_map())
    data = await fetch_ads_dashboard(
        start_date=start_date,
        end_date=end_date,
        detail_scope=detail_scope,
        detail_id=detail_id,
        api_key=api_key,
        advertiser_name=advertiser_name,
        advertiser_industry=advertiser_industry,
    )
    return apply_campaign_objective_overrides(data, get_ads_campaign_objective_override_map())


@app.get("/api/admin/ads-dashboard/hourly", include_in_schema=False)
async def admin_ads_dashboard_hourly(request: Request) -> dict[str, Any]:
    from admin_store import get_ads_api_key_credential
    from rag_chatbot.ads_api import fetch_resource_hourly_insights

    _require_admin(request)
    advertiser_name = str(request.query_params.get("advertiser_name") or "").strip()
    resource_scope = str(request.query_params.get("resource_scope") or "campaign").strip()
    resource_id = str(request.query_params.get("resource_id") or "").strip()
    resource_name = str(request.query_params.get("resource_name") or "").strip()
    campaign_id = str(request.query_params.get("campaign_id") or "").strip()
    campaign_name = str(request.query_params.get("campaign_name") or "").strip()
    if not resource_id and campaign_id:
        resource_scope = "campaign"
        resource_id = campaign_id
        resource_name = resource_name or campaign_name
    since_hour = str(request.query_params.get("since") or "").strip()
    until_hour = str(request.query_params.get("until") or "").strip()
    if not advertiser_name:
        raise HTTPException(status_code=400, detail="광고주명이 필요합니다.")
    credential = get_ads_api_key_credential(advertiser_name)
    api_key = credential.get("api_key", "")
    if not api_key:
        return {
            "ok": False,
            "configured": False,
            "advertiser_name": advertiser_name,
            "campaign_id": campaign_id,
            "resource_scope": resource_scope,
            "resource_id": resource_id,
            "error": f"{advertiser_name} Ads API 키가 등록되어 있지 않거나 비활성 상태입니다.",
        }
    try:
        return await fetch_resource_hourly_insights(
            advertiser_name=advertiser_name,
            resource_scope=resource_scope,
            resource_id=resource_id,
            resource_name=resource_name,
            since_hour=since_hour,
            until_hour=until_hour,
            api_key=api_key,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.get("/api/admin/official-changes", include_in_schema=False)
def admin_official_changes(request: Request) -> dict[str, Any]:
    from admin_store import list_official_guide_changes

    _require_admin(request)
    try:
        limit = int(str(request.query_params.get("limit") or "15"))
    except ValueError:
        limit = 15
    try:
        page = int(str(request.query_params.get("page") or "1"))
    except ValueError:
        page = 1
    return list_official_guide_changes(
        limit=limit,
        page=page,
        start_date=str(request.query_params.get("start_date") or ""),
        end_date=str(request.query_params.get("end_date") or ""),
    )


@app.get("/api/admin/mail-review", include_in_schema=False)
def admin_mail_review(request: Request) -> dict[str, Any]:
    from admin_store import list_mail_review_rows

    _require_admin(request)
    status_filter = str(request.query_params.get("status") or "")
    try:
        limit = int(str(request.query_params.get("limit") or "10"))
    except ValueError:
        limit = 10
    try:
        page = int(str(request.query_params.get("page") or "1"))
    except ValueError:
        page = 1
    return list_mail_review_rows(status_filter=status_filter, limit=limit, page=page)


@app.post("/api/admin/mail-review/update", include_in_schema=False)
def admin_update_mail_review(request: Request, update: MailReviewUpdateRequest) -> dict[str, Any]:
    from admin_store import update_mail_review_row

    _require_admin(request)
    try:
        return update_mail_review_row(update.model_dump())
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.get("/api/admin/manual-rag", include_in_schema=False)
def admin_manual_rag_items(request: Request) -> dict[str, Any]:
    from admin_store import list_manual_rag_items

    _require_admin(request)
    include_deleted = str(request.query_params.get("include_deleted") or "").lower() in {"1", "true", "yes"}
    try:
        limit = int(str(request.query_params.get("limit") or "200"))
    except ValueError:
        limit = 200
    return list_manual_rag_items(include_deleted=include_deleted, limit=limit)


@app.post("/api/admin/manual-rag", include_in_schema=False)
def admin_create_manual_rag(request: Request, item: ManualRagItemRequest) -> dict[str, Any]:
    from admin_store import create_manual_rag_item

    _require_admin(request)
    try:
        return create_manual_rag_item(item.model_dump())
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/api/admin/manual-rag/{item_id}/update", include_in_schema=False)
def admin_update_manual_rag(request: Request, item_id: str, item: ManualRagItemRequest) -> dict[str, Any]:
    from admin_store import update_manual_rag_item

    _require_admin(request)
    try:
        return update_manual_rag_item(item_id, item.model_dump())
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/api/admin/manual-rag/{item_id}/delete", include_in_schema=False)
def admin_delete_manual_rag(request: Request, item_id: str) -> dict[str, Any]:
    from admin_store import delete_manual_rag_item

    _require_admin(request)
    try:
        return delete_manual_rag_item(item_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.get("/api/admin/campaign-intakes", include_in_schema=False)
def admin_campaign_intakes(request: Request) -> dict[str, Any]:
    from admin_store import list_campaign_intake_items

    _require_admin(request)
    return list_campaign_intake_items()


@app.post("/api/admin/campaign-intakes/update", include_in_schema=False)
def admin_update_campaign_intake(
    request: Request,
    update: CampaignIntakeOpsRequest,
) -> dict[str, Any]:
    from admin_store import update_campaign_intake_ops

    _require_admin(request)
    try:
        return update_campaign_intake_ops(update.model_dump())
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/chat", response_model=ChatResponse)
def chat(request: ChatRequest) -> ChatResponse:
    try:
        from rag_chatbot.qa import answer_question

        result = answer_question(request.question)
        try:
            from admin_store import record_chat_question

            record_chat_question(request.question, result.get("answer", ""), result.get("sources", []))
        except Exception:
            pass
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    return ChatResponse(**result)


@app.post("/check", response_model=list[CheckResultResponse])
async def check(request: CheckRequest) -> list[CheckResultResponse]:
    try:
        from checker import check_urls

        results = await check_urls(request.urls)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    return [CheckResultResponse(**result.to_dict()) for result in results]


@app.post("/check-favicon", response_model=list[FaviconCheckResultResponse])
async def check_favicon(
    request: FaviconCheckRequest,
) -> list[FaviconCheckResultResponse]:
    try:
        from favicon_checker import check_favicon_urls

        results = await check_favicon_urls(request.urls)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    return [FaviconCheckResultResponse(**result.to_dict()) for result in results]


@app.post("/intake", response_model=IntakeResponse)
async def intake(request: Request) -> IntakeResponse:
    try:
        from intake import IntakeSubmission, forward_intake_to_sheet

        payload = await request.json()
        submission = IntakeSubmission.model_validate(payload)
        client_key = request.client.host if request.client else "unknown"
        result = await forward_intake_to_sheet(submission, client_key=client_key)
    except ValidationError as exc:
        raise HTTPException(
            status_code=400,
            detail=_validation_error_details(exc),
        ) from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    return IntakeResponse(**result)


@app.post("/intake/workbook", include_in_schema=False)
async def intake_workbook(request: Request) -> StreamingResponse:
    try:
        from intake import IntakeSubmission, create_workbook_bytes

        payload = await request.json()
        submission = IntakeSubmission.model_validate(payload)
        workbook_bytes = create_workbook_bytes(submission)
    except ValidationError as exc:
        raise HTTPException(
            status_code=400,
            detail=_validation_error_details(exc),
        ) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc

    filename = "openai_ads_bulk_upload.xlsx"
    return StreamingResponse(
        BytesIO(workbook_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/intake/inspect-workbook", response_model=WorkbookInspectResponse, include_in_schema=False)
async def inspect_intake_workbook(file: UploadFile = File(...)) -> WorkbookInspectResponse:
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="xlsx 파일만 업로드할 수 있습니다.")
    try:
        from intake import inspect_workbook_bytes

        content = await file.read()
        if len(content) > 8 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="파일은 8MB 이하만 업로드할 수 있습니다.")
        result = inspect_workbook_bytes(content)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"워크북을 읽을 수 없습니다: {exc}") from exc
    return WorkbookInspectResponse(**result)


@app.post("/intake/upload-image", response_model=ImageUploadResponse, include_in_schema=False)
async def upload_intake_image(file: UploadFile = File(...)) -> ImageUploadResponse:
    supabase_url = os.getenv("SUPABASE_URL", "").strip().rstrip("/")
    service_role_key = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "").strip()
    bucket = os.getenv("SUPABASE_STORAGE_BUCKET", "openai-ad-assets").strip()
    if not supabase_url or not service_role_key:
        raise HTTPException(
            status_code=503,
            detail="이미지 업로드 저장소가 아직 설정되지 않았습니다. 공개 직접 이미지 URL을 입력해 주세요.",
        )

    content_type = file.content_type or ""
    if content_type not in {"image/png", "image/jpeg"}:
        raise HTTPException(status_code=400, detail="PNG 또는 JPG 이미지만 첨부할 수 있습니다.")
    data = await file.read()
    if len(data) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="이미지 파일은 5MB 이하만 첨부할 수 있습니다.")

    try:
        from PIL import Image

        image = Image.open(BytesIO(data))
        width, height = image.size
    except Exception as exc:
        raise HTTPException(status_code=400, detail="이미지 파일을 열 수 없습니다.") from exc
    if width != height:
        raise HTTPException(status_code=400, detail="이미지는 정사각형이어야 합니다.")
    if width < 640 or height < 640 or width > 1200 or height > 1200:
        raise HTTPException(status_code=400, detail="이미지는 640×640 이상, 1200×1200 이하이어야 합니다.")

    suffix = ".png" if content_type == "image/png" else ".jpg"
    safe_stem = re.sub(r"[^0-9A-Za-z._-]+", "-", Path(file.filename or "image").stem).strip("-")
    object_name = f"creative-upload/{uuid4().hex}-{safe_stem or 'image'}{suffix}"
    upload_url = f"{supabase_url}/storage/v1/object/{bucket}/{object_name}"
    headers = {
        "apikey": service_role_key,
        "Authorization": f"Bearer {service_role_key}",
        "Content-Type": content_type,
        "x-upsert": "false",
    }
    async with httpx.AsyncClient(timeout=30, follow_redirects=True) as client:
        response = await client.post(upload_url, content=data, headers=headers)
    if response.status_code >= 400:
        raise HTTPException(status_code=502, detail=f"Supabase Storage 업로드 실패(HTTP {response.status_code})")

    public_url = f"{supabase_url}/storage/v1/object/public/{bucket}/{object_name}"
    return ImageUploadResponse(
        image_url=public_url,
        width=width,
        height=height,
        content_type=content_type,
    )


@app.post("/admin/reindex", response_model=IngestResponse, include_in_schema=False)
async def admin_reindex(request: Request) -> IngestResponse:
    token = os.getenv("INGEST_TOKEN")
    provided = request.headers.get("x-ingest-token")
    if not token or provided != token:
        raise HTTPException(status_code=404, detail="Not found")

    try:
        from rag_chatbot.ingestion import ingest_collections

        counts = await asyncio.to_thread(ingest_collections)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    return IngestResponse(counts=counts)

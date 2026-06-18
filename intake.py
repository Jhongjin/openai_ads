from __future__ import annotations

from io import BytesIO
import json
from datetime import date, datetime
import os
import re
import threading
import time
from typing import Any, Literal
from urllib.parse import urlparse
from zoneinfo import ZoneInfo

import httpx
from pydantic import BaseModel, Field, field_validator, model_validator


KST = ZoneInfo("Asia/Seoul")
_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
_SAFE_NAME_RE = re.compile(r"^(?=.*[0-9A-Za-z])[0-9A-Za-z _-]+$")
_RECEIPT_LOCK = threading.Lock()
_RECEIPT_COUNTERS: dict[str, int] = {}
_RATE_LIMIT_SECONDS = 10
_LAST_SUBMIT_BY_IP: dict[str, float] = {}


def _clean_text(value: str) -> str:
    return str(value or "").strip()


def _require_text(value: str) -> str:
    text = _clean_text(value)
    if not text:
        raise ValueError("필수값이 비어 있습니다.")
    return text


def _valid_http_url(value: str, label: str = "URL") -> str:
    text = _require_text(value)
    if any(character.isspace() for character in text):
        raise ValueError(f"{label}은 공백 없이 입력해야 합니다.")
    parsed = urlparse(text)
    if parsed.scheme not in {"http", "https"}:
        raise ValueError(f"{label}은 http:// 또는 https://로 시작해야 합니다.")
    hostname = parsed.hostname or ""
    if (
        not parsed.netloc
        or not hostname
        or hostname.startswith(".")
        or hostname.endswith(".")
        or ".." in hostname
        or not re.fullmatch(r"[A-Za-z0-9.-]+", hostname)
    ):
        raise ValueError(f"{label}의 도메인 형식을 확인해 주세요.")
    return text


def _valid_safe_name(value: str, label: str) -> str:
    text = _require_text(value)
    if not _SAFE_NAME_RE.fullmatch(text):
        raise ValueError(
            f"{label}에는 영문, 숫자, 공백, 하이픈(-), 언더스코어(_)만 사용할 수 있습니다. "
            "점(.), 슬래시(/), 괄호, 이모지, 한글 등은 벌크 업로드 오류 방지를 위해 피해주세요."
        )
    return text


def _number_for_sheet(value: float | None) -> str:
    if value is None:
        return ""
    number = int(value) if float(value).is_integer() else value
    return str(number)


def _date_for_sheet(value: date | None) -> str:
    return value.isoformat() if value else ""


class OpsMeta(BaseModel):
    execution_route: Literal["openai_cbt", "criteo"] = Field(default="openai_cbt", alias="executionRoute")
    advertiser_name: str = Field(..., alias="advertiserName")
    ads_manager_account: str = Field(default="", alias="adsManagerAccount")

    sales_owner: str = Field(..., alias="salesOwner")
    sales_owner_email: str = Field(..., alias="salesOwnerEmail")
    owner_headquarters: str = Field(..., alias="ownerHeadquarters")
    owner_office: str = Field(..., alias="ownerOffice")
    owner_team: str = Field(..., alias="ownerTeam")
    notes: str = ""
    ads_manager_ready: bool = Field(default=False, alias="adsManagerReady")
    payment_ready: bool = Field(default=False, alias="paymentReady")
    crawler_ready: bool = Field(default=False, alias="crawlerReady")
    favicon_ready: bool = Field(default=False, alias="faviconReady")

    honeypot: str = ""
    form_started_at: int | None = Field(default=None, alias="formStartedAt")

    @field_validator(
        "advertiser_name",
        "sales_owner",
        "sales_owner_email",
        "owner_headquarters",
        "owner_office",
        "owner_team",
    )
    @classmethod
    def _required_text(cls, value: str) -> str:
        return _require_text(value)

    @field_validator(
        "notes",
        "honeypot",
        "ads_manager_account",
    )
    @classmethod
    def _optional_text(cls, value: str) -> str:
        return _clean_text(value)

    @field_validator("sales_owner_email")
    @classmethod
    def _valid_email(cls, value: str) -> str:
        text = _clean_text(value)
        if not text:
            return ""
        if not _EMAIL_RE.match(text):
            raise ValueError("이메일 형식이 올바르지 않습니다.")
        return text


class CampaignInfo(BaseModel):
    campaign_name: str
    budget_max: float
    budget_type: Literal["lifetime", "daily"] = "lifetime"
    launch_date: date | None = None
    end_date: date | None = None
    objective: Literal["views", "clicks"]
    target_countries: list[str] = Field(default_factory=list)

    @field_validator("campaign_name")
    @classmethod
    def _valid_campaign_name(cls, value: str) -> str:
        return _valid_safe_name(value, "캠페인명")

    @field_validator("budget_max")
    @classmethod
    def _valid_budget(cls, value: float) -> float:
        if value <= 0:
            raise ValueError("예산 금액은 0보다 커야 합니다.")
        return value

    @field_validator("target_countries")
    @classmethod
    def _valid_countries(cls, values: list[str]) -> list[str]:
        cleaned = [str(value or "").strip().upper() for value in values if str(value or "").strip()]
        invalid = [value for value in cleaned if not re.fullmatch(r"[A-Z]{2}", value)]
        if invalid:
            raise ValueError("국가는 2자리 ISO 코드로 저장해야 합니다.")
        return cleaned


class AdGroupInfo(BaseModel):
    campaign_name: str = ""
    adgroup_name: str
    max_bid: float | None = None
    keywords: list[str] = Field(default_factory=list)

    @field_validator("campaign_name")
    @classmethod
    def _clean_campaign_name(cls, value: str) -> str:
        return _clean_text(value)

    @field_validator("adgroup_name")
    @classmethod
    def _valid_adgroup_name(cls, value: str) -> str:
        return _valid_safe_name(value, "광고그룹명")

    @field_validator("max_bid", mode="before")
    @classmethod
    def _blank_max_bid(cls, value: Any) -> Any:
        if value == "":
            return None
        return value

    @field_validator("max_bid")
    @classmethod
    def _valid_max_bid(cls, value: float | None) -> float | None:
        if value is not None and value <= 0:
            raise ValueError("최대 입찰가는 0보다 커야 합니다.")
        return value

    @field_validator("keywords")
    @classmethod
    def _clean_keywords(cls, values: list[str]) -> list[str]:
        return [_clean_text(value) for value in values if _clean_text(value)]


class AdInfo(BaseModel):
    adgroup_name: str = ""
    title: str
    ad_copy: str = Field(..., alias="copy")
    link: str
    image_link: str

    @field_validator("adgroup_name")
    @classmethod
    def _clean_adgroup_name(cls, value: str) -> str:
        return _clean_text(value)

    @field_validator("title", "ad_copy")
    @classmethod
    def _required_copy(cls, value: str) -> str:
        return _require_text(value)

    @field_validator("link")
    @classmethod
    def _valid_link(cls, value: str) -> str:
        return _valid_http_url(value, "랜딩 URL")

    @field_validator("image_link")
    @classmethod
    def _valid_image_link(cls, value: str) -> str:
        return _valid_http_url(value, "image_link")


class AdGroupSubmission(BaseModel):
    adgroup: AdGroupInfo
    ads: list[AdInfo] = Field(..., min_length=1)


class CampaignSubmission(BaseModel):
    campaign: CampaignInfo
    adgroups: list[AdGroupSubmission] = Field(..., min_length=1)


class IntakeSubmission(BaseModel):
    ops_meta: OpsMeta = Field(..., alias="opsMeta")
    campaigns: list[CampaignSubmission] = Field(..., min_length=1)

    @model_validator(mode="before")
    @classmethod
    def _coerce_legacy_payload(cls, values: Any) -> Any:
        if not isinstance(values, dict) or values.get("campaigns"):
            return values
        campaign = values.get("campaign")
        adgroup = values.get("adgroup")
        ads = values.get("ads")
        if campaign and adgroup and ads:
            values = dict(values)
            values["campaigns"] = [{
                "campaign": campaign,
                "adgroups": [{
                    "adgroup": adgroup,
                    "ads": ads,
                }],
            }]
        return values

    @model_validator(mode="after")
    def _valid_submission(self) -> "IntakeSubmission":
        title_max = 30 if self.ops_meta.execution_route == "criteo" else 24
        copy_max = 60 if self.ops_meta.execution_route == "criteo" else 48
        campaign_names: set[str] = set()
        adgroup_names: set[str] = set()

        for campaign_index, campaign_item in enumerate(self.campaigns, start=1):
            campaign = campaign_item.campaign
            if campaign.campaign_name in campaign_names:
                raise ValueError(f"캠페인명 '{campaign.campaign_name}'이 중복되었습니다.")
            campaign_names.add(campaign.campaign_name)

            if campaign.end_date and campaign.launch_date:
                if campaign.end_date < campaign.launch_date:
                    raise ValueError(f"캠페인 {campaign_index}의 종료일은 시작일 이후여야 합니다.")
            if self.ops_meta.execution_route == "criteo" and campaign.objective != "views":
                raise ValueError("크리테오 경유는 CPM(Views) 목표만 선택할 수 있습니다.")

            for group_index, group_item in enumerate(campaign_item.adgroups, start=1):
                if campaign.objective == "views" and group_item.adgroup.max_bid is not None:
                    raise ValueError(
                        f"캠페인 {campaign_index} / 광고그룹 {group_index}: "
                        "Views(CPM) 캠페인은 max_bid를 비워야 합니다. max_bid는 Clicks(CPC) 캠페인에서만 입력하세요."
                    )
                adgroup_name = group_item.adgroup.adgroup_name
                if adgroup_name in adgroup_names:
                    raise ValueError(f"광고그룹명 '{adgroup_name}'이 중복되었습니다.")
                adgroup_names.add(adgroup_name)

                for ad_index, ad in enumerate(group_item.ads, start=1):
                    label = f"캠페인 {campaign_index} / 광고그룹 {group_index} / 소재 {ad_index}"
                    if len(ad.title) > title_max:
                        raise ValueError(f"{label}의 광고 제목은 최대 {title_max}자입니다.")
                    if len(ad.ad_copy) > copy_max:
                        raise ValueError(f"{label}의 광고 설명은 최대 {copy_max}자입니다.")
        return self


def _receipt_number(now: datetime) -> str:
    day = now.strftime("%Y%m%d")
    with _RECEIPT_LOCK:
        next_number = _RECEIPT_COUNTERS.get(day, 0) + 1
        _RECEIPT_COUNTERS[day] = next_number
    return f"KT-OAI-{day}-{next_number:03d}"


def _enforce_spam_controls(submission: IntakeSubmission, client_key: str) -> None:
    if submission.ops_meta.honeypot:
        raise ValueError("스팸 제출로 판단되어 접수하지 않았습니다.")
    if submission.ops_meta.form_started_at:
        elapsed = int(time.time() * 1000) - int(submission.ops_meta.form_started_at)
        if elapsed < 2500:
            raise ValueError("제출이 너무 빠릅니다. 내용을 확인한 뒤 다시 제출해 주세요.")

    now = time.monotonic()
    last_submit = _LAST_SUBMIT_BY_IP.get(client_key)
    if last_submit and now - last_submit < _RATE_LIMIT_SECONDS:
        raise ValueError("연속 제출 간격이 너무 짧습니다. 잠시 후 다시 시도해 주세요.")
    _LAST_SUBMIT_BY_IP[client_key] = now


def build_sheet_payload(
    submission: IntakeSubmission,
    *,
    shared_secret: str,
) -> dict[str, Any]:
    campaigns: list[dict[str, Any]] = []
    adgroups: list[dict[str, Any]] = []
    ads: list[dict[str, Any]] = []

    for campaign_item in submission.campaigns:
        campaign_name = campaign_item.campaign.campaign_name
        campaigns.append({
            "campaign_name": campaign_name,
            "budget_max": _number_for_sheet(campaign_item.campaign.budget_max),
            "budget_type": campaign_item.campaign.budget_type,
            "launch_date": _date_for_sheet(campaign_item.campaign.launch_date),
            "end_date": _date_for_sheet(campaign_item.campaign.end_date),
            "objective": campaign_item.campaign.objective,
            "target_countries": campaign_item.campaign.target_countries,
        })
        for group_item in campaign_item.adgroups:
            adgroup_name = group_item.adgroup.adgroup_name
            adgroups.append({
                "campaign_name": campaign_name,
                "adgroup_name": adgroup_name,
                "max_bid": _number_for_sheet(group_item.adgroup.max_bid),
                "keywords": group_item.adgroup.keywords,
            })
            ads.extend(
                {
                    "adgroup_name": adgroup_name,
                    "title": ad.title,
                    "copy": ad.ad_copy,
                    "link": ad.link,
                    "image_link": ad.image_link,
                }
                for ad in group_item.ads
            )

    primary_campaign = campaigns[0]
    ops = {
        "advertiser_name": submission.ops_meta.advertiser_name,
        "brand_name": submission.ops_meta.ads_manager_account,
        "sales_owner": submission.ops_meta.sales_owner,
        "sales_owner_email": submission.ops_meta.sales_owner_email,
        "owner_headquarters": submission.ops_meta.owner_headquarters,
        "owner_office": submission.ops_meta.owner_office,
        "owner_team": submission.ops_meta.owner_team,
        "note": submission.ops_meta.notes,
    }
    return {
        "secret": shared_secret,
        "data": {
            # Keep both shapes while the Apps Script web app is being redeployed.
            # Older script versions only read data.campaign; newer versions read data.campaigns.
            "campaign": campaigns,
            "primary_campaign": primary_campaign,
            "campaigns": campaigns,
            "adgroups": adgroups,
            "ads": ads,
            "ops": ops,
        },
    }


WORKBOOK_COLUMNS = {
    "campaigns": [
        "campaign_name",
        "budget_max",
        "budget_type",
        "launch_date",
        "end_date",
        "objective",
        "target_countries",
    ],
    "adgroups": ["campaign_name", "adgroup_name", "max_bid", "keywords"],
    "ads": ["adgroup_name", "title", "copy", "link", "image_link"],
}


def _sheet_value(value: Any) -> Any:
    if isinstance(value, list) and not value:
        return ""
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return "" if value is None else value


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    return str(value).strip()


def _is_template_guide_row(first_cell: str) -> bool:
    return (
        first_cell in {"Required"}
        or first_cell.startswith("How we will")
        or first_cell.startswith("The ")
    )


def _is_template_sample_value(value: Any) -> bool:
    return _cell_text(value).lower().startswith("oaitest")


def _is_http_url(value: str) -> bool:
    try:
        _valid_http_url(value)
    except ValueError:
        return False
    return True


def _json_array(value: Any, *, label: str, row_number: int, errors: list[str]) -> list[Any] | None:
    text = _cell_text(value)
    if not text:
        return None
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError:
        errors.append(f"{label} 행 {row_number}: JSON 배열 형식이 아닙니다.")
        return None
    if not isinstance(parsed, list):
        errors.append(f"{label} 행 {row_number}: JSON 배열이어야 합니다.")
        return None
    return parsed


def _json_array_quiet(value: Any) -> list[Any]:
    text = _cell_text(value)
    if not text:
        return []
    try:
        parsed = json.loads(text)
    except (TypeError, json.JSONDecodeError):
        return []
    return parsed if isinstance(parsed, list) else []


def _normalise_choice(value: Any) -> str:
    return _cell_text(value).lower()


def _normalise_objective(value: Any) -> str:
    # The template marks objective optional, but Ads Manager treats a blank
    # objective like the default Views flow for max_bid validation.
    return _normalise_choice(value) or "views"


def _to_positive_number(value: Any, *, label: str, row_number: int, errors: list[str]) -> float | None:
    text = _cell_text(value)
    if not text:
        return None
    try:
        number = float(text.replace(",", ""))
    except ValueError:
        errors.append(f"{label} 행 {row_number}: 숫자로 입력해야 합니다.")
        return None
    if number <= 0:
        errors.append(f"{label} 행 {row_number}: 0보다 큰 숫자를 입력해야 합니다.")
        return None
    return number


def create_workbook_bytes(submission: IntakeSubmission) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    sheet_payload = build_sheet_payload(submission, shared_secret="workbook")
    data = sheet_payload["data"]
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)

    for sheet_name in ("campaigns", "adgroups", "ads"):
        ws = wb.create_sheet(sheet_name)
        columns = WORKBOOK_COLUMNS[sheet_name]
        ws.append(columns)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in data[sheet_name]:
            ws.append([_sheet_value(row.get(column)) for column in columns])
        for column_cells in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 14), 44)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def inspect_workbook_bytes(content: bytes) -> dict[str, Any]:
    from openpyxl import load_workbook

    errors: list[str] = []
    warnings: list[str] = []
    summary: dict[str, Any] = {"sheets": {}, "errors": errors, "warnings": warnings}
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    parsed_rows: dict[str, list[dict[str, Any]]] = {sheet: [] for sheet in WORKBOOK_COLUMNS}
    for sheet_name, required_columns in WORKBOOK_COLUMNS.items():
        if sheet_name not in workbook.sheetnames:
            errors.append(f"{sheet_name} 시트가 없습니다.")
            continue
        ws = workbook[sheet_name]
        header = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        missing = [column for column in required_columns if column not in header]
        if missing:
            errors.append(f"{sheet_name} 시트에 필수 컬럼이 없습니다: {', '.join(missing)}")
        rows = 0
        sample_rows = 0
        guide_rows = 0
        for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            first = _cell_text(row[0])
            if not any(cell not in (None, "") for cell in row):
                continue
            if _is_template_guide_row(first):
                guide_rows += 1
                continue
            if _is_template_sample_value(first):
                sample_rows += 1
                continue
            rows += 1
            row_map = {
                column: row[header.index(column)] if column in header and header.index(column) < len(row) else None
                for column in required_columns
            }
            row_map["_row_number"] = row_number
            parsed_rows[sheet_name].append(row_map)
        if sample_rows and rows == 0:
            warnings.append(
                f"{sheet_name} 시트에 공식 템플릿 샘플(oaitest...) 행만 있습니다. "
                "샘플값을 실제 campaign/adgroup 이름으로 바꿔야 업로드 데이터로 인식됩니다."
            )
        if rows == 0:
            warnings.append(f"{sheet_name} 시트에서 실제 데이터 행을 찾지 못했습니다.")
        summary["sheets"][sheet_name] = {
            "columns": header,
            "rows": rows,
            "sample_rows": sample_rows,
            "guide_rows": guide_rows,
            "missing_columns": missing,
        }

    summary["data"] = {
        "campaigns": [
            {
                "campaign_name": _cell_text(row.get("campaign_name")),
                "budget_max": _cell_text(row.get("budget_max")),
                "budget_type": _normalise_choice(row.get("budget_type")) or "lifetime",
                "launch_date": _cell_text(row.get("launch_date")),
                "end_date": _cell_text(row.get("end_date")),
                "objective": _normalise_objective(row.get("objective")),
                "target_countries": _json_array_quiet(row.get("target_countries")),
            }
            for row in parsed_rows["campaigns"]
        ],
        "adgroups": [
            {
                "campaign_name": _cell_text(row.get("campaign_name")),
                "adgroup_name": _cell_text(row.get("adgroup_name")),
                "max_bid": _cell_text(row.get("max_bid")),
                "keywords": _json_array_quiet(row.get("keywords")),
            }
            for row in parsed_rows["adgroups"]
        ],
        "ads": [
            {
                "adgroup_name": _cell_text(row.get("adgroup_name")),
                "title": _cell_text(row.get("title")),
                "copy": _cell_text(row.get("copy")),
                "link": _cell_text(row.get("link")),
                "image_link": _cell_text(row.get("image_link")),
            }
            for row in parsed_rows["ads"]
        ],
    }

    campaign_objectives: dict[str, str] = {}
    for row in parsed_rows["campaigns"]:
        row_number = int(row["_row_number"])
        campaign_name = _cell_text(row.get("campaign_name"))
        budget_max = _cell_text(row.get("budget_max"))
        objective = _normalise_objective(row.get("objective"))
        budget_type = _normalise_choice(row.get("budget_type"))
        target_countries = _cell_text(row.get("target_countries"))

        if not campaign_name:
            errors.append(f"campaigns 행 {row_number}: campaign_name은 필수입니다.")
        elif _is_template_sample_value(campaign_name):
            errors.append(f"campaigns 행 {row_number}: 공식 샘플 campaign_name(oaitest...)을 실제 캠페인명으로 바꿔 주세요.")
        elif not _SAFE_NAME_RE.fullmatch(campaign_name):
            errors.append(
                f"campaigns 행 {row_number}: campaign_name은 영문, 숫자, 공백, 하이픈(-), 언더스코어(_)만 사용하세요. "
                "점(.), 슬래시(/), 괄호, 이모지, 한글 등은 벌크 업로드 오류 방지를 위해 피해주세요."
            )
        elif campaign_name in campaign_objectives:
            errors.append(f"campaigns 행 {row_number}: campaign_name '{campaign_name}'이 중복되었습니다.")
        else:
            campaign_objectives[campaign_name] = objective

        if not budget_max:
            errors.append(f"campaigns 행 {row_number}: budget_max는 필수입니다.")
        else:
            _to_positive_number(budget_max, label="campaigns budget_max", row_number=row_number, errors=errors)
        if budget_type and budget_type not in {"lifetime", "daily"}:
            errors.append(f"campaigns 행 {row_number}: budget_type은 lifetime 또는 daily여야 합니다.")
        if objective not in {"views", "clicks"}:
            errors.append(f"campaigns 행 {row_number}: objective는 views 또는 clicks여야 합니다.")
        if target_countries:
            countries = _json_array(target_countries, label="campaigns target_countries", row_number=row_number, errors=errors)
            if countries is not None:
                invalid = [str(country) for country in countries if not re.fullmatch(r"[A-Z]{2}", str(country).strip().upper())]
                if invalid:
                    errors.append(f"campaigns 행 {row_number}: target_countries는 2자리 국가 코드 배열이어야 합니다.")
                normalized = {str(country).strip().upper() for country in countries}
                if "KR" in normalized:
                    errors.append(
                        f"campaigns 행 {row_number}: 현재 OpenAI Ads Manager 벌크 업로드에서 "
                        "target_countries에 KR을 넣으면 오류가 확인되었습니다. 한국 집행은 이 칸을 빈칸(NULL)으로 두세요."
                    )
        else:
            warnings.append(
                f"campaigns 행 {row_number}: target_countries는 빈칸입니다. "
                "현재 한국 집행 벌크 업로드는 빈칸(NULL)으로 두면 Ads Manager에서 ALL_COUNTRIES로 표시되며 업로드됩니다."
            )

    adgroup_campaigns: dict[str, str] = {}
    for row in parsed_rows["adgroups"]:
        row_number = int(row["_row_number"])
        campaign_name = _cell_text(row.get("campaign_name"))
        adgroup_name = _cell_text(row.get("adgroup_name"))
        max_bid = _cell_text(row.get("max_bid"))
        keywords = _cell_text(row.get("keywords"))
        campaign_objective = campaign_objectives.get(campaign_name)

        if not campaign_name:
            errors.append(f"adgroups 행 {row_number}: campaign_name은 필수입니다.")
        elif campaign_name not in campaign_objectives:
            errors.append(f"adgroups 행 {row_number}: 존재하지 않는 campaign_name '{campaign_name}'을 참조합니다.")
        if not adgroup_name:
            errors.append(f"adgroups 행 {row_number}: adgroup_name은 필수입니다.")
        elif _is_template_sample_value(adgroup_name):
            errors.append(f"adgroups 행 {row_number}: 공식 샘플 adgroup_name(oaitest...)을 실제 광고그룹명으로 바꿔 주세요.")
        elif not _SAFE_NAME_RE.fullmatch(adgroup_name):
            errors.append(
                f"adgroups 행 {row_number}: adgroup_name은 영문, 숫자, 공백, 하이픈(-), 언더스코어(_)만 사용하세요. "
                "점(.), 슬래시(/), 괄호, 이모지, 한글 등은 벌크 업로드 오류 방지를 위해 피해주세요."
            )
        elif adgroup_name in adgroup_campaigns:
            errors.append(f"adgroups 행 {row_number}: adgroup_name '{adgroup_name}'이 중복되었습니다.")
        else:
            adgroup_campaigns[adgroup_name] = campaign_name

        if max_bid:
            if campaign_objective == "views":
                errors.append(f"adgroups 행 {row_number}: Views(CPM) 캠페인의 광고그룹은 max_bid를 비워야 합니다. max_bid는 Clicks(CPC) 캠페인에서만 입력하세요.")
            elif campaign_objective == "clicks":
                _to_positive_number(max_bid, label="adgroups max_bid", row_number=row_number, errors=errors)
            else:
                warnings.append(f"adgroups 행 {row_number}: campaign objective를 확인할 수 없어 max_bid 적용 여부를 판단할 수 없습니다.")
        if keywords:
            _json_array(keywords, label="adgroups keywords", row_number=row_number, errors=errors)

    for row in parsed_rows["ads"]:
        row_number = int(row["_row_number"])
        adgroup_name = _cell_text(row.get("adgroup_name"))
        title = _cell_text(row.get("title"))
        copy = _cell_text(row.get("copy"))
        link = _cell_text(row.get("link"))
        image_link = _cell_text(row.get("image_link"))

        if not adgroup_name:
            errors.append(f"ads 행 {row_number}: adgroup_name은 필수입니다.")
        elif adgroup_name not in adgroup_campaigns:
            errors.append(f"ads 행 {row_number}: 존재하지 않는 adgroup_name '{adgroup_name}'을 참조합니다.")
        if not title:
            errors.append(f"ads 행 {row_number}: title은 필수입니다.")
        elif len(title) > 24:
            errors.append(f"ads 행 {row_number}: title은 최대 24자입니다.")
        if not copy:
            errors.append(f"ads 행 {row_number}: copy는 필수입니다.")
        elif len(copy) > 48:
            errors.append(f"ads 행 {row_number}: copy는 최대 48자입니다.")
        if not link:
            errors.append(f"ads 행 {row_number}: link는 필수입니다.")
        elif not _is_http_url(link):
            errors.append(f"ads 행 {row_number}: link는 http:// 또는 https://로 시작해야 합니다.")
        if not image_link:
            errors.append(f"ads 행 {row_number}: image_link는 필수입니다.")
        elif not _is_http_url(image_link):
            errors.append(f"ads 행 {row_number}: image_link는 http:// 또는 https://로 시작하는 정상 URL 형식이어야 합니다.")

    summary["ok"] = not errors and all(
        item.get("rows", 0) > 0 for item in summary["sheets"].values()
    )
    return summary


async def forward_intake_to_sheet(
    submission: IntakeSubmission,
    *,
    client_key: str,
) -> dict[str, Any]:
    _enforce_spam_controls(submission, client_key)

    webhook_url = os.getenv("GOOGLE_SHEETS_WEBHOOK_URL", "").strip()
    shared_secret = os.getenv("SHEETS_SHARED_SECRET", "").strip()
    if not webhook_url or not shared_secret:
        raise RuntimeError(
            "GOOGLE_SHEETS_WEBHOOK_URL 또는 SHEETS_SHARED_SECRET 환경변수가 설정되지 않았습니다."
        )

    now = datetime.now(KST)
    fallback_receipt_number = _receipt_number(now)
    fallback_submitted_at_kst = now.strftime("%Y-%m-%d %H:%M:%S")
    payload = build_sheet_payload(
        submission,
        shared_secret=shared_secret,
    )

    async with httpx.AsyncClient(timeout=15, follow_redirects=True) as client:
        response = await client.post(webhook_url, json=payload)
        if response.status_code >= 400:
            raise RuntimeError(f"Google Sheets 기록 실패(HTTP {response.status_code})")
        try:
            response_payload = response.json()
        except ValueError:
            response_payload = {}
        if isinstance(response_payload, dict) and response_payload.get("ok") is False:
            error = response_payload.get("error") or "Apps Script가 접수를 거부했습니다."
            raise RuntimeError(f"Google Sheets 기록 실패({error})")

    return {
        "receipt_number": str(
            response_payload.get("receiptNumber")
            or response_payload.get("receipt_number")
            or fallback_receipt_number
        ),
        "submitted_at_kst": str(
            response_payload.get("submittedAtKst")
            or response_payload.get("submitted_at_kst")
            or fallback_submitted_at_kst
        ),
        "mail_sent": response_payload.get("mailSent")
        if isinstance(response_payload.get("mailSent"), bool)
        else None,
        "mail_error": str(response_payload.get("mailError") or ""),
    }
